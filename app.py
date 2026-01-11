import os
import secrets
from datetime import datetime
import pytz   # <--- add this (install via `pip install pytz`)
import psycopg2
import psycopg2.extras

from flask import Flask, request, jsonify, session, send_from_directory, make_response
#from flask_login import login_required
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from cryptography.fernet import Fernet
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

# --------------------------------------------------------------------------------------
# Flask app setup
# --------------------------------------------------------------------------------------

app = Flask(__name__, static_folder='static', static_url_path='')
app.secret_key = os.environ.get('SESSION_SECRET', secrets.token_hex(32))

# --------------------------------------------------------------------------------------
# PostgreSQL helpers
# --------------------------------------------------------------------------------------
from functools import wraps
from flask import session, jsonify

def login_required_local(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'username' not in session:
            return jsonify({"error": "Unauthorized"}), 401
        return f(*args, **kwargs)
    return decorated

def get_db():
    """
    Get a new PostgreSQL connection.
    Configure via environment variables:

      PGHOST, PGPORT, PGUSER, PGPASSWORD, PGDATABASE

    For example:
      export PGHOST=localhost
      export PGPORT=5432
      export PGUSER=gtn
      export PGPASSWORD=secret
      export PGDATABASE=contract_review
    """
    conn = psycopg2.connect(
        host=os.environ.get("PGHOST", "localhost"),
        port=int(os.environ.get("PGPORT", 5432)),
        user=os.environ.get("PGUSER", "ContractDMS"),
        password=os.environ.get("PGPASSWORD", "gtn@123"),
        dbname=os.environ.get("PGDATABASE", "CR"),
    )
    conn.autocommit = False
    return conn

IST = pytz.timezone("Asia/Kolkata")

def now_ist():
    """Return current time in India (IST) as aware datetime."""
    return datetime.now(IST)

# --------------------------------------------------------------------------------------
# Encryption key storage (encryption_keys table)
# --------------------------------------------------------------------------------------

def get_encryption_key():
    """Get or generate stable encryption key stored in DB."""
    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS encryption_keys (
                    id   INTEGER PRIMARY KEY,
                    key  TEXT NOT NULL
                )
            """)
            conn.commit()

            cur.execute("SELECT key FROM encryption_keys WHERE id = 1")
            row = cur.fetchone()
            if row:
                key = row[0].encode()
            else:
                key = Fernet.generate_key()
                cur.execute(
                    "INSERT INTO encryption_keys (id, key) VALUES (%s, %s)",
                    (1, key.decode())
                )
                conn.commit()
        conn.close()
        return key
    except Exception as e:
        print(f"Error getting encryption key: {e}")
        conn.close()
        return Fernet.generate_key()

def encrypt_password(password):
    """Encrypt SMTP password"""
    f = Fernet(get_encryption_key())
    return f.encrypt(password.encode()).decode()

def decrypt_password(encrypted_password):
    """Decrypt SMTP password"""
    f = Fernet(get_encryption_key())
    return f.decrypt(encrypted_password.encode()).decode()

# --------------------------------------------------------------------------------------
# Email sending
# --------------------------------------------------------------------------------------

def send_notification_email(recipient_email, subject, message_body):
    """Send email notification using stored SMTP configuration"""
    if not recipient_email:
        return False

    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart

    conn = get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                "SELECT * FROM email_config ORDER BY id DESC LIMIT 1"
            )
            config = cur.fetchone()
        conn.close()

        if not config or not config["email_enabled"]:
            return False

        decrypted_password = decrypt_password(config["smtp_password"])

        msg = MIMEMultipart()
        msg["From"] = config["from_email"]
        msg["To"] = recipient_email
        msg["Subject"] = subject

        html_body = f'''
        <html>
        <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
            <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
                <div style="background: #0b5ed7; color: white; padding: 20px; border-radius: 8px 8px 0 0;">
                    <h2 style="margin: 0;">GTN ENGINEERING (INDIA) LIMITED</h2>
                    <p style="margin: 5px 0 0 0; font-size: 14px;">Contract Review Dashboard</p>
                </div>
                <div style="background: #f8f9fa; padding: 30px; border-radius: 0 0 8px 8px;">
                    {message_body}
                </div>
                <div style="margin-top: 20px; padding: 15px; background: #fff; border: 1px solid #e5e7eb; border-radius: 6px; font-size: 12px; color: #6b7280;">
                    <p style="margin: 0;"><strong>Note:</strong> This is an automated notification from the Contract Review Dashboard. Please do not reply to this email.</p>
                </div>
            </div>
        </body>
        </html>
        '''

        msg.attach(MIMEText(html_body, "html"))

        server = smtplib.SMTP(config["smtp_host"], config["smtp_port"])
        if config["use_tls"]:
            server.starttls()
        server.login(config["smtp_username"], decrypted_password)
        server.send_message(msg)
        server.quit()

        return True
    except Exception as e:
        print(f"Failed to send email: {e}")
        return False

#------------------------------------------------------------
#
#-----------------------------------------------------------
# ---------- department-check endpoint (Option A, strict sequential) ----------
import json
from flask import jsonify

def _is_nonblank(val):
    return bool(str(val).strip())

def _compute_cr_completed_depts(conn, cr_form_id):
    """
    Return set of completed department keys for a CR form.
    Option A: department completed only when every cell in that department's
    columns for every row is non-blank.
    """
    groups = [
        ('engineering', 13),
        ('manufacturing', 8),
        ('materials', 8),
        ('purchase', 5),
        ('special-process', 3),
        ('welding', 3),
        ('assembly', 4),
        ('quality', 10),
        ('painting', 5),
        ('customer-service', 4),
        ('commercial', 1)
    ]

    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute("SELECT cycles FROM cr_form_rows WHERE cr_form_id = %s ORDER BY id", (cr_form_id,))
        rows = cur.fetchall()

    if not rows:
        return set()

    cycles_rows = []
    for r in rows:
        try:
            c = json.loads(r['cycles']) if r['cycles'] else []
        except Exception:
            c = []
        cycles_rows.append(c)

    completed = set()
    start = 0
    for key, count in groups:
        end = start + count - 1
        all_ok = True
        for crow in cycles_rows:
            for idx in range(start, end + 1):
                val = ''
                if idx < len(crow):
                    val = crow[idx]
                if not _is_nonblank(val):
                    all_ok = False
                    break
            if not all_ok:
                break
        if all_ok:
            completed.add(key)
        start += count
    return completed

def _compute_ped_completed_depts(conn, ped_form_id):
    """
    Return set of completed department keys for a PED form.
    For strict Option A this requires the pedCycles columns (group ranges)
    to be fully non-blank for every row.
    """
    groups = [
        ('engineering', 7),
        ('manufacturing', 1),
        ('materials', 1),
        ('purchase', 1)
    ]

    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute("SELECT ped_cycles FROM ped_form_rows WHERE ped_form_id = %s ORDER BY id", (ped_form_id,))
        rows = cur.fetchall()

    if not rows:
        return set()

    ped_rows = []
    for r in rows:
        try:
            p = json.loads(r['ped_cycles']) if r['ped_cycles'] else []
        except Exception:
            p = []
        ped_rows.append(p)

    completed = set()
    start = 0
    for key, count in groups:
        end = start + count - 1
        all_ok = True
        for crow in ped_rows:
            for idx in range(start, end + 1):
                val = ''
                if idx < len(crow):
                    val = crow[idx]
                if not _is_nonblank(val):
                    all_ok = False
                    break
            if not all_ok:
                break
        if all_ok:
            completed.add(key)
        start += count
    return completed

def _compute_lead_completed_depts(conn, po_key):
    """
    LEAD completion: a department is completed if it has signed for this po_key
    in lead_department_signatures.
    """
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute("""
            SELECT department
            FROM lead_department_signatures
            WHERE po_key = %s
        """, (po_key,))
        rows = cur.fetchall()

    completed = set()
    for r in rows:
        d = (r.get("department") or "").strip()
        if d:
            completed.add(d)
    return completed

@app.route("/api/forms/check-departments", methods=["GET"])
@login_required_local
def api_check_departments():
    """
    Return:
      { "allowedDepartments": [...], "completedDepartments": [...] }

    Behavior (Option A strict):
      - Find first dept in order that is NOT fully completed -> return that dept as allowed.
      - If no form exists yet for poKey, allow 'engineering' (initial).
      - If all depts complete, allowedDepartments = [].
    """
    form_type = (request.args.get("formType") or "").strip().upper()
    po_key = (request.args.get("poKey") or "").strip()
    if not form_type or not po_key:
        return jsonify({"error": "formType and poKey required"}), 400

    conn = get_db()
    try:
        if form_type == "CR":
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("SELECT id FROM cr_forms WHERE po_key = %s", (po_key,))
                f = cur.fetchone()
            if not f:
                return jsonify({"allowedDepartments": ["engineering"], "completedDepartments": []})
            completed = _compute_cr_completed_depts(conn, f['id'])
            for d in CR_DEPT_ORDER:
                if d not in completed:
                    return jsonify({"allowedDepartments": [d], "completedDepartments": list(completed)})
            return jsonify({"allowedDepartments": [], "completedDepartments": list(completed)})

        elif form_type == "PED":
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("SELECT id FROM ped_forms WHERE po_key = %s", (po_key,))
                f = cur.fetchone()
            if not f:
                return jsonify({"allowedDepartments": ["engineering"], "completedDepartments": []})
            completed = _compute_ped_completed_depts(conn, f['id'])
            for d in PED_DEPT_ORDER:
                if d not in completed:
                    return jsonify({"allowedDepartments": [d], "completedDepartments": list(completed)})
            return jsonify({"allowedDepartments": [], "completedDepartments": list(completed)})

        elif form_type == "LEAD":
            # If no lead form exists yet, allow first dept (css)
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("SELECT id FROM lead_forms WHERE po_key = %s", (po_key,))
                f = cur.fetchone()

            if not f:
                return jsonify({"allowedDepartments": ["css"], "completedDepartments": []})

            completed = _compute_lead_completed_depts(conn, po_key)

            for d in LEAD_DEPT_ORDER:
                if d not in completed:
                    return jsonify({"allowedDepartments": [d], "completedDepartments": list(completed)})

            return jsonify({"allowedDepartments": [], "completedDepartments": list(completed)})

        else:
            return jsonify({"error": "Unsupported formType"}), 400


    except Exception as e:
        conn.close()
        print("check-departments error:", e)
        return jsonify({"error": str(e)}), 500





# --------------------------------------------------------------------------------------
# DB initialization (schema in PostgreSQL)
# --------------------------------------------------------------------------------------

def init_db():
    conn = get_db()
    try:
        with conn.cursor() as cur:
            # users
            cur.execute("""
                CREATE TABLE IF NOT EXISTS users (
                    id               SERIAL PRIMARY KEY,
                    username         TEXT UNIQUE NOT NULL,
                    password_hash    TEXT NOT NULL,
                    name             TEXT NOT NULL,
                    department       TEXT NOT NULL,
                    is_admin         BOOLEAN NOT NULL DEFAULT FALSE,
                    lead_form_access BOOLEAN NOT NULL DEFAULT FALSE,
                    email            TEXT,
                    created_at       TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP
                )
            """)

            # ---- MIGRATIONS / SAFE ALTERs (IMPORTANT) ----
            # If the table was created earlier, CREATE TABLE IF NOT EXISTS will NOT add new columns.
            # These ALTERs make sure newer columns exist.
            cur.execute("""
                ALTER TABLE users
                ADD COLUMN IF NOT EXISTS email TEXT
            """)
            cur.execute("""
                ALTER TABLE users
                ADD COLUMN IF NOT EXISTS lead_form_access BOOLEAN NOT NULL DEFAULT FALSE
            """)
            # (optional) if you ever added is_admin later in older DBs:
            cur.execute("""
                ALTER TABLE users
                ADD COLUMN IF NOT EXISTS is_admin BOOLEAN NOT NULL DEFAULT FALSE
            """)

            # pos
            cur.execute("""
                CREATE TABLE IF NOT EXISTS pos (
                    id         SERIAL PRIMARY KEY,
                    customer   TEXT NOT NULL,
                    bid        TEXT NOT NULL,
                    po         TEXT NOT NULL,
                    cr         TEXT NOT NULL,
                    created_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP
                )
            """)
            # CR Department Signatures
            cur.execute("""
            CREATE TABLE IF NOT EXISTS cr_department_signatures (
                id SERIAL PRIMARY KEY,
                po_key TEXT NOT NULL,
                department TEXT NOT NULL,
                signed_by TEXT NOT NULL,
                signed_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP,
                UNIQUE (po_key, department)
            )
            """)

            # cr_forms
            cur.execute("""
                CREATE TABLE IF NOT EXISTS cr_forms (
                    id               SERIAL PRIMARY KEY,
                    po_key           TEXT UNIQUE NOT NULL,
                    customer         TEXT,
                    bid              TEXT,
                    po               TEXT,
                    cr               TEXT,
                    record_no        TEXT,
                    record_date      TEXT,
                    last_modified_by TEXT,
                    last_modified_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP,
                    amendment_details TEXT
                )
            """)

            # cr_form_rows
            cur.execute("""
                CREATE TABLE IF NOT EXISTS cr_form_rows (
                    id               SERIAL PRIMARY KEY,
                    cr_form_id       INTEGER NOT NULL REFERENCES cr_forms(id) ON DELETE CASCADE,
                    item_no          TEXT NOT NULL,
                    part_number      TEXT,
                    part_description TEXT,
                    rev              TEXT,
                    qty              TEXT,
                    cycles           TEXT,
                    remarks          TEXT
                )
            """)

            #PED Signature
            cur.execute ("""
                CREATE TABLE IF NOT EXISTS ped_department_signatures (
                    id SERIAL PRIMARY KEY,
                    po_key TEXT NOT NULL,
                    department TEXT NOT NULL,
                    signed_by TEXT NOT NULL,
                    signed_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE (po_key, department)
                )
            """)

            # ped_forms
            cur.execute("""
                CREATE TABLE IF NOT EXISTS ped_forms (
                    id               SERIAL PRIMARY KEY,
                    po_key           TEXT UNIQUE NOT NULL,
                    customer         TEXT,
                    bid              TEXT,
                    po               TEXT,
                    cr               TEXT,
                    record_no        TEXT,
                    record_date      TEXT,
                    amendment_details TEXT,
                    last_modified_by TEXT,
                    last_modified_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP
                )
            """)

            # ped_form_rows
            cur.execute("""
                CREATE TABLE IF NOT EXISTS ped_form_rows (
                    id               SERIAL PRIMARY KEY,
                    ped_form_id      INTEGER NOT NULL REFERENCES ped_forms(id) ON DELETE CASCADE,
                    item_no          TEXT NOT NULL,
                    part_number      TEXT,
                    part_description TEXT,
                    rev              TEXT,
                    qty              TEXT,
                    ped_cycles       TEXT,
                    notes            TEXT,
                    remarks          TEXT
                )
            """)

            # lead_forms
            cur.execute("""
                CREATE TABLE IF NOT EXISTS lead_forms (
                    id               SERIAL PRIMARY KEY,
                    po_key           TEXT UNIQUE NOT NULL,
                    customer         TEXT,
                    bid              TEXT,
                    po               TEXT,
                    cr               TEXT,
                    record_no        TEXT,
                    record_date      TEXT,
                    prepared_by      TEXT,     -- NEW
                    general_remarks  TEXT,
                    last_modified_by TEXT,
                    last_modified_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP
                )
            """)

            # lead_form_rows
            cur.execute("""
                CREATE TABLE IF NOT EXISTS lead_form_rows (
                    id                    SERIAL PRIMARY KEY,
                    lead_form_id          INTEGER NOT NULL REFERENCES lead_forms(id) ON DELETE CASCADE,
                    item_no               TEXT NOT NULL,
                    part_number           TEXT,
                    part_description      TEXT,
                    rev                   TEXT,
                    qty                   TEXT,
                    customer_required_date TEXT,
                    standard_lead_time    TEXT,
                    gtn_agreed_date       TEXT,
                    remarks               TEXT
                )
            """)

            #LEAD Sign

            cur.execute("""
                        CREATE TABLE IF NOT EXISTS lead_department_signatures (
                            id SERIAL PRIMARY KEY,
                            po_key TEXT NOT NULL,
                            department TEXT NOT NULL,
                            signed_by TEXT NOT NULL,
                            signed_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP,
                            UNIQUE (po_key, department)
                        )
                        """)

            # comments tables
            cur.execute("""
                CREATE TABLE IF NOT EXISTS cr_comments (
                    id          SERIAL PRIMARY KEY,
                    cr_form_id  INTEGER NOT NULL REFERENCES cr_forms(id) ON DELETE CASCADE,
                    username    TEXT NOT NULL,
                    department  TEXT NOT NULL,
                    comment_text TEXT NOT NULL,
                    created_at  TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP
                )
            """)

            cur.execute("""
                CREATE TABLE IF NOT EXISTS ped_comments (
                    id          SERIAL PRIMARY KEY,
                    ped_form_id INTEGER NOT NULL REFERENCES ped_forms(id) ON DELETE CASCADE,
                    username    TEXT NOT NULL,
                    department  TEXT NOT NULL,
                    comment_text TEXT NOT NULL,
                    created_at  TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP
                )
            """)

            cur.execute("""
                CREATE TABLE IF NOT EXISTS lead_comments (
                    id          SERIAL PRIMARY KEY,
                    lead_form_id INTEGER NOT NULL REFERENCES lead_forms(id) ON DELETE CASCADE,
                    username    TEXT NOT NULL,
                    department  TEXT NOT NULL,
                    comment_text TEXT NOT NULL,
                    created_at  TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP
                )
            """)

            # notifications
            cur.execute("""
                CREATE TABLE IF NOT EXISTS notifications (
                    id                SERIAL PRIMARY KEY,
                    recipient_user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    actor_user_id     INTEGER REFERENCES users(id) ON DELETE SET NULL,
                    event_type        TEXT NOT NULL,
                    po_id             INTEGER,
                    form_type         TEXT,
                    form_id           INTEGER,
                    message           TEXT NOT NULL,
                    metadata          TEXT,
                    is_read           BOOLEAN DEFAULT FALSE,
                    email_sent        BOOLEAN DEFAULT FALSE,
                    created_at        TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP
                )
            """)

            # form_completion_states
            cur.execute("""
                CREATE TABLE IF NOT EXISTS form_completion_states (
                    id                  SERIAL PRIMARY KEY,
                    form_type           TEXT NOT NULL,
                    form_id             INTEGER NOT NULL,
                    status              TEXT DEFAULT 'in_progress',
                    completion_snapshot TEXT,
                    completed_by        INTEGER REFERENCES users(id) ON DELETE SET NULL,
                    completed_at        TIMESTAMPTZ,
                    created_at          TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(form_type, form_id)
                )
            """)

            # per-department completion states
            cur.execute("""
                CREATE TABLE IF NOT EXISTS form_department_states (
                    id           SERIAL PRIMARY KEY,
                    form_type    TEXT NOT NULL,
                    form_id      INTEGER NOT NULL,
                    department   TEXT NOT NULL,
                    status       TEXT DEFAULT 'in_progress',
                    completed_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                    completed_at TIMESTAMPTZ,
                    UNIQUE(form_type, form_id, department)
                )
            """)

            # master_signatures
            cur.execute("""
                CREATE TABLE IF NOT EXISTS master_signatures (
                    id                 SERIAL PRIMARY KEY,
                    user_id            INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    username           TEXT NOT NULL,
                    department         TEXT NOT NULL,

                    signature_file_name TEXT NOT NULL,
                    signature_path      TEXT NOT NULL,
                    signature_type      TEXT NOT NULL,
                    signature_size_kb   INTEGER NOT NULL,
                    checksum_sha256     TEXT NOT NULL,

                    is_active          BOOLEAN DEFAULT TRUE,
                    remarks            TEXT,

                    uploaded_by        TEXT NOT NULL,
                    uploaded_at        TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP,
                    updated_at         TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP,

                    UNIQUE (user_id, department)
                )
            """)

            # email_config
            cur.execute("""
                CREATE TABLE IF NOT EXISTS email_config (
                    id           SERIAL PRIMARY KEY,
                    smtp_host    TEXT NOT NULL,
                    smtp_port    INTEGER NOT NULL,
                    smtp_username TEXT NOT NULL,
                    smtp_password TEXT NOT NULL,
                    from_email   TEXT NOT NULL,
                    use_tls      BOOLEAN DEFAULT TRUE,
                    email_enabled BOOLEAN DEFAULT FALSE,
                    updated_by   TEXT,
                    updated_at   TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP
                )
            """)

            # ensure default admin user
            cur.execute(
                "SELECT COUNT(*) FROM users WHERE username = %s",
                ("admin",)
            )
            count = cur.fetchone()[0]
            if count == 0:
                cur.execute(
                    """
                    INSERT INTO users (username, password_hash, name, department, is_admin)
                    VALUES (%s, %s, %s, %s, %s)
                    """,
                    ("admin", generate_password_hash("admin"),
                     "IT Administrator", "it", True)
                )

        conn.commit()
    except Exception as e:
        conn.rollback()
        print("init_db error:", e)
    finally:
        conn.close()

# ---------- Department order definitions ----------

CR_DEPT_ORDER = [
    "engineering", "manufacturing", "materials", "purchase",
    "special-process", "welding", "assembly", "quality",
    "painting", "customer-service", "commercial"
]

PED_DEPT_ORDER = [
    "engineering", "manufacturing", "materials", "purchase"
]

LEAD_DEPT_ORDER = ["css", "materials", "technical-operations", "quality", "operations"]

NEXT_FORM_AFTER = {
    "CR": "PED",
    "PED": "LEAD",
    "LEAD": None
}

# --------------------------------------------------------------------------------------
# Auth decorators
# --------------------------------------------------------------------------------------

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user_id" not in session:
            return jsonify({"error": "Unauthorized"}), 401
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user_id" not in session:
            return jsonify({"error": "Unauthorized"}), 401
        conn = get_db()
        try:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT is_admin FROM users WHERE id = %s",
                    (session["user_id"],)
                )
                row = cur.fetchone()
            conn.close()
            if not row or not row[0]:
                return jsonify({"error": "Admin access required"}), 403
            return f(*args, **kwargs)
        except Exception:
            conn.close()
            return jsonify({"error": "Admin access required"}), 403
    return decorated_function


# --------------------------------------------------------------------------------------
# Notifications helpers
# --------------------------------------------------------------------------------------

def create_notification_for_all_users(
    event_type,
    message,
    actor_user_id=None,
    po_id=None,
    form_type=None,
    form_id=None,
    exclude_user_id=None
):
    """Create a notification for all users (optionally excluding one user)"""
    conn = get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT id, email, name FROM users")
            users = cur.fetchall()

            for user in users:
                if exclude_user_id and user["id"] == exclude_user_id:
                    continue
                cur.execute(
                    """
                    INSERT INTO notifications
                      (recipient_user_id, actor_user_id, event_type, po_id, form_type, form_id, message)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    """,
                    (user["id"], actor_user_id, event_type, po_id, form_type, form_id, message)
                )

        conn.commit()

        # send emails (separate cursor/connection)
        conn2 = get_db()
        try:
            with conn2.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur2:
                for user in users:
                    if exclude_user_id and user["id"] == exclude_user_id:
                        continue
                    if not user["email"]:
                        continue

                    if event_type == "po_created":
                        subject = "New PO Created - Contract Review Dashboard"
                        email_body = f'''
                        <h3 style="color:#0b5ed7;margin-top:0;">New Purchase Order Created</h3>
                        <p style="font-size:16px;color:#111;">{message}</p>
                        <p style="margin-top:20px;">Please log in to the Contract Review Dashboard to view the details and start working on the associated forms.</p>
                        <div style="margin-top:30px;">
                          <a href="http://192.160.207.224:1000" style="background:#0b5ed7;color:white;padding:12px 24px;text-decoration:none;border-radius:6px;display:inline-block;">View Dashboard</a>
                        </div>
                        '''
                    elif event_type == "form_completed":
                        subject = f"{form_type} Form Completed - Contract Review Dashboard"
                        email_body = f'''
                        <h3 style="color:#22c55e;margin-top:0;">Form Completed</h3>
                        <p style="font-size:16px;color:#111;">{message}</p>
                        <p style="margin-top:20px;">The {form_type} form has been completed and is now ready for review.</p>
                        <div style="margin-top:30px;">
                          <a href="http://192.160.207.224:1000" style="background:#22c55e;color:white;padding:12px 24px;text-decoration:none;border-radius:6px;display:inline-block;">View Form</a>
                        </div>
                        '''
                    else:
                        subject = "Notification - Contract Review Dashboard"
                        email_body = f'''
                        <h3 style="color:#0b5ed7;margin-top:0;">New Notification</h3>
                        <p style="font-size:16px;color:#111;">{message}</p>
                        '''

                    sent = send_notification_email(user["email"], subject, email_body)
                    if sent:
                        cur2.execute(
                            """
                            UPDATE notifications
                               SET email_sent = TRUE
                             WHERE recipient_user_id = %s
                               AND form_id = %s
                               AND (form_type = %s OR %s IS NULL)
                            """,
                            (user["id"], form_id, form_type, form_type)
                        )
            conn2.commit()
        finally:
            conn2.close()

        return True
    except Exception as e:
        conn.rollback()
        conn.close()
        print(f"Error creating notifications: {e}")
        return False

# --------------------------------------------------------------------------------------
# Form completion checks
# --------------------------------------------------------------------------------------

def check_form_completion(form_type, form_data):
    """Check if a form meets completion criteria."""
    import json

    if form_type == "CR":
        if not all([form_data.get("customer"), form_data.get("bid"),
                    form_data.get("po"), form_data.get("cr")]):
            return False
        if not form_data.get("recordNo") or not form_data.get("recordDate"):
            return False
        rows = form_data.get("rows", [])
        if not rows:
            return False
        for row in rows:
            cycles = row.get("cycles", [])
            if isinstance(cycles, str):
                try:
                    cycles = json.loads(cycles)
                except Exception:
                    return False
            if not cycles or not isinstance(cycles, list):
                return False
            if not any(str(c).strip() for c in cycles if c):
                return False
        return True

    if form_type == "PED":
        if not all([form_data.get("customer"), form_data.get("bid"),
                    form_data.get("po"), form_data.get("cr")]):
            return False
        if not form_data.get("recordNo") or not form_data.get("recordDate"):
            return False
        rows = form_data.get("rows", [])
        if not rows:
            return False
        for row in rows:
            ped_cycles = row.get("pedCycles", [])
            if isinstance(ped_cycles, str):
                try:
                    ped_cycles = json.loads(ped_cycles)
                except Exception:
                    ped_cycles = []

            notes = row.get("notes", [])
            if isinstance(notes, str):
                try:
                    notes = json.loads(notes)
                except Exception:
                    notes = []

            remarks = str(row.get("remarks", "")).strip()

            has_cycles = isinstance(ped_cycles, list) and any(str(c).strip() for c in ped_cycles if c)
            has_notes = isinstance(notes, list) and any(str(n).strip() for n in notes if n)
            has_remarks = bool(remarks)

            if not (has_cycles or has_notes or has_remarks):
                return False
        return True

    if form_type == "LEAD":
        if not all([form_data.get("customer"), form_data.get("bid"),
                    form_data.get("po"), form_data.get("cr")]):
            return False
        if not form_data.get("recordNo") or not form_data.get("recordDate"):
            return False
        rows = form_data.get("rows", [])
        if not rows:
            return False
        for row in rows:
            cust_date = str(row.get("customerRequiredDate", "")).strip()
            std_lead = str(row.get("standardLeadTime", "")).strip()
            gtn_date = str(row.get("gtnAgreedDate", "")).strip()
            if not cust_date or not std_lead or not gtn_date:
                return False
        return True

    return False

def handle_form_completion_notification(form_type, form_id, form_data, user_id, username):
    """Check and notify if form is newly completed"""
    conn = get_db()
    try:
        is_complete = check_form_completion(form_type, form_data)

        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT status
                  FROM form_completion_states
                 WHERE form_type = %s
                   AND form_id   = %s
                """,
                (form_type, form_id)
            )
            existing = cur.fetchone()

            if is_complete:
                if not existing or existing["status"] != "complete":
                    completed_at = now_ist().strftime("%Y-%m-%d %H:%M:%S")
                    cur.execute(
                        """
                        INSERT INTO form_completion_states
                          (form_type, form_id, status, completed_by, completed_at)
                        VALUES (%s, %s, 'complete', %s, %s)
                        ON CONFLICT(form_type, form_id)
                        DO UPDATE SET
                          status       = EXCLUDED.status,
                          completed_by = EXCLUDED.completed_by,
                          completed_at = EXCLUDED.completed_at
                        """,
                        (form_type, form_id, user_id, completed_at)
                    )
                    conn.commit()

                    po_info = f"{form_data.get('customer')} - {form_data.get('po')}"
                    message = f"{username} completed {form_type} form for PO: {po_info}"

                    create_notification_for_all_users(
                        event_type="form_completed",
                        message=message,
                        actor_user_id=user_id,
                        form_type=form_type,
                        form_id=form_id,
                        exclude_user_id=user_id
                    )
            else:
                if existing and existing["status"] == "complete":
                    cur.execute(
                        """
                        UPDATE form_completion_states
                           SET status = 'in_progress'
                         WHERE form_type = %s
                           AND form_id   = %s
                        """,
                        (form_type, form_id)
                    )
                    conn.commit()
    except Exception as e:
        conn.rollback()
        print("Error handling form completion:", e)
    finally:
        conn.close()

# --------------------------------------------------------------------------------------
# Per-department completion helpers
# --------------------------------------------------------------------------------------

def upsert_department_state(form_type, form_id, department, completed_by):
    conn = get_db()
    try:
        now = now_ist().isoformat(sep=" ")
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO form_department_states
                  (form_type, form_id, department, status, completed_by, completed_at)
                VALUES (%s, %s, %s, 'completed', %s, %s)
                ON CONFLICT(form_type, form_id, department)
                DO UPDATE SET
                  status       = 'completed',
                  completed_by = EXCLUDED.completed_by,
                  completed_at = EXCLUDED.completed_at
                """,
                (form_type, form_id, department, completed_by, now)
            )
        conn.commit()
    except Exception as e:
        conn.rollback()
        print("upsert_department_state error:", e)
    finally:
        conn.close()

def department_is_completed(form_type, form_id, department):
    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT status
                  FROM form_department_states
                 WHERE form_type = %s
                   AND form_id   = %s
                   AND department = %s
                """,
                (form_type, form_id, department)
            )
            row = cur.fetchone()
        conn.close()
        return bool(row and row[0] == "completed")
    except Exception as e:
        conn.close()
        print("department_is_completed error:", e)
        return False

def all_departments_completed(form_type, form_id, dept_order):
    for d in dept_order:
        if not department_is_completed(form_type, form_id, d):
            return False
    return True

def next_department_after(current_dept, dept_order):
    try:
        idx = dept_order.index(current_dept)
    except ValueError:
        return None
    if idx + 1 < len(dept_order):
        return dept_order[idx + 1]
    return None

def notify_users_of_department(
    department, subject, message_html,
    actor_user_id=None, form_type=None, form_id=None,
    exclude_user_id=None
):
    """Notify all users in a particular department (db + email)."""
    conn = get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                "SELECT id, email, name FROM users WHERE department = %s",
                (department,)
            )
            users = cur.fetchall()

            for u in users:
                if exclude_user_id and u["id"] == exclude_user_id:
                    continue
                cur.execute(
                    """
                    INSERT INTO notifications
                      (recipient_user_id, actor_user_id, event_type, form_type, form_id, message)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    """,
                    (u["id"], actor_user_id, f"{form_type.lower()}_dept_completed",
                     form_type, form_id, message_html)
                )
        conn.commit()
        conn.close()

        conn2 = get_db()
        try:
            with conn2.cursor() as cur2:
                for u in users:
                    if exclude_user_id and u["id"] == exclude_user_id:
                        continue
                    if not u["email"]:
                        continue
                    try:
                        sent = send_notification_email(u["email"], subject, message_html)
                        if sent:
                            cur2.execute(
                                """
                                UPDATE notifications
                                   SET email_sent = TRUE
                                 WHERE recipient_user_id = %s
                                   AND form_id = %s
                                   AND form_type = %s
                                """,
                                (u["id"], form_id, form_type)
                            )
                    except Exception as e:
                        print("email send failed:", e)
                conn2.commit()
        finally:
            conn2.close()
    except Exception as e:
        conn.rollback()
        conn.close()
        print("notify_users_of_department error:", e)

def notify_all_departments(
    subject, message_html,
    actor_user_id=None, form_type=None, form_id=None,
    exclude_user_id=None
):
    """Notify all users in the system (db + email)."""
    conn = get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT id, email, name FROM users")
            users = cur.fetchall()

            for u in users:
                if exclude_user_id and u["id"] == exclude_user_id:
                    continue
                cur.execute(
                    """
                    INSERT INTO notifications
                      (recipient_user_id, actor_user_id, event_type, form_type, form_id, message)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    """,
                    (u["id"], actor_user_id, f"{form_type.lower()}_all_completed",
                     form_type, form_id, message_html)
                )
        conn.commit()
        conn.close()

        conn2 = get_db()
        try:
            with conn2.cursor() as cur2:
                for u in users:
                    if exclude_user_id and u["id"] == exclude_user_id:
                        continue
                    if not u["email"]:
                        continue
                    try:
                        sent = send_notification_email(u["email"], subject, message_html)
                        if sent:
                            cur2.execute(
                                """
                                UPDATE notifications
                                   SET email_sent = TRUE
                                 WHERE recipient_user_id = %s
                                   AND form_id = %s
                                   AND form_type = %s
                                """,
                                (u["id"], form_id, form_type)
                            )
                    except Exception as e:
                        print("email send failed:", e)
                conn2.commit()
        finally:
            conn2.close()
    except Exception as e:
        conn.rollback()
        conn.close()
        print("notify_all_departments error:", e)

# --------------------------------------------------------------------------------------
# CR common items map
# --------------------------------------------------------------------------------------

def get_cr_common_item_map(conn, po_key):
    """
    Return a dict keyed by item_no for CR form of this po_key with core item fields:
    { item_no: { 'part_number': ..., 'part_description': ..., 'rev': ..., 'qty': ... }, ... }
    """
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                "SELECT id FROM cr_forms WHERE po_key = %s",
                (po_key,)
            )
            cr_form = cur.fetchone()
            if not cr_form:
                return {}
            cur.execute(
                """
                SELECT item_no, part_number, part_description, rev, qty
                  FROM cr_form_rows
                 WHERE cr_form_id = %s
                 ORDER BY id
                """,
                (cr_form["id"],)
            )
            rows = cur.fetchall()
        result = {}
        for r in rows:
            key = (r["item_no"] or "").strip()
            if not key:
                continue
            result[key] = {
                "part_number": r["part_number"] or "",
                "part_description": r["part_description"] or "",
                "rev": r["rev"] or "",
                "qty": r["qty"] or "",
            }
        return result
    except Exception as e:
        print("get_cr_common_item_map error:", e)
        return {}

#------------------------------------------------------
# Signature Master
#------------------------------------------------------
@app.route("/api/admin/signatures", methods=["GET"])
@admin_required
def list_master_signatures():
    conn = get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("""
                SELECT id, username, department, signature_path,
                       uploaded_by, uploaded_at, is_active
                FROM master_signatures
                ORDER BY department, username
            """)
            rows = cur.fetchall()
        return jsonify({"signatures": rows})
    finally:
        conn.close()

import hashlib

@app.route("/api/admin/signature/upload", methods=["POST"])
@admin_required
def upload_master_signature():
    file = request.files.get("signature")
    user_id = request.form.get("user_id")
    remarks = request.form.get("remarks", "")

    if not file or not user_id:
        return jsonify({"error": "Missing file or user"}), 400

    if file.mimetype not in ("image/png", "image/jpeg"):
        return jsonify({"error": "Only PNG/JPG allowed"}), 400

    data = file.read()
    size_kb = len(data) // 1024
    checksum = hashlib.sha256(data).hexdigest()

    filename = f"sig_{user_id}.png"
    save_path = os.path.join(app.static_folder, "signatures", filename)
    os.makedirs(os.path.dirname(save_path), exist_ok=True)

    with open(save_path, "wb") as f:
        f.write(data)

    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT username, department FROM users WHERE id=%s
            """, (user_id,))
            u = cur.fetchone()

            cur.execute("""
                INSERT INTO master_signatures
                  (user_id, username, department,
                   signature_file_name, signature_path,
                   signature_type, signature_size_kb,
                   checksum_sha256, remarks, uploaded_by)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (user_id, department)
                DO UPDATE SET
                  signature_file_name = EXCLUDED.signature_file_name,
                  signature_path = EXCLUDED.signature_path,
                  signature_type = EXCLUDED.signature_type,
                  signature_size_kb = EXCLUDED.signature_size_kb,
                  checksum_sha256 = EXCLUDED.checksum_sha256,
                  remarks = EXCLUDED.remarks,
                  updated_at = CURRENT_TIMESTAMP,
                  uploaded_by = EXCLUDED.uploaded_by
            """, (
                user_id, u[0], u[1],
                file.filename, f"/static/signatures/{filename}",
                file.mimetype, size_kb,
                checksum, remarks, session["username"]
            ))
        conn.commit()
        return jsonify({"success": True})
    finally:
        conn.close()
# ------------------------------------------------------
# Admin: List users (for Master Signature dropdown)
# ------------------------------------------------------
@app.route("/api/admin/users", methods=["GET"])
@admin_required
def admin_list_users():
    conn = get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("""
                SELECT id, name, department
                FROM users
                ORDER BY department, name
            """)
            users = cur.fetchall()
        return jsonify(users)
    finally:
        conn.close()
# ------------------------------------------------------
# Admin page: Master Signature
# ------------------------------------------------------
@app.route("/master-signature.html")
@admin_required
def master_signature_page():
    return send_from_directory(app.static_folder, "master-signature.html")


def get_cr_engineering_signature(conn, po_key):
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute("""
            SELECT
                cds.signed_by,
                cds.signed_at,
                ms.signature_path
            FROM cr_department_signatures cds
            JOIN master_signatures ms
              ON ms.username = cds.signed_by
             AND ms.department = cds.department
             AND ms.is_active = TRUE
            WHERE cds.po_key = %s
              AND cds.department = 'engineering'
        """, (po_key,))
        return cur.fetchone()
#-----------
#sign-department
#------------
@app.route('/api/forms/sign-department', methods=['POST'])
@login_required
def sign_department():
    data = request.get_json() or {}
    form_type = (data.get('formType') or '').strip().upper()
    po_key = (data.get('poKey') or '').strip()
    department = (data.get('department') or '').strip()

    # Validate form type
    if form_type not in ('CR', 'PED', 'LEAD'):
        return jsonify({'error': f'Invalid form type: {form_type}'}), 400

    if not po_key or not department:
        return jsonify({'error': 'Missing PO key or department'}), 400

    username = session.get('username')
    if not username:
        return jsonify({'error': 'Unauthorized'}), 401

    # Permissions
    is_admin = bool(session.get('user_is_admin') or session.get('is_admin'))
    user_dept = (session.get('user_department') or '').strip()

    # Users can sign only their own dept (admin can sign any dept)
    if not is_admin and user_dept != department:
        return jsonify({'error': 'Unauthorized: You can only sign for your own department'}), 403

    # Determine table name based on form type
    table_map = {
        'CR': 'cr_department_signatures',
        'PED': 'ped_department_signatures',
        'LEAD': 'lead_department_signatures'
    }
    table_name = table_map[form_type]

    conn = get_db()
    try:
        with conn.cursor() as cur:
            # Auto-create tables if they don't exist
            if form_type == 'PED':
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS ped_department_signatures (
                        id SERIAL PRIMARY KEY,
                        po_key TEXT NOT NULL,
                        department TEXT NOT NULL,
                        signed_by TEXT NOT NULL,
                        signed_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP,
                        UNIQUE (po_key, department)
                    )
                """)
            elif form_type == 'LEAD':
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS lead_department_signatures (
                        id SERIAL PRIMARY KEY,
                        po_key TEXT NOT NULL,
                        department TEXT NOT NULL,
                        signed_by TEXT NOT NULL,
                        signed_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP,
                        UNIQUE (po_key, department)
                    )
                """)

            # Check if already signed
            cur.execute(
                f"SELECT signed_by, signed_at FROM {table_name} WHERE po_key = %s AND department = %s",
                (po_key, department)
            )
            existing = cur.fetchone()
            if existing:
                # keep same behavior: cannot sign twice
                return jsonify({'error': f'{department} is already signed for {form_type}'}), 409

            # Insert signature
            cur.execute(
                f"INSERT INTO {table_name} (po_key, department, signed_by) VALUES (%s, %s, %s)",
                (po_key, department, username)
            )

        conn.commit()

        return jsonify({
            'success': True,
            'formType': form_type,
            'department': department,
            'signedBy': username,
            'timestamp': datetime.now().isoformat()
        })

    except Exception as e:
        conn.rollback()
        print(f"sign-department error ({form_type}):", e)
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()

#-------------------------------------------
#FORM Signature
#-------------------------------------------
@app.route('/api/forms/cr-signed-departments')
@login_required
def get_cr_signed_departments():
    po_key = (request.args.get('poKey') or '').strip()
    if not po_key:
        return jsonify({'signed': []})

    conn = get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("""
                SELECT department, signed_by, signed_at
                FROM cr_department_signatures
                WHERE po_key = %s
                ORDER BY signed_at
            """, (po_key,))
            rows = cur.fetchall()
        return jsonify({
            'signed': [
                {
                    'department': r['department'],
                    'signedBy': r['signed_by'],
                    'signedAt': r['signed_at']
                } for r in rows
            ]
        })
    except Exception as e:
        print("get_cr_signed_departments error:", e)
        return jsonify({'signed': []}), 500
    finally:
        conn.close()

@app.route('/api/forms/ped-signed-departments')
@login_required
def get_ped_signed_departments():
    po_key = (request.args.get('poKey') or '').strip()
    if not po_key:
        return jsonify({'signed': []})

    conn = get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("""
                SELECT department, signed_by, signed_at
                FROM ped_department_signatures
                WHERE po_key = %s
                ORDER BY signed_at
            """, (po_key,))
            rows = cur.fetchall()
        return jsonify({
            'signed': [
                {
                    'department': r['department'],
                    'signedBy': r['signed_by'],
                    'signedAt': r['signed_at']
                } for r in rows
            ]
        })
    except Exception as e:
        return jsonify({'signed': []}), 500
    finally:
        conn.close()

@app.route('/api/forms/lead-signed-departments')
@login_required
def get_lead_signed_departments():
    po_key = (request.args.get('poKey') or '').strip()
    if not po_key:
        return jsonify({'signed': []})

    conn = get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            # Ensure table exists so query doesn't crash on first run
            cur.execute("""
                CREATE TABLE IF NOT EXISTS lead_department_signatures (
                    id SERIAL PRIMARY KEY,
                    po_key TEXT NOT NULL,
                    department TEXT NOT NULL,
                    signed_by TEXT NOT NULL,
                    signed_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE (po_key, department)
                )
            """)
            conn.commit()

            cur.execute("""
                SELECT department, signed_by, signed_at
                FROM lead_department_signatures
                WHERE po_key = %s
                ORDER BY signed_at
            """, (po_key,))
            rows = cur.fetchall()
        return jsonify({
            'signed': [
                {
                    'department': r['department'],
                    'signedBy': r['signed_by'],
                    'signedAt': r['signed_at']
                } for r in rows
            ]
        })
    except Exception as e:
        return jsonify({'signed': []}), 500
    finally:
        conn.close()

# --------------------------------------------------------------------------------------
# Admin helpers for dashboard
# --------------------------------------------------------------------------------------

def get_form_department_status_for_admin():
    conn = get_db()
    results = []

    def dept_states(form_type, form_ids):
        if not form_ids:
            return {}
        placeholders = ",".join(["%s"] * len(form_ids))
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                f"""
                SELECT form_type, form_id, department, status
                  FROM form_department_states
                 WHERE form_type = %s
                   AND form_id IN ({placeholders})
                """,
                [form_type] + list(form_ids)
            )
            rows = cur.fetchall()
        out = {}
        for r in rows:
            key = r["form_id"]
            if key not in out:
                out[key] = {}
            out[key][r["department"]] = r["status"]
        return out

    def pick_activity_date(row, fallback_field):
        raw = (
            row.get("completed_at")
            or row.get("last_modified_at")
            or row.get(fallback_field)
        )
        if not raw:
            return ""
        return str(raw)[:10]

    try:
        # CR
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT f.id,
                       f.po_key, f.customer, f.bid, f.po, f.cr,
                       f.last_modified_at,
                       s.status AS overall_status, s.completed_at
                  FROM cr_forms f
             LEFT JOIN form_completion_states s
                    ON s.form_type = 'CR' AND s.form_id = f.id
                 ORDER BY f.id DESC
                """
            )
            cr_forms = cur.fetchall()
        cr_ids = [r["id"] for r in cr_forms]
        cr_dept = dept_states("CR", cr_ids)

        for f in cr_forms:
            dept_map = cr_dept.get(f["id"], {})
            dept_info = []
            for d in CR_DEPT_ORDER:
                state = dept_map.get(d, "in_progress")
                dept_info.append({"department": d, "status": state})

            base_status = f["overall_status"] or "in_progress"
            all_depts_done = all(
                dept_map.get(d) == "completed" for d in CR_DEPT_ORDER
            ) if CR_DEPT_ORDER else True

            if base_status == "complete" and all_depts_done:
                dashboard_status = "complete"
            else:
                dashboard_status = "in_progress"

            results.append({
                "formType": "CR",
                "formId": f["id"],
                "poKey": f["po_key"],
                "customer": f["customer"],
                "bid": f["bid"],
                "po": f["po"],
                "cr": f["cr"],
                "overallStatus": dashboard_status,
                "completedAt": f["completed_at"],
                "date": pick_activity_date(f, "last_modified_at"),
                "departments": dept_info,
            })

        # PED
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT f.id,
                       f.po_key, f.customer, f.bid, f.po, f.cr,
                       f.last_modified_at,
                       s.status AS overall_status, s.completed_at
                  FROM ped_forms f
             LEFT JOIN form_completion_states s
                    ON s.form_type = 'PED' AND s.form_id = f.id
                 ORDER BY f.id DESC
                """
            )
            ped_forms = cur.fetchall()
        ped_ids = [r["id"] for r in ped_forms]
        ped_dept = dept_states("PED", ped_ids)

        for f in ped_forms:
            dept_map = ped_dept.get(f["id"], {})
            dept_info = []
            for d in PED_DEPT_ORDER:
                state = dept_map.get(d, "in_progress")
                dept_info.append({"department": d, "status": state})

            base_status = f["overall_status"] or "in_progress"
            all_depts_done = all(
                dept_map.get(d) == "completed" for d in PED_DEPT_ORDER
            ) if PED_DEPT_ORDER else True

            if base_status == "complete" and all_depts_done:
                dashboard_status = "complete"
            else:
                dashboard_status = "in_progress"

            results.append({
                "formType": "PED",
                "formId": f["id"],
                "poKey": f["po_key"],
                "customer": f["customer"],
                "bid": f["bid"],
                "po": f["po"],
                "cr": f["cr"],
                "overallStatus": dashboard_status,
                "completedAt": f["completed_at"],
                "date": pick_activity_date(f, "last_modified_at"),
                "departments": dept_info,
            })

        # LEAD
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT f.id,
                       f.po_key, f.customer, f.bid, f.po, f.cr,
                       f.last_modified_at,
                       s.status AS overall_status, s.completed_at
                  FROM lead_forms f
             LEFT JOIN form_completion_states s
                    ON s.form_type = 'LEAD' AND s.form_id = f.id
                 ORDER BY f.id DESC
                """
            )
            lead_forms = cur.fetchall()

        for f in lead_forms:
            results.append({
                "formType": "LEAD",
                "formId": f["id"],
                "poKey": f["po_key"],
                "customer": f["customer"],
                "bid": f["bid"],
                "po": f["po"],
                "cr": f["cr"],
                "overallStatus": f["overall_status"] or "in_progress",
                "completedAt": f["completed_at"],
                "date": pick_activity_date(f, "last_modified_at"),
                "departments": [],
            })
    except Exception as e:
        print("get_form_department_status_for_admin error:", e)
    finally:
        conn.close()
    return results

# --------------------------------------------------------------------------------------
# PO-level overview
# --------------------------------------------------------------------------------------

def get_po_overview():
    conn = get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT id, customer, bid, po, cr
                  FROM pos
                 ORDER BY id DESC
                """
            )
            po_rows = cur.fetchall()

            cur.execute("SELECT id, po_key, last_modified_at FROM cr_forms")
            cr_forms = cur.fetchall()
            cur.execute("SELECT id, po_key, last_modified_at FROM ped_forms")
            ped_forms = cur.fetchall()
            cur.execute("SELECT id, po_key, last_modified_at FROM lead_forms")
            lead_forms = cur.fetchall()

            cur.execute(
                """
                SELECT form_type, form_id, status, completed_at
                  FROM form_completion_states
                """
            )
            states = cur.fetchall()

        def index_forms(rows):
            out = {}
            for r in rows:
                key = r["po_key"]
                if key not in out:
                    out[key] = []
                out[key].append(r)
            return out

        cr_by_key = index_forms(cr_forms)
        ped_by_key = index_forms(ped_forms)
        lead_by_key = index_forms(lead_forms)

        state_index = {}
        for s in states:
            state_index[(s["form_type"], s["form_id"])] = {
                "status": s["status"],
                "completed_at": s["completed_at"],
            }

        def summarize_form_list(form_type, forms_for_po):
            if not forms_for_po:
                return "not_started", None, None
            any_in_progress = False
            any_complete = False
            dates = []
            for frm in forms_for_po:
                st = state_index.get((form_type, frm["id"]), {})
                status = st.get("status") or "in_progress"
                if status == "complete":
                    any_complete = True
                else:
                    any_in_progress = True
                if st.get("completed_at"):
                    dates.append(str(st["completed_at"]))
                if frm["last_modified_at"]:
                    dates.append(str(frm["last_modified_at"]))
            if any_in_progress:
                summary_status = "in_progress"
            elif any_complete:
                summary_status = "complete"
            else:
                summary_status = "in_progress"
            if dates:
                dates_sorted = sorted(dates)
                oldest = dates_sorted[0][:10]
                newest = dates_sorted[-1][:10]
            else:
                oldest = newest = None
            return summary_status, oldest, newest

        overview = []
        for po in po_rows:
            po_key = f"{po['customer']}|{po['bid']}|{po['po']}|{po['cr']}"
            cr_list = cr_by_key.get(po_key, [])
            ped_list = ped_by_key.get(po_key, [])
            lead_list = lead_by_key.get(po_key, [])

            cr_status, cr_oldest, cr_newest = summarize_form_list("CR", cr_list)
            ped_status, ped_oldest, ped_newest = summarize_form_list("PED", ped_list)
            lead_status, lead_oldest, lead_newest = summarize_form_list("LEAD", lead_list)

            all_dates = [
                d for d in
                [cr_oldest, cr_newest, ped_oldest, ped_newest, lead_oldest, lead_newest]
                if d
            ]
            if all_dates:
                all_dates_sorted = sorted(all_dates)
                oldest_activity = all_dates_sorted[0]
                latest_activity = all_dates_sorted[-1]
            else:
                oldest_activity = ""
                latest_activity = ""

            overview.append({
                "poId": po["id"],
                "customer": po["customer"],
                "bid": po["bid"],
                "po": po["po"],
                "cr": po["cr"],
                "crStatus": cr_status,
                "pedStatus": ped_status,
                "leadStatus": lead_status,
                "oldestActivity": oldest_activity,
                "latestActivity": latest_activity,
            })
        return overview
    except Exception as e:
        print("get_po_overview error:", e)
        return []
    finally:
        conn.close()

# --------------------------------------------------------------------------------------
# Admin endpoints (status, reminders, etc.)
# --------------------------------------------------------------------------------------

@app.route("/api/admin/forms/status", methods=["GET"])
@admin_required
def admin_forms_status():
    form_type_filter = (request.args.get("formType") or "").strip().upper()
    only_pending = (request.args.get("onlyPending") or "").lower() == "true"
    try:
        data = get_form_department_status_for_admin()
        if form_type_filter in ("CR", "PED", "LEAD"):
            data = [f for f in data if f["formType"] == form_type_filter]
        if only_pending:
            data = [f for f in data if f["overallStatus"] != "complete"]
        return jsonify({"forms": data})
    except Exception as e:
        print("admin_forms_status error:", e)
        return jsonify({"error": str(e)}), 500

@app.route("/api/admin/forms/status-by-date", methods=["GET"])
@admin_required
def admin_forms_status_by_date():
    form_type_filter = (request.args.get("formType") or "").strip().upper()
    only_pending = (request.args.get("onlyPending") or "").lower() == "true"
    date_filter = (request.args.get("date") or "").strip()

    try:
        all_data = get_form_department_status_for_admin()
        for f in all_data:
            d = (f.get("date") or "")[:10]
            f["date"] = d

        forms = all_data
        if form_type_filter in ("CR", "PED", "LEAD"):
            forms = [f for f in forms if f["formType"] == form_type_filter]
        if only_pending:
            forms = [f for f in forms if f["overallStatus"] != "complete"]
        if date_filter:
            forms = [f for f in forms if f["date"] == date_filter]

        from collections import defaultdict
        date_stats = defaultdict(lambda: {
            "total": 0,
            "pending": 0,
            "completed": 0,
            "byType": {
                "CR": {"total": 0, "pending": 0, "completed": 0},
                "PED": {"total": 0, "pending": 0, "completed": 0},
                "LEAD": {"total": 0, "pending": 0, "completed": 0},
            },
        })

        for f in forms:
            d = f["date"] or "Unknown"
            t = f["formType"]
            st = f["overallStatus"] or "in_progress"
            bucket = date_stats[d]
            bucket["total"] += 1
            if st == "complete":
                bucket["completed"] += 1
                bucket["byType"][t]["completed"] += 1
            else:
                bucket["pending"] += 1
                bucket["byType"][t]["pending"] += 1
            bucket["byType"][t]["total"] += 1

        dates_sorted = sorted(date_stats.items(), key=lambda x: x[0], reverse=True)
        dates_payload = [{"date": d, "stats": s} for d, s in dates_sorted]

        return jsonify({"dates": dates_payload, "forms": forms})
    except Exception as e:
        print("admin_forms_status_by_date error:", e)
        return jsonify({"error": str(e)}), 500

@app.route("/api/admin/po-overview", methods=["GET"])
@admin_required
def admin_po_overview():
    only_pending = (request.args.get("onlyPending") or "").lower() == "true"
    try:
        rows = get_po_overview()
        if only_pending:
            rows = [
                r for r in rows
                if r["crStatus"] != "complete"
                or r["pedStatus"] != "complete"
                or r["leadStatus"] != "complete"
            ]
        return jsonify({"pos": rows})
    except Exception as e:
        print("admin_po_overview error:", e)
        return jsonify({"error": str(e)}), 500

@app.route("/api/admin/forms/remind", methods=["POST"])
@admin_required
def admin_remind_department():
    data = request.get_json() or {}
    form_type = (data.get("formType") or "").strip().upper()
    form_id = data.get("formId")
    department = (data.get("department") or "").strip().lower()
    user_id_target = data.get("userId")
    custom_message = (data.get("customMessage") or "").strip()

    if form_type not in ("CR", "PED", "LEAD"):
        return jsonify({"error": "Invalid formType"}), 400
    if not form_id:
        return jsonify({"error": "formId is required"}), 400

    conn = get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            if form_type == "CR":
                cur.execute(
                    "SELECT po_key, customer, po, cr FROM cr_forms WHERE id = %s",
                    (form_id,)
                )
                form = cur.fetchone()
                valid_depts = CR_DEPT_ORDER
            elif form_type == "PED":
                cur.execute(
                    "SELECT po_key, customer, po, cr FROM ped_forms WHERE id = %s",
                    (form_id,)
                )
                form = cur.fetchone()
                valid_depts = PED_DEPT_ORDER
            else:
                cur.execute(
                    "SELECT po_key, customer, po, cr FROM lead_forms WHERE id = %s",
                    (form_id,)
                )
                form = cur.fetchone()
                valid_depts = []

        if not form:
            conn.close()
            return jsonify({"error": f"{form_type} form not found"}), 404

        actor_user_id = session.get("user_id")
        actor_name = session.get("user_name", session.get("username", "Admin"))
        po_info = f"{form['customer']} - {form['po']} (CR: {form['cr']})"

        if form_type in ("CR", "PED"):
            if department not in valid_depts:
                conn.close()
                return jsonify({"error": "Invalid or unsupported department for this form type"}), 400

            conn2 = get_db()
            try:
                with conn2.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur2:
                    if user_id_target:
                        cur2.execute(
                            """
                            SELECT id, email, name
                              FROM users
                             WHERE id = %s AND department = %s
                            """,
                            (user_id_target, department)
                        )
                    else:
                        cur2.execute(
                            """
                            SELECT id, email, name
                              FROM users
                             WHERE department = %s
                            """,
                            (department,)
                        )
                    users = cur2.fetchall()

                    if not users:
                        conn2.close()
                        return jsonify({"error": "No users found for this department / userId"}), 404

                    subject = f"Reminder: {form_type} Review Pending ({department.title()})"
                    if custom_message:
                        msg_html = f"<p>{custom_message}</p>"
                    else:
                        msg_html = (
                            f"<p>Dear {department.title()} Team,</p>"
                            f"<p>This is a reminder that your review is pending "
                            f"for {form_type} form of PO: <strong>{po_info}</strong>.</p>"
                            f"<p>Please complete your section in the Contract Review Dashboard.</p>"
                            f"<p>Regards,<br>{actor_name}</p>"
                        )

                    sent_ids = []
                    for u in users:
                        notif_message = (
                            f"{actor_name} sent a reminder to {department} for "
                            f"{form_type} form ({po_info})."
                        )
                        cur2.execute(
                            """
                            INSERT INTO notifications
                              (recipient_user_id, actor_user_id, event_type, form_type, form_id, message)
                            VALUES (%s, %s, %s, %s, %s, %s)
                            """,
                            (
                                u["id"], actor_user_id,
                                f"{form_type.lower()}_dept_reminder",
                                form_type, form_id, notif_message
                            )
                        )
                        if u["email"]:
                            try:
                                sent = send_notification_email(u["email"], subject, msg_html)
                                if sent:
                                    cur2.execute(
                                        """
                                        UPDATE notifications
                                           SET email_sent = TRUE
                                         WHERE recipient_user_id = %s
                                           AND form_type = %s
                                           AND form_id   = %s
                                        """,
                                        (u["id"], form_type, form_id)
                                    )
                            except Exception as e:
                                print("Reminder email send failed:", e)
                        sent_ids.append(u["id"])
                    conn2.commit()
                conn2.close()
                return jsonify({"success": True, "sentTo": sent_ids})
            except Exception as e:
                conn2.rollback()
                conn2.close()
                print("admin_remind_department CR/PED error:", e)
                return jsonify({"error": str(e)}), 500

        # LEAD reminders
        conn3 = get_db()
        try:
            with conn3.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur3:
                if user_id_target:
                    cur3.execute(
                        "SELECT id, email, name FROM users WHERE id = %s",
                        (user_id_target,)
                    )
                else:
                    cur3.execute("SELECT id, email, name FROM users")
                users = cur3.fetchall()
                if not users:
                    conn3.close()
                    return jsonify({"error": "No users found"}), 404

                subject = "Reminder: LEAD Form Pending"
                if custom_message:
                    msg_html = f"<p>{custom_message}</p>"
                else:
                    msg_html = (
                        f"<p>This is a reminder that the LEAD form for "
                        f"PO: <strong>{po_info}</strong> is still pending.</p>"
                        f"<p>Please review and complete it in the Contract Review Dashboard.</p>"
                        f"<p>Regards,<br>{actor_name}</p>"
                    )

                sent_ids = []
                for u in users:
                    notif_message = f"{actor_name} sent a reminder for LEAD form ({po_info})."
                    cur3.execute(
                        """
                        INSERT INTO notifications
                          (recipient_user_id, actor_user_id, event_type, form_type, form_id, message)
                        VALUES (%s, %s, %s, %s, %s, %s)
                        """,
                        (u["id"], actor_user_id, "lead_form_reminder", form_type, form_id, notif_message)
                    )
                    if u["email"]:
                        try:
                            sent = send_notification_email(u["email"], subject, msg_html)
                            if sent:
                                cur3.execute(
                                    """
                                    UPDATE notifications
                                       SET email_sent = TRUE
                                     WHERE recipient_user_id = %s
                                       AND form_type = %s
                                       AND form_id   = %s
                                    """,
                                    (u["id"], form_type, form_id)
                                )
                        except Exception as e:
                            print("Reminder email send failed:", e)
                    sent_ids.append(u["id"])
                conn3.commit()
            conn3.close()
            return jsonify({"success": True, "sentTo": sent_ids})
        except Exception as e:
            conn3.rollback()
            conn3.close()
            print("admin_remind_department LEAD error:", e)
            return jsonify({"error": str(e)}), 500

    except Exception as e:
        conn.close()
        print("admin_remind_department error:", e)
        return jsonify({"error": str(e)}), 500

# --------------------------------------------------------------------------------------
# Static + auth endpoints
# --------------------------------------------------------------------------------------

@app.route("/")
def index():
    return send_from_directory("static", "login.html")

@app.route("/<path:path>")
def serve_static(path):
    return send_from_directory("static", path)

@app.route("/api/login", methods=["POST"])
def login():
    data = request.get_json()
    username = data.get("username", "").strip()
    password = data.get("password", "")

    if not username or not password:
        return jsonify({"error": "Username and password required"}), 400

    conn = get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                "SELECT * FROM users WHERE username = %s",
                (username,)
            )
            user = cur.fetchone()
        conn.close()
    except Exception as e:
        conn.close()
        print("login error:", e)
        return jsonify({"error": "Server error"}), 500

    if user and check_password_hash(user["password_hash"], password):
        session["user_id"] = user["id"]
        session["username"] = user["username"]
        session["user_department"] = user["department"]
        session["user_name"] = user["name"]
        session["user_is_admin"] = bool(user["is_admin"])

        return jsonify({
            "success": True,
            "user": {
                "username": user["username"],
                "name": user["name"],
                "department": user["department"],
                "isAdmin": bool(user["is_admin"]),
            },
        })

    return jsonify({"error": "Invalid username or password"}), 401

@app.route("/api/logout", methods=["POST"])
def logout():
    session.clear()
    return jsonify({"success": True})

@app.route("/api/session", methods=["GET"])
def get_session():
    if "user_id" in session:
        return jsonify({
            "loggedIn": True,
            "user": {
                "username": session.get("username"),
                "name": session.get("user_name"),
                "department": session.get("user_department"),
                "isAdmin": session.get("user_is_admin", False),
            },
        })
    return jsonify({"loggedIn": False})

# --------------------------------------------------------------------------------------
# User management
# --------------------------------------------------------------------------------------

@app.route("/api/users", methods=["GET"])
@admin_required
def get_users():
    conn = get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT id, username, name, department, email, is_admin, lead_form_access
                  FROM users
                 ORDER BY id
                """
            )
            users = cur.fetchall()
        conn.close()
        return jsonify([{
            "id": u["id"],
            "username": u["username"],
            "name": u["name"],
            "department": u["department"],
            "email": u["email"] or "",
            "isAdmin": bool(u["is_admin"]),
            "leadFormAccess": bool(u["lead_form_access"]),
        } for u in users])
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500

@app.route("/api/users", methods=["POST"])
@admin_required
def create_user():
    import traceback

    data = request.get_json(silent=True) or {}
    username = (data.get("username") or "").strip()
    password = data.get("password") or ""
    name = (data.get("name") or "").strip()
    department = (data.get("department") or "").strip()
    email = (data.get("email") or "").strip()
    is_admin = bool(data.get("isAdmin", False))
    lead_form_access = bool(data.get("leadFormAccess", False))

    if not all([username, password, name, department]):
        return jsonify({"error": "Username, password, name, and department are required"}), 400

    conn = get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT id FROM users WHERE username = %s", (username,))
            existing = cur.fetchone()
            if existing:
                conn.close()
                return jsonify({"error": "Username already exists"}), 400

            if email:
                cur.execute(
                    "SELECT id FROM users WHERE email = %s AND email <> ''",
                    (email,)
                )
                existing_email = cur.fetchone()
                if existing_email:
                    conn.close()
                    return jsonify({"error": "Email already exists"}), 400

            cur.execute(
                """
                INSERT INTO users
                  (username, password_hash, name, department, email, is_admin, lead_form_access)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                RETURNING id
                """,
                (
                    username,
                    generate_password_hash(password),
                    name,
                    department,
                    email,
                    True if is_admin else False,
                    True if lead_form_access else False,
                )
            )
            row = cur.fetchone()
            user_id = row["id"]  # <-- FIX: dict access

        conn.commit()
        conn.close()
        return jsonify({
            "success": True,
            "user": {
                "id": user_id,
                "username": username,
                "name": name,
                "department": department,
                "email": email,
                "isAdmin": is_admin,
                "leadFormAccess": lead_form_access,
            },
        })
    except Exception as e:
        conn.rollback()
        traceback.print_exc()
        conn.close()
        return jsonify({"error": str(e)}), 500

@app.route("/api/users/<int:user_id>", methods=["PUT"])
@admin_required
def update_user(user_id):
    data = request.get_json()
    name = data.get("name", "").strip()
    department = data.get("department", "").strip()
    email = data.get("email", "").strip()
    is_admin = data.get("isAdmin", False)
    lead_form_access = data.get("leadFormAccess", False)
    password = data.get("password", "").strip()

    if not all([name, department]):
        return jsonify({"error": "Name and department required"}), 400

    conn = get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                "SELECT username, is_admin FROM users WHERE id = %s",
                (user_id,)
            )
            user = cur.fetchone()
            if not user:
                conn.close()
                return jsonify({"error": "User not found"}), 404

            if email:
                cur.execute(
                    """
                    SELECT id
                      FROM users
                     WHERE email = %s
                       AND email <> ''
                       AND id <> %s
                    """,
                    (email, user_id)
                )
                existing_email = cur.fetchone()
                if existing_email:
                    conn.close()
                    return jsonify({"error": "Email already exists"}), 400

            if user["username"] == "admin" and user["is_admin"] and not is_admin:
                cur.execute(
                    "SELECT COUNT(*) FROM users WHERE is_admin = TRUE"
                )
                admin_count = cur.fetchone()[0]
                if admin_count <= 1:
                    conn.close()
                    return jsonify(
                        {"error": "Cannot remove admin status from last admin user"}
                    ), 400

            if password:
                cur.execute(
                    """
                    UPDATE users
                       SET name = %s,
                           department = %s,
                           email = %s,
                           is_admin = %s,
                           lead_form_access = %s,
                           password_hash = %s
                     WHERE id = %s
                    """,
                    (
                        name, department, email,
                        True if is_admin else False,
                        True if lead_form_access else False,
                        generate_password_hash(password),
                        user_id,
                    )
                )
            else:
                cur.execute(
                    """
                    UPDATE users
                       SET name = %s,
                           department = %s,
                           email = %s,
                           is_admin = %s,
                           lead_form_access = %s
                     WHERE id = %s
                    """,
                    (
                        name, department, email,
                        True if is_admin else False,
                        True if lead_form_access else False,
                        user_id,
                    )
                )
        conn.commit()
        conn.close()
        return jsonify({
            "success": True,
            "user": {
                "id": user_id,
                "username": user["username"],
                "name": name,
                "department": department,
                "email": email,
                "isAdmin": is_admin,
                "leadFormAccess": lead_form_access,
            },
        })
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({"error": str(e)}), 500

@app.route("/api/users/<int:user_id>", methods=["DELETE"])
@admin_required
def delete_user(user_id):
    conn = get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                "SELECT username, is_admin FROM users WHERE id = %s",
                (user_id,)
            )
            user = cur.fetchone()
            if not user:
                conn.close()
                return jsonify({"error": "User not found"}), 404

            if user["username"] == "admin" and user["is_admin"]:
                conn.close()
                return jsonify({"error": "Cannot delete default admin user"}), 400

            cur.execute(
                "SELECT COUNT(*) FROM users WHERE is_admin = TRUE"
            )
            admin_count = cur.fetchone()[0]
            if user["is_admin"] and admin_count <= 1:
                conn.close()
                return jsonify({"error": "Cannot delete last admin user"}), 400

            cur.execute("DELETE FROM users WHERE id = %s", (user_id,))
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({"error": str(e)}), 500

# --------------------------------------------------------------------------------------
# PO endpoints
# --------------------------------------------------------------------------------------

@app.route("/api/pos", methods=["GET"])
@login_required
def get_pos():
    conn = get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT * FROM pos ORDER BY id DESC")
            pos_list = cur.fetchall()
        conn.close()
        return jsonify([{
            "id": po["id"],
            "customer": po["customer"],
            "bid": po["bid"],
            "po": po["po"],
            "cr": po["cr"],
        } for po in pos_list])
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500

@app.route("/api/pos", methods=["POST"])
@admin_required
def create_po():
    data = request.get_json()
    customer = data.get("customer", "").strip()
    bid = data.get("bid", "").strip()
    po = data.get("po", "").strip()
    cr = data.get("cr", "").strip()

    if not all([customer, bid, po, cr]):
        return jsonify({"error": "All fields required"}), 400

    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO pos (customer, bid, po, cr)
                VALUES (%s, %s, %s, %s)
                RETURNING id
                """,
                (customer, bid, po, cr)
            )
            po_id = cur.fetchone()[0]
        conn.commit()
        conn.close()

        actor_user_id = session.get("user_id")
        actor_username = session.get("username", "Admin")
        message = f"New PO created by {actor_username}: {customer} - {po} (CR: {cr})"

        create_notification_for_all_users(
            event_type="po_created",
            message=message,
            actor_user_id=actor_user_id,
            po_id=po_id,
            exclude_user_id=actor_user_id,
        )

        return jsonify({
            "success": True,
            "po": {
                "id": po_id,
                "customer": customer,
                "bid": bid,
                "po": po,
                "cr": cr,
            },
        })
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({"error": str(e)}), 500

@app.route("/api/pos/<int:po_id>", methods=["PUT"])
@admin_required
def update_po(po_id):
    data = request.get_json()
    customer = data.get("customer", "").strip()
    bid = data.get("bid", "").strip()
    po = data.get("po", "").strip()
    cr = data.get("cr", "").strip()

    if not all([customer, bid, po, cr]):
        return jsonify({"error": "All fields required"}), 400

    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                UPDATE pos
                   SET customer = %s,
                       bid      = %s,
                       po       = %s,
                       cr       = %s,
                       updated_at = CURRENT_TIMESTAMP
                 WHERE id = %s
                """,
                (customer, bid, po, cr, po_id)
            )
        conn.commit()
        conn.close()
        return jsonify({
            "success": True,
            "po": {
                "id": po_id,
                "customer": customer,
                "bid": bid,
                "po": po,
                "cr": cr,
            },
        })
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({"error": str(e)}), 500

@app.route("/api/pos/<int:po_id>", methods=["DELETE"])
@admin_required
def delete_po(po_id):
    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM pos WHERE id = %s", (po_id,))
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({"error": str(e)}), 500

# --------------------------------------------------------------------------------------
# Backup / restore (users + pos)
# --------------------------------------------------------------------------------------

@app.route("/api/backup", methods=["GET"])
@admin_required
def backup_data():
    conn = get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                "SELECT id, username, password_hash, name, department, is_admin FROM users"
            )
            users = cur.fetchall()
            cur.execute("SELECT * FROM pos")
            pos_list = cur.fetchall()
        conn.close()

        return jsonify({
            "meta": {
                "app": "GTN-ContractReview",
                "version": "1.0",
                "exportedAt": now_ist().isoformat(),
                "by": session.get("username", ""),
            },
            "users": [{
                "username": u["username"],
                "password_hash": u["password_hash"],
                "name": u["name"],
                "department": u["department"],
                "isAdmin": bool(u["is_admin"]),
            } for u in users],
            "pos": [{
                "customer": p["customer"],
                "bid": p["bid"],
                "po": p["po"],
                "cr": p["cr"],
            } for p in pos_list],
        })
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500

@app.route("/api/restore", methods=["POST"])
@admin_required
def restore_data():
    data = request.get_json()

    if not data or not isinstance(data, dict):
        return jsonify({"error": "Invalid data format"}), 400

    if not data.get("users") or not isinstance(data["users"], list):
        return jsonify({"error": "Missing or invalid users array"}), 400

    if not data.get("pos") or not isinstance(data["pos"], list):
        return jsonify({"error": "Missing or invalid pos array"}), 400

    for user in data["users"]:
        if not all(k in user for k in ["username", "password_hash", "name", "department"]):
            return jsonify({"error": "Invalid user data: missing required fields"}), 400
        if not user.get("password_hash") or not user["password_hash"].startswith(
            ("pbkdf2:", "scrypt:", "bcrypt:")
        ):
            return jsonify(
                {"error": "Invalid user data: password_hash must be a valid hash"}
            ), 400

    has_admin = any(u.get("isAdmin") for u in data["users"])
    if not has_admin:
        return jsonify({"error": "Backup must contain at least one admin user"}), 400

    for po in data["pos"]:
        if not all(k in po for k in ["customer", "bid", "po", "cr"]):
            return jsonify({"error": "Invalid PO data: missing required fields"}), 400

    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM users")
            cur.execute("DELETE FROM pos")

            for user in data["users"]:
                cur.execute(
                    """
                    INSERT INTO users (username, password_hash, name, department, is_admin)
                    VALUES (%s, %s, %s, %s, %s)
                    """,
                    (
                        user["username"],
                        user["password_hash"],
                        user["name"],
                        user["department"],
                        True if user.get("isAdmin") else False,
                    )
                )

            for po in data["pos"]:
                cur.execute(
                    """
                    INSERT INTO pos (customer, bid, po, cr)
                    VALUES (%s, %s, %s, %s)
                    """,
                    (po["customer"], po["bid"], po["po"], po["cr"])
                )

        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({"error": str(e)}), 500

# ========== CR / PED / LEAD SAVE & LOAD (POSTGRES) ==========

@app.route('/api/cr-form/save', methods=['POST'])
@login_required
def save_cr_form():
    import json
    data = request.get_json()

    po_key = data.get('poKey', '').strip()
    if not po_key:
        return jsonify({'error': 'PO key required'}), 400

    customer = data.get('customer', '').strip()
    bid = data.get('bid', '').strip()
    po = data.get('po', '').strip()
    cr = data.get('cr', '').strip()
    record_no = data.get('recordNo', '').strip()
    record_date = data.get('recordDate', '').strip()
    rows = data.get('rows', [])

    username = session.get('username', 'unknown')
    is_admin = session.get('user_is_admin', False)

    conn = get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            # lock row for update if exists
            cur.execute(
                'SELECT id, amendment_details FROM cr_forms WHERE po_key = %s FOR UPDATE',
                (po_key,)
            )
            form = cur.fetchone()

            if is_admin:
                amendment_details = data.get('amendmentDetails', '').strip()
            else:
                amendment_details = form['amendment_details'] if form else ''

            if form:
                form_id = form['id']
                cur.execute('''
                    UPDATE cr_forms 
                       SET customer = %s, bid = %s, po = %s, cr = %s,
                           record_no = %s, record_date = %s,
                           amendment_details = %s,
                           last_modified_by = %s,
                           last_modified_at = CURRENT_TIMESTAMP
                     WHERE id = %s
                ''', (customer, bid, po, cr, record_no, record_date,
                      amendment_details, username, form_id))

                cur.execute('DELETE FROM cr_form_rows WHERE cr_form_id = %s', (form_id,))
            else:
                cur.execute('''
                    INSERT INTO cr_forms
                      (po_key, customer, bid, po, cr, record_no, record_date,
                       amendment_details, last_modified_by)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    RETURNING id
                ''', (po_key, customer, bid, po, cr, record_no, record_date,
                      amendment_details, username))
                form_id = cur.fetchone()['id']

            for row in rows:
                item_no = row.get('key', '')
                if not item_no:
                    continue
                cycles_json = json.dumps(row.get('cycles', []))
                cur.execute('''
                    INSERT INTO cr_form_rows 
                      (cr_form_id, item_no, part_number, part_description, rev, qty, cycles, remarks)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                ''', (
                    form_id,
                    item_no,
                    row.get('part', ''),
                    row.get('desc', ''),
                    row.get('rev', ''),
                    row.get('qty', ''),
                    cycles_json,
                    row.get('remarks', '')
                ))

        conn.commit()
        conn.close()

        user_id = session.get('user_id')
        handle_form_completion_notification('CR', form_id, data, user_id, username)

        return jsonify({
            'success': True,
            'lastModifiedBy': username,
            'lastModifiedAt': now_ist().isoformat()
        })
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': str(e)}), 500

@app.route('/api/cr-form/load', methods=['GET'])
@login_required
def load_cr_form():
    import json
    po_key = request.args.get('poKey', '').strip()
    if not po_key:
        return jsonify({'error': 'PO key required'}), 400

    conn = get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute('''
                SELECT id, customer, bid, po, cr, record_no, record_date,
                       amendment_details, last_modified_by, last_modified_at
                  FROM cr_forms
                 WHERE po_key = %s
            ''', (po_key,))
            form = cur.fetchone()

            if not form:
                conn.close()
                return jsonify({'exists': False})

            form_id = form['id']

            cur.execute('''
                SELECT item_no, part_number, part_description, rev, qty, cycles, remarks
                  FROM cr_form_rows
                 WHERE cr_form_id = %s
                 ORDER BY id
            ''', (form_id,))
            rows_db = cur.fetchall()

        rows = []
        for row in rows_db:
            cycles = json.loads(row['cycles']) if row['cycles'] else []
            rows.append({
                'key': row['item_no'],
                'part': row['part_number'] or '',
                'desc': row['part_description'] or '',
                'rev': row['rev'] or '',
                'qty': row['qty'] or '',
                'cycles': cycles,
                'remarks': row['remarks'] or ''
            })

        conn.close()

        return jsonify({
            'exists': True,
            'formId': form_id,
            'customer': form['customer'] or '',
            'bid': form['bid'] or '',
            'po': form['po'] or '',
            'cr': form['cr'] or '',
            'recordNo': form['record_no'] or '',
            'recordDate': form['record_date'] or '',
            'amendmentDetails': form['amendment_details'] or '',
            'rows': rows,
            'lastModifiedBy': form['last_modified_by'] or '',
            'lastModifiedAt': form['last_modified_at'] or ''
        })
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 500

@app.route('/api/ped-form/save', methods=['POST'])
@login_required
def save_ped_form():
    import json
    data = request.get_json()

    po_key = data.get('poKey', '').strip()
    if not po_key:
        return jsonify({'error': 'PO key required'}), 400

    customer = data.get('customer', '').strip()
    bid = data.get('bid', '').strip()
    po = data.get('po', '').strip()
    cr = data.get('cr', '').strip()
    record_no = data.get('recordNo', '').strip()
    record_date = data.get('recordDate', '').strip()
    rows = data.get('rows', [])

    username = session.get('username', 'unknown')
    is_admin = session.get('user_is_admin', False)

    conn = get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                'SELECT id, amendment_details FROM ped_forms WHERE po_key = %s FOR UPDATE',
                (po_key,)
            )
            form = cur.fetchone()

            if is_admin:
                amendment_details = data.get('amendmentDetails', '').strip()
            else:
                amendment_details = form['amendment_details'] if form else ''

            if form:
                form_id = form['id']
                cur.execute('''
                    UPDATE ped_forms 
                       SET customer = %s, bid = %s, po = %s, cr = %s,
                           record_no = %s, record_date = %s,
                           amendment_details = %s,
                           last_modified_by = %s,
                           last_modified_at = CURRENT_TIMESTAMP
                     WHERE id = %s
                ''', (customer, bid, po, cr, record_no, record_date,
                      amendment_details, username, form_id))

                cur.execute('DELETE FROM ped_form_rows WHERE ped_form_id = %s', (form_id,))
            else:
                cur.execute('''
                    INSERT INTO ped_forms
                      (po_key, customer, bid, po, cr, record_no, record_date,
                       amendment_details, last_modified_by)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    RETURNING id
                ''', (po_key, customer, bid, po, cr, record_no, record_date,
                      amendment_details, username))
                form_id = cur.fetchone()['id']

            # sync shared item fields from CR
            cr_item_map = get_cr_common_item_map(conn, po_key)

            for row in rows:
                item_no = row.get('key', '')
                if not item_no:
                    continue

                part_number = row.get('part', '')
                part_description = row.get('desc', '')
                rev = row.get('rev', '')
                qty = row.get('qty', '')

                cr_item = cr_item_map.get(str(item_no).strip())
                if cr_item:
                    if not part_number:
                        part_number = cr_item['part_number']
                    if not part_description:
                        part_description = cr_item['part_description']
                    if not rev:
                        rev = cr_item['rev']
                    if not qty:
                        qty = cr_item['qty']

                ped_cycles_json = json.dumps(row.get('pedCycles', []))
                notes_json = json.dumps(row.get('notes', []))

                cur.execute('''
                    INSERT INTO ped_form_rows 
                      (ped_form_id, item_no, part_number, part_description, rev, qty, ped_cycles, notes, remarks)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                ''', (
                    form_id,
                    item_no,
                    part_number,
                    part_description,
                    rev,
                    qty,
                    ped_cycles_json,
                    notes_json,
                    row.get('remarks', '')
                ))

        conn.commit()
        conn.close()

        user_id = session.get('user_id')
        handle_form_completion_notification('PED', form_id, data, user_id, username)

        return jsonify({
            'success': True,
            'lastModifiedBy': username,
            'lastModifiedAt': now_ist().isoformat()
        })
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': str(e)}), 500

@app.route('/api/ped-form/load', methods=['GET'])
@login_required
def load_ped_form():
    import json
    po_key = request.args.get('poKey', '').strip()
    if not po_key:
        return jsonify({'error': 'PO key required'}), 400

    conn = get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute('''
                SELECT id, customer, bid, po, cr, record_no, record_date,
                       amendment_details, last_modified_by, last_modified_at
                  FROM ped_forms
                 WHERE po_key = %s
            ''', (po_key,))
            form = cur.fetchone()

            if not form:
                conn.close()
                return jsonify({'exists': False})

            form_id = form['id']

            cur.execute('''
                SELECT item_no, part_number, part_description, rev, qty,
                       ped_cycles, notes, remarks
                  FROM ped_form_rows
                 WHERE ped_form_id = %s
                 ORDER BY id
            ''', (form_id,))
            rows_db = cur.fetchall()

        rows = []
        for row in rows_db:
            ped_cycles = json.loads(row['ped_cycles']) if row['ped_cycles'] else []
            notes = json.loads(row['notes']) if row['notes'] else []
            rows.append({
                'key': row['item_no'],
                'part': row['part_number'] or '',
                'desc': row['part_description'] or '',
                'rev': row['rev'] or '',
                'qty': row['qty'] or '',
                'pedCycles': ped_cycles,
                'notes': notes,
                'remarks': row['remarks'] or ''
            })

        conn.close()

        return jsonify({
            'exists': True,
            'formId': form_id,
            'customer': form['customer'] or '',
            'bid': form['bid'] or '',
            'po': form['po'] or '',
            'cr': form['cr'] or '',
            'recordNo': form['record_no'] or '',
            'recordDate': form['record_date'] or '',
            'amendmentDetails': form['amendment_details'] or '',
            'rows': rows,
            'lastModifiedBy': form['last_modified_by'] or '',
            'lastModifiedAt': form['last_modified_at'] or ''
        })
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 500

@app.route('/api/lead-form/save', methods=['POST'])
@login_required
def save_lead_form():
    data = request.get_json()

    po_key = data.get('poKey', '').strip()
    if not po_key:
        return jsonify({'error': 'PO key required'}), 400

    customer = data.get('customer', '').strip()
    bid = data.get('bid', '').strip()
    po = data.get('po', '').strip()
    cr = data.get('cr', '').strip()
    record_no = data.get('recordNo', '').strip()
    record_date = data.get('recordDate', '').strip()
    prepared_by = (data.get('preparedBy') or '').strip()   # NEW
    general_remarks = data.get('generalRemarks', '').strip()
    rows = data.get('rows', [])

    username = session.get('username', 'unknown')

    conn = get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                'SELECT id FROM lead_forms WHERE po_key = %s FOR UPDATE',
                (po_key,)
            )
            form = cur.fetchone()

            if form:
                form_id = form['id']
                cur.execute('''
                    UPDATE lead_forms 
                       SET customer = %s, bid = %s, po = %s, cr = %s,
                           record_no = %s, record_date = %s,
                           prepared_by = %s,              -- NEW
                           general_remarks = %s,
                           last_modified_by = %s,
                           last_modified_at = CURRENT_TIMESTAMP
                     WHERE id = %s
                ''', (customer, bid, po, cr, record_no, record_date,
                      prepared_by, general_remarks, username, form_id))

                cur.execute('DELETE FROM lead_form_rows WHERE lead_form_id = %s', (form_id,))
            else:
                cur.execute('''
                    INSERT INTO lead_forms
                      (po_key, customer, bid, po, cr, record_no, record_date,
                       prepared_by,                     -- NEW
                       general_remarks, last_modified_by)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    RETURNING id
                ''', (po_key, customer, bid, po, cr, record_no, record_date,
                      prepared_by, general_remarks, username))
                form_id = cur.fetchone()['id']

            cr_item_map = get_cr_common_item_map(conn, po_key)

            for row in rows:
                item_no = row.get('itemNo', '')
                if not item_no:
                    continue

                part_number = row.get('part', '')
                part_description = row.get('desc', '')
                rev = row.get('rev', '')
                qty = row.get('qty', '')

                cr_item = cr_item_map.get(str(item_no).strip())
                if cr_item:
                    if not part_number:
                        part_number = cr_item['part_number']
                    if not part_description:
                        part_description = cr_item['part_description']
                    if not rev:
                        rev = cr_item['rev']
                    if not qty:
                        qty = cr_item['qty']

                cur.execute('''
                    INSERT INTO lead_form_rows 
                      (lead_form_id, item_no, part_number, part_description, rev, qty, 
                       customer_required_date, standard_lead_time, gtn_agreed_date, remarks)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ''', (
                    form_id,
                    item_no,
                    part_number,
                    part_description,
                    rev,
                    qty,
                    row.get('customerRequiredDate', ''),
                    row.get('standardLeadTime', ''),
                    row.get('gtnAgreedDate', ''),
                    row.get('remarks', '')
                ))

        conn.commit()
        conn.close()

        user_id = session.get('user_id')
        handle_form_completion_notification('LEAD', form_id, data, user_id, username)

        return jsonify({
            'success': True,
            'lastModifiedBy': username,
            'lastModifiedAt': now_ist().isoformat()
        })
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': str(e)}), 500

@app.route('/api/lead-form/load', methods=['GET'])
@login_required
def load_lead_form():
    po_key = request.args.get('poKey', '').strip()
    if not po_key:
        return jsonify({'error': 'PO key required'}), 400

    conn = get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute('''
                SELECT id, customer, bid, po, cr, record_no, record_date,
                       prepared_by,                      -- NEW
                       general_remarks, last_modified_by, last_modified_at
                  FROM lead_forms
                 WHERE po_key = %s
            ''', (po_key,))
            form = cur.fetchone()

            if not form:
                conn.close()
                return jsonify({'exists': False})

            form_id = form['id']

            cur.execute('''
                SELECT item_no, part_number, part_description, rev, qty, 
                       customer_required_date, standard_lead_time, gtn_agreed_date, remarks
                  FROM lead_form_rows
                 WHERE lead_form_id = %s
                 ORDER BY id
            ''', (form_id,))
            rows_db = cur.fetchall()

        rows = []
        for row in rows_db:
            rows.append({
                'itemNo': row['item_no'],
                'part': row['part_number'] or '',
                'desc': row['part_description'] or '',
                'rev': row['rev'] or '',
                'qty': row['qty'] or '',
                'customerRequiredDate': row['customer_required_date'] or '',
                'standardLeadTime': row['standard_lead_time'] or '',
                'gtnAgreedDate': row['gtn_agreed_date'] or '',
                'remarks': row['remarks'] or ''
            })

        conn.close()

        return jsonify({
            'exists': True,
            'formId': form_id,
            'customer': form['customer'] or '',
            'bid': form['bid'] or '',
            'po': form['po'] or '',
            'cr': form['cr'] or '',
            'recordNo': form['record_no'] or '',
            'recordDate': form['record_date'] or '',
            'preparedBy': form.get('prepared_by') or '',     # NEW
            'generalRemarks': form['general_remarks'] or '',
            'rows': rows,
            'lastModifiedBy': form['last_modified_by'] or '',
            'lastModifiedAt': form['last_modified_at'] or ''
        })
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 500

# ========== EXPORT & COMMENTS & NOTIFICATIONS (CR / PED / LEAD) ==========
# (Your long export_cr_to_excel / export_ped_to_excel / export_lead_to_excel
#  code from your last message has already been ported to psycopg2 style there;
#  I’m leaving it unchanged here for brevity since you pasted the working version.)

# ------------------ CR/PED/LEAD export & comments & notifications ------------------

# ------------------ CR/PED/LEAD export & comments & notifications ------------------

def build_cr_comments_excel(conn, form, form_id):
    import openpyxl
    from io import BytesIO

    template_path = "attached_assets/Comments.xlsx"
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    ws.title = "CR COMMENTS"

    # ---- Header ----
    ws['AB1'] = form['record_no'] or ''
    ws['AB2'] = form['record_date'] or ''
    ws['C3']  = form['customer'] or ''
    ws['F3']  = form['bid'] or ''
    ws['S3']  = form['po'] or ''
    ws['AC3'] = form['cr'] or ''

    # ---- Amendment Details (NEW) ----
    ws['C26'] = form['amendment_details'] or ''

    # ---- Comments ----
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute("""
            SELECT department, username, comment_text
            FROM cr_comments
            WHERE cr_form_id = %s
            ORDER BY created_at
        """, (form_id,))
        comments = cur.fetchall()

    start_row = 5
    for idx, c in enumerate(comments, start=1):
        r = start_row + idx - 1
        ws[f"A{r}"] = idx
        ws[f"B{r}"] = c['department']
        ws[f"C{r}"] = "CR"
        ws[f"D{r}"] = c['username']
        ws[f"E{r}"] = c['comment_text']

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

@app.route('/api/cr-export-excel', methods=['GET'])
@login_required
def export_cr_to_excel():
    import json
    import openpyxl
    from openpyxl.drawing.image import Image
    from openpyxl.styles import Alignment
    from io import BytesIO
    import zipfile
    from flask import make_response, jsonify, request
    import os

    # --------------------------------------------------
    # Helpers
    # --------------------------------------------------
    def build_merged_cell_map(ws):
        merged_map = {}
        for merged_range in ws.merged_cells.ranges:
            min_row, min_col = merged_range.min_row, merged_range.min_col
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    merged_map[(row, col)] = (min_row, min_col)
        return merged_map

    def write_cell(ws, row, col, value, merged_map):
        if (row, col) in merged_map:
            ar, ac = merged_map[(row, col)]
            ws.cell(row=ar, column=ac).value = value
        else:
            ws.cell(row=row, column=col).value = value

    # --------------------------------------------------
    # Engineering Signature (JOIN master_signatures)
    # --------------------------------------------------
    def get_engineering_signature(conn, po_key):
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("""
                SELECT
                    cds.signed_by,
                    cds.signed_at,
                    ms.signature_path
                FROM cr_department_signatures cds
                JOIN master_signatures ms
                  ON ms.username = cds.signed_by
                 AND ms.department = cds.department
                 AND ms.is_active = TRUE
                WHERE cds.po_key = %s
                  AND cds.department = 'engineering'
            """, (po_key,))
            return cur.fetchone()

    # --------------------------------------------------
    # Header Mapping
    # --------------------------------------------------
    cr_header_cells = {
        'CR_1': {
            'record_no_col': 26,
            'record_date_col': 26,
            'customer_row_col': (3, 3),
            'bid_row_col': (3, 5),
            'po_row_col': (3, 15),
            'cr_row_col': (3, 25),
            'amendment_row_col': (41, 3),
        },
        'CR_2': {
            'record_no_col': 28,
            'record_date_col': 28,
            'customer_row_col': (3, 3),
            'bid_row_col': (3, 5),
            'po_row_col': (3, 18),
            'cr_row_col': (3, 28),
            'amendment_row_col': (41, 3),
        },
        'CR_3': {
            'record_no_col': 26,
            'record_date_col': 26,
            'customer_row_col': (3, 3),
            'bid_row_col': (3, 5),
            'po_row_col': (3, 15),
            'cr_row_col': (3, 25),
            'amendment_row_col': (41, 3),
        },
    }

    templates = {
        'CR_1': 'attached_assets/CR_1762338481711.xlsx',
        'CR_2': 'attached_assets/CR 2_1762338481710.xlsx',
        'CR_3': 'attached_assets/CR 3_1762338481711.xlsx'
    }

    logo_path = 'attached_assets/GTN_LOGO_1762400078631.png'

    cycle_mapping = {
        'CR_1': (0, 21),
        'CR_2': (21, 44),
        'CR_3': (44, 64),
    }

    for p in templates.values():
        if not os.path.exists(p):
            return jsonify({'error': f'Template missing: {p}'}), 404
    if not os.path.exists(logo_path):
        return jsonify({'error': 'Logo missing'}), 404

    # --------------------------------------------------
    # PO Filter
    # --------------------------------------------------
    po_ids_param = (request.args.get('po_ids') or '').strip()
    po_ids = [int(x) for x in po_ids_param.split(',') if x.isdigit()] if po_ids_param else []

    conn = get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            if po_ids:
                placeholders = ",".join(["%s"] * len(po_ids))
                cur.execute(
                    f"SELECT customer, bid, po, cr FROM pos WHERE id IN ({placeholders})",
                    po_ids
                )
                pos_rows = cur.fetchall()
                po_keys = [f"{p['customer']}|{p['bid']}|{p['po']}|{p['cr']}" for p in pos_rows]

                cur.execute(f"""
                    SELECT id, customer, bid, po, cr,
                           record_no, record_date, amendment_details
                    FROM cr_forms
                    WHERE po_key IN ({",".join(["%s"] * len(po_keys))})
                    ORDER BY id
                """, po_keys)
                forms = cur.fetchall()
            else:
                cur.execute("""
                    SELECT id, customer, bid, po, cr,
                           record_no, record_date, amendment_details
                    FROM cr_forms
                    ORDER BY id
                """)
                forms = cur.fetchall()

        if not forms:
            return jsonify({'error': 'No CR forms found'}), 404

        zip_buffer = BytesIO()

        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                for idx, form in enumerate(forms):
                    form_id = form['id']
                    po_key = f"{form['customer']}|{form['bid']}|{form['po']}|{form['cr']}"
                    safe_customer = ''.join(
                        c for c in (form['customer'] or 'Customer')
                        if c.isalnum() or c in (' ', '_', '-')
                    )[:30]

                    # ==================================================
                    # 1) CR EXCEL FILES (3 templates)
                    # ==================================================
                    for template_key, template_path in templates.items():
                        wb = openpyxl.load_workbook(template_path)
                        ws = wb.active
                        ws.title = "CR"
                        merged_map = build_merged_cell_map(ws)

                        # Logo
                        try:
                            logo = Image(logo_path)
                            logo.width = 80
                            logo.height = 60
                            ws.add_image(logo, 'A1')
                        except Exception:
                            pass

                        cells = cr_header_cells[template_key]
                        write_cell(ws, 1, cells['record_no_col'], form['record_no'], merged_map)
                        write_cell(ws, 2, cells['record_date_col'], form['record_date'], merged_map)
                        write_cell(ws, *cells['customer_row_col'], form['customer'], merged_map)
                        write_cell(ws, *cells['bid_row_col'], form['bid'], merged_map)
                        write_cell(ws, *cells['po_row_col'], form['po'], merged_map)
                        write_cell(ws, *cells['cr_row_col'], form['cr'], merged_map)
                        write_cell(ws, *cells['amendment_row_col'], form['amendment_details'], merged_map)

                        cur.execute("""
                            SELECT item_no, part_number, part_description,
                                   rev, qty, cycles, remarks
                            FROM cr_form_rows
                            WHERE cr_form_id = %s
                            ORDER BY id
                        """, (form_id,))
                        rows = cur.fetchall()

                        cycle_start, cycle_end = cycle_mapping[template_key]
                        excel_row = 8

                        for r in rows:
                            if excel_row > 12:
                                break

                            write_cell(ws, excel_row, 1, r['item_no'], merged_map)
                            write_cell(ws, excel_row, 2, r['part_number'], merged_map)
                            write_cell(ws, excel_row, 3, r['part_description'], merged_map)
                            write_cell(ws, excel_row, 4, r['rev'], merged_map)
                            write_cell(ws, excel_row, 5, r['qty'], merged_map)

                            cycles = json.loads(r['cycles']) if r['cycles'] else []
                            for i, v in enumerate(cycles[cycle_start:cycle_end]):
                                write_cell(ws, excel_row, 6 + i, v, merged_map)

                            excel_row += 1

                        # Engineering signature → CR_1 → J38:N38
                        if template_key == 'CR_1':
                            sign = get_engineering_signature(conn, po_key)
                            if sign and sign.get('signature_path'):
                                try:
                                    sig_rel_path = sign["signature_path"].lstrip("/")
                                    sig_abs_path = os.path.join(app.root_path, sig_rel_path)

                                    if os.path.exists(sig_abs_path):
                                        tmp_path = None
                                        try:
                                            # TRY: use Pillow to auto-crop whitespace/alpha, scale to ~70% of span width, then insert
                                            from PIL import Image as PILImage, ImageChops
                                            import tempfile

                                            pil = PILImage.open(sig_abs_path)

                                            # normalize transparency to RGBA
                                            if pil.mode in ("RGBA", "LA") or (
                                                    pil.mode == "P" and "transparency" in pil.info):
                                                bg = PILImage.new("RGBA", pil.size, (255, 255, 255, 0))
                                                bg.paste(pil, mask=pil.convert("RGBA").split()[-1])
                                                pil = bg
                                            pil = pil.convert("RGBA")

                                            # trim transparent / blank margins
                                            try:
                                                bbox = ImageChops.invert(pil.convert("L")).getbbox()
                                                if bbox:
                                                    pil = pil.crop(bbox)
                                            except Exception:
                                                pass

                                            # target span: columns J..N (10..14), row 38
                                            start_col = 10  # J
                                            end_col = 14  # N
                                            start_row = 38
                                            end_row = 38

                                            # helpers to estimate pixel sizes
                                            from openpyxl.utils import get_column_letter

                                            def col_width_to_pixels(ws, col_idx):
                                                col_letter = get_column_letter(col_idx)
                                                w = ws.column_dimensions.get(col_letter).width
                                                if w is None:
                                                    w = 8.43
                                                if w < 1:
                                                    return int(w * 12)
                                                return int((w - 0.75) * 7 + 5)

                                            def row_height_to_pixels(ws, row_idx):
                                                h = ws.row_dimensions.get(row_idx).height
                                                if h is None:
                                                    h = 15
                                                return int(h * 96.0 / 72.0)

                                            # compute span width/height (J..N)
                                            span_width_px = 0
                                            for c in range(start_col, end_col + 1):
                                                span_width_px += col_width_to_pixels(ws, c)
                                            span_height_px = 0
                                            for r in range(start_row, end_row + 1):
                                                span_height_px += row_height_to_pixels(ws, r)
                                            if span_height_px == 0:
                                                span_height_px = row_height_to_pixels(ws, start_row)

                                            # scale to ~70% of span width, preserving aspect ratio
                                            target_w = max(1, int(span_width_px * 0.7))
                                            orig_w, orig_h = pil.size
                                            if orig_w > 0 and target_w < orig_w:
                                                target_h = int((target_w / orig_w) * orig_h)
                                                pil = pil.resize((target_w, target_h), PILImage.LANCZOS)

                                            # save processed image to temp file for openpyxl
                                            tf = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
                                            tmp_path = tf.name
                                            pil.save(tmp_path, format="PNG")
                                            tf.close()

                                            # create openpyxl Image and set explicit width/height (px)
                                            img = Image(tmp_path)
                                            try:
                                                img.width, img.height = pil.size
                                            except Exception:
                                                pass

                                            # insert anchored at J38 then adjust anchor to span J..N and center
                                            ws.add_image(img, "J38")

                                            try:
                                                from openpyxl.utils.units import pixels_to_EMU

                                                anch = getattr(img, "anchor", None)
                                                if anch is not None and hasattr(anch, "_from"):
                                                    fr = anch._from
                                                    to = anch._to
                                                    # set anchor span (zero-based)
                                                    fr.col = start_col - 1
                                                    fr.row = start_row - 1
                                                    to.col = end_col
                                                    to.row = end_row

                                                    left_px = max(
                                                        int((span_width_px - (img.width or pil.size[0])) / 2.0), 0)
                                                    top_px = max(
                                                        int((span_height_px - (img.height or pil.size[1])) / 2.0), 0)
                                                    try:
                                                        fr.colOff = pixels_to_EMU(left_px)
                                                        fr.rowOff = pixels_to_EMU(top_px)
                                                    except Exception:
                                                        # older/newer openpyxl may behave differently; ignore if fails
                                                        pass
                                            except Exception:
                                                # ignore EMU/anchor adjustment failures and keep default placement
                                                pass

                                        except Exception:
                                            # Pillow missing or processing failed -> fallback to simple insertion
                                            try:
                                                img = Image(sig_abs_path)
                                                img.width = 120
                                                img.height = 50
                                                ws.add_image(img, "J38")
                                                # best-effort: try to center using same span math (if anchor available)
                                                try:
                                                    from openpyxl.utils import get_column_letter
                                                    from openpyxl.utils.units import pixels_to_EMU

                                                    def col_width_to_pixels(ws, col_idx):
                                                        col_letter = get_column_letter(col_idx)
                                                        w = ws.column_dimensions.get(col_letter).width
                                                        if w is None:
                                                            w = 8.43
                                                        if w < 1:
                                                            return int(w * 12)
                                                        return int((w - 0.75) * 7 + 5)

                                                    def row_height_to_pixels(ws, row_idx):
                                                        h = ws.row_dimensions.get(row_idx).height
                                                        if h is None:
                                                            h = 15
                                                        return int(h * 96.0 / 72.0)

                                                    span_width_px = sum(
                                                        col_width_to_pixels(ws, c) for c in range(10, 15))  # J..N
                                                    span_height_px = row_height_to_pixels(ws, 38)
                                                    anch = getattr(img, "anchor", None)
                                                    if anch is not None and hasattr(anch, "_from"):
                                                        fr = anch._from
                                                        to = anch._to
                                                        fr.col = 10 - 1  # J
                                                        fr.row = 38 - 1
                                                        to.col = 14  # N
                                                        to.row = 38
                                                        left_px = max(int((span_width_px - img.width) / 2.0), 0)
                                                        top_px = max(int((span_height_px - img.height) / 2.0), 0)
                                                        try:
                                                            fr.colOff = pixels_to_EMU(left_px)
                                                            fr.rowOff = pixels_to_EMU(top_px)
                                                        except Exception:
                                                            pass
                                                except Exception:
                                                    pass
                                            except Exception:
                                                pass
                                        finally:
                                            # cleanup temp file if created
                                            try:
                                                if tmp_path:
                                                    import os
                                                    os.unlink(tmp_path)
                                            except Exception:
                                                pass

                                except Exception as e:
                                    print("Signature image error:", e)

                            # Centre the signed_by | date text across J39..N39
                            try:
                                signed_at = sign.get('signed_at') if sign else None
                                date_str = ''
                                if signed_at:
                                    try:
                                        date_str = signed_at.strftime('%d-%m-%Y')
                                    except Exception:
                                        date_str = str(signed_at)
                                signed_by = sign.get('signed_by') if sign else ''
                                try:
                                    ws.merge_cells(start_row=39, start_column=10, end_row=39, end_column=14)  # J39:N39
                                except Exception:
                                    pass
                                ws.cell(row=39, column=10).value = f"{signed_by} | {date_str}"
                                ws.cell(row=39, column=10).alignment = Alignment(
                                    wrap_text=True,
                                    horizontal='center',
                                    vertical='center'
                                )
                            except Exception:
                                pass

                        # Manufacturing signature → CR_1 → S38
                        # --- Manufacturing signature → CR_1 → S38..Z38 (with date in S39..Z39) ---
                        # Insert this block in export_cr_to_excel right after the engineering signature block.
                        if template_key == 'CR_1':
                            # fetch manufacturing signature for this PO
                            try:
                                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as _cur:
                                    _cur.execute("""
                                        SELECT
                                            cds.signed_by,
                                            cds.signed_at,
                                            ms.signature_path
                                        FROM cr_department_signatures cds
                                        JOIN master_signatures ms
                                          ON ms.username = cds.signed_by
                                         AND ms.department = cds.department
                                         AND ms.is_active = TRUE
                                        WHERE cds.po_key = %s
                                          AND cds.department = %s
                                        """, (po_key, 'manufacturing'))
                                    msign = _cur.fetchone()
                            except Exception:
                                msign = None

                            if msign and msign.get('signature_path'):
                                try:
                                    sig_rel_path = msign["signature_path"].lstrip("/")
                                    sig_abs_path = os.path.join(app.root_path, sig_rel_path)

                                    if os.path.exists(sig_abs_path):
                                        tmp_path = None
                                        try:
                                            # Pillow-based processing (auto-crop + scale)
                                            from PIL import Image as PILImage, ImageChops
                                            import tempfile

                                            pil = PILImage.open(sig_abs_path)

                                            if pil.mode in ("RGBA", "LA") or (
                                                    pil.mode == "P" and "transparency" in pil.info):
                                                bg = PILImage.new("RGBA", pil.size, (255, 255, 255, 0))
                                                bg.paste(pil, mask=pil.convert("RGBA").split()[-1])
                                                pil = bg
                                            pil = pil.convert("RGBA")

                                            try:
                                                bbox = ImageChops.invert(pil.convert("L")).getbbox()
                                                if bbox:
                                                    pil = pil.crop(bbox)
                                            except Exception:
                                                pass

                                            # target span: columns S..Z (19..26), row 38
                                            start_col = 19  # S
                                            end_col = 26  # Z
                                            start_row = 38
                                            end_row = 38

                                            from openpyxl.utils import get_column_letter
                                            def col_width_to_pixels(ws, col_idx):
                                                col_letter = get_column_letter(col_idx)
                                                w = ws.column_dimensions.get(col_letter).width
                                                if w is None:
                                                    w = 8.43
                                                if w < 1:
                                                    return int(w * 12)
                                                return int((w - 0.75) * 7 + 5)

                                            def row_height_to_pixels(ws, row_idx):
                                                h = ws.row_dimensions.get(row_idx).height
                                                if h is None:
                                                    h = 15
                                                return int(h * 96.0 / 72.0)

                                            span_width_px = sum(
                                                col_width_to_pixels(ws, c) for c in range(start_col, end_col + 1))
                                            span_height_px = sum(
                                                row_height_to_pixels(ws, r) for r in range(start_row, end_row + 1))
                                            if span_height_px == 0:
                                                span_height_px = row_height_to_pixels(ws, start_row)

                                            # scale to ~70% of span width
                                            target_w = max(1, int(span_width_px * 0.7))
                                            orig_w, orig_h = pil.size
                                            if orig_w > 0 and target_w < orig_w:
                                                target_h = int((target_w / orig_w) * orig_h)
                                                pil = pil.resize((target_w, target_h), PILImage.LANCZOS)

                                            tf = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
                                            tmp_path = tf.name
                                            pil.save(tmp_path, format="PNG")
                                            tf.close()

                                            img = Image(tmp_path)
                                            try:
                                                img.width, img.height = pil.size
                                            except Exception:
                                                pass

                                            ws.add_image(img, "S38")

                                            try:
                                                from openpyxl.utils.units import pixels_to_EMU
                                                anch = getattr(img, "anchor", None)
                                                if anch is not None and hasattr(anch, "_from"):
                                                    fr = anch._from
                                                    to = anch._to
                                                    fr.col = start_col - 1
                                                    fr.row = start_row - 1
                                                    to.col = end_col
                                                    to.row = end_row

                                                    left_px = max(
                                                        int((span_width_px - (img.width or pil.size[0])) / 2.0), 0)
                                                    top_px = max(
                                                        int((span_height_px - (img.height or pil.size[1])) / 2.0), 0)
                                                    try:
                                                        fr.colOff = pixels_to_EMU(left_px)
                                                        fr.rowOff = pixels_to_EMU(top_px)
                                                    except Exception:
                                                        pass
                                            except Exception:
                                                pass

                                        except Exception:
                                            # Pillow fallback: simple insert + best-effort centering
                                            try:
                                                img = Image(sig_abs_path)
                                                img.width = 120
                                                img.height = 50
                                                ws.add_image(img, "S38")
                                                try:
                                                    from openpyxl.utils import get_column_letter
                                                    from openpyxl.utils.units import pixels_to_EMU
                                                    def col_width_to_pixels(ws, col_idx):
                                                        col_letter = get_column_letter(col_idx)
                                                        w = ws.column_dimensions.get(col_letter).width
                                                        if w is None:
                                                            w = 8.43
                                                        if w < 1:
                                                            return int(w * 12)
                                                        return int((w - 0.75) * 7 + 5)

                                                    def row_height_to_pixels(ws, row_idx):
                                                        h = ws.row_dimensions.get(row_idx).height
                                                        if h is None:
                                                            h = 15
                                                        return int(h * 96.0 / 72.0)

                                                    span_width_px = sum(
                                                        col_width_to_pixels(ws, c) for c in range(19, 27))
                                                    span_height_px = row_height_to_pixels(ws, 38)
                                                    anch = getattr(img, "anchor", None)
                                                    if anch is not None and hasattr(anch, "_from"):
                                                        fr = anch._from
                                                        to = anch._to
                                                        fr.col = 19 - 1
                                                        fr.row = 38 - 1
                                                        to.col = 26
                                                        to.row = 38
                                                        left_px = max(int((span_width_px - img.width) / 2.0), 0)
                                                        top_px = max(int((span_height_px - img.height) / 2.0), 0)
                                                        try:
                                                            fr.colOff = pixels_to_EMU(left_px)
                                                            fr.rowOff = pixels_to_EMU(top_px)
                                                        except Exception:
                                                            pass
                                                except Exception:
                                                    pass
                                            except Exception:
                                                pass
                                        finally:
                                            try:
                                                if tmp_path:
                                                    import os
                                                    os.unlink(tmp_path)
                                            except Exception:
                                                pass

                                except Exception as e:
                                    print("Manufacturing signature image error:", e)

                            # Centre the signed_by | date text across S39..Z39
                            try:
                                if msign:
                                    signed_at = msign.get('signed_at')
                                    signed_by = msign.get('signed_by', '')
                                else:
                                    signed_at = None
                                    signed_by = ''
                                date_str = ''
                                if signed_at:
                                    try:
                                        date_str = signed_at.strftime('%d-%m-%Y')
                                    except Exception:
                                        date_str = str(signed_at)
                                try:
                                    ws.merge_cells(start_row=39, start_column=19, end_row=39, end_column=26)  # S39:Z39
                                except Exception:
                                    pass
                                ws.cell(row=39, column=19).value = f"{signed_by} | {date_str}"
                                ws.cell(row=39, column=19).alignment = Alignment(wrap_text=True, horizontal='center',
                                                                                 vertical='center')
                            except Exception:
                                pass

                        #materials/planning signature for this PO
                        if template_key == 'CR_2':
                            # fetch materials/planning signature for this PO
                            try:
                                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as _cur:
                                    _cur.execute("""
                                        SELECT
                                            cds.signed_by,
                                            cds.signed_at,
                                            ms.signature_path
                                        FROM cr_department_signatures cds
                                        JOIN master_signatures ms
                                          ON ms.username = cds.signed_by
                                         AND ms.department = cds.department
                                         AND ms.is_active = TRUE
                                        WHERE cds.po_key = %s
                                          AND cds.department = %s
                                        """, (po_key, 'materials'))
                                    mat_sign = _cur.fetchone()
                            except Exception:
                                mat_sign = None

                            if mat_sign and mat_sign.get('signature_path'):
                                try:
                                    sig_rel_path = mat_sign["signature_path"].lstrip("/")
                                    sig_abs_path = os.path.join(app.root_path, sig_rel_path)

                                    if os.path.exists(sig_abs_path):
                                        tmp_path = None
                                        try:
                                            # Pillow-based processing: autocrop, scale to ~70% of span, then insert
                                            from PIL import Image as PILImage, ImageChops
                                            import tempfile

                                            pil = PILImage.open(sig_abs_path)

                                            # normalize transparency
                                            if pil.mode in ("RGBA", "LA") or (
                                                    pil.mode == "P" and "transparency" in pil.info):
                                                bg = PILImage.new("RGBA", pil.size, (255, 255, 255, 0))
                                                bg.paste(pil, mask=pil.convert("RGBA").split()[-1])
                                                pil = bg
                                            pil = pil.convert("RGBA")

                                            # trim transparent/blank margins
                                            try:
                                                bbox = ImageChops.invert(pil.convert("L")).getbbox()
                                                if bbox:
                                                    pil = pil.crop(bbox)
                                            except Exception:
                                                pass

                                            # target span: columns F..M (6..13), row 38
                                            start_col = 6  # F
                                            end_col = 13  # M
                                            start_row = 38
                                            end_row = 38

                                            from openpyxl.utils import get_column_letter
                                            def col_width_to_pixels(ws, col_idx):
                                                col_letter = get_column_letter(col_idx)
                                                w = ws.column_dimensions.get(col_letter).width
                                                if w is None:
                                                    w = 8.43
                                                if w < 1:
                                                    return int(w * 12)
                                                return int((w - 0.75) * 7 + 5)

                                            def row_height_to_pixels(ws, row_idx):
                                                h = ws.row_dimensions.get(row_idx).height
                                                if h is None:
                                                    h = 15
                                                return int(h * 96.0 / 72.0)

                                            # compute span pixel dimensions
                                            span_width_px = sum(
                                                col_width_to_pixels(ws, c) for c in range(start_col, end_col + 1))
                                            span_height_px = sum(
                                                row_height_to_pixels(ws, r) for r in range(start_row, end_row + 1))
                                            if span_height_px == 0:
                                                span_height_px = row_height_to_pixels(ws, start_row)

                                            # resize image to ~70% of span width while preserving aspect ratio
                                            target_w = max(1, int(span_width_px * 0.7))
                                            orig_w, orig_h = pil.size
                                            if orig_w > 0 and target_w < orig_w:
                                                target_h = int((target_w / orig_w) * orig_h)
                                                pil = pil.resize((target_w, target_h), PILImage.LANCZOS)

                                            # save processed image to temp file
                                            tf = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
                                            tmp_path = tf.name
                                            pil.save(tmp_path, format="PNG")
                                            tf.close()

                                            img = Image(tmp_path)
                                            try:
                                                img.width, img.height = pil.size
                                            except Exception:
                                                pass

                                            # insert and centre across the span
                                            ws.add_image(img, "F38")
                                            try:
                                                from openpyxl.utils.units import pixels_to_EMU
                                                anch = getattr(img, "anchor", None)
                                                if anch is not None and hasattr(anch, "_from"):
                                                    fr = anch._from
                                                    to = anch._to
                                                    fr.col = start_col - 1
                                                    fr.row = start_row - 1
                                                    to.col = end_col
                                                    to.row = end_row

                                                    left_px = max(
                                                        int((span_width_px - (img.width or pil.size[0])) / 2.0), 0)
                                                    top_px = max(
                                                        int((span_height_px - (img.height or pil.size[1])) / 2.0), 0)
                                                    try:
                                                        fr.colOff = pixels_to_EMU(left_px)
                                                        fr.rowOff = pixels_to_EMU(top_px)
                                                    except Exception:
                                                        # ignore EMU assignment issues
                                                        pass
                                            except Exception:
                                                pass

                                        except Exception:
                                            # fallback if Pillow not available or processing fails
                                            try:
                                                img = Image(sig_abs_path)
                                                img.width = 120
                                                img.height = 50
                                                ws.add_image(img, "F38")
                                                # best-effort center
                                                try:
                                                    from openpyxl.utils import get_column_letter
                                                    from openpyxl.utils.units import pixels_to_EMU
                                                    def col_width_to_pixels(ws, col_idx):
                                                        col_letter = get_column_letter(col_idx)
                                                        w = ws.column_dimensions.get(col_letter).width
                                                        if w is None:
                                                            w = 8.43
                                                        if w < 1:
                                                            return int(w * 12)
                                                        return int((w - 0.75) * 7 + 5)

                                                    def row_height_to_pixels(ws, row_idx):
                                                        h = ws.row_dimensions.get(row_idx).height
                                                        if h is None:
                                                            h = 15
                                                        return int(h * 96.0 / 72.0)

                                                    span_width_px = sum(
                                                        col_width_to_pixels(ws, c) for c in range(6, 14))
                                                    span_height_px = row_height_to_pixels(ws, 38)
                                                    anch = getattr(img, "anchor", None)
                                                    if anch is not None and hasattr(anch, "_from"):
                                                        fr = anch._from
                                                        to = anch._to
                                                        fr.col = 6 - 1
                                                        fr.row = 38 - 1
                                                        to.col = 13
                                                        to.row = 38
                                                        left_px = max(int((span_width_px - img.width) / 2.0), 0)
                                                        top_px = max(int((span_height_px - img.height) / 2.0), 0)
                                                        try:
                                                            fr.colOff = pixels_to_EMU(left_px)
                                                            fr.rowOff = pixels_to_EMU(top_px)
                                                        except Exception:
                                                            pass
                                                except Exception:
                                                    pass
                                            except Exception:
                                                pass
                                        finally:
                                            # cleanup temp file if created
                                            try:
                                                if tmp_path:
                                                    import os
                                                    os.unlink(tmp_path)
                                            except Exception:
                                                pass

                                except Exception as e:
                                    print("Materials/Planning signature image error:", e)

                            # Write date + time into F39 (centered)
                            try:
                                if mat_sign:
                                    signed_at = mat_sign.get('signed_at')
                                    signed_by = mat_sign.get('signed_by', '')
                                else:
                                    signed_at = None
                                    signed_by = ''
                                dt_str = ''
                                if signed_at:
                                    try:
                                        # signed_at may be datetime; include time
                                        dt_str = signed_at.strftime('%d-%m-%Y %H:%M')
                                    except Exception:
                                        dt_str = str(signed_at)
                                ws.cell(row=39, column=6).value = f"{signed_by} | {dt_str}"
                                ws.cell(row=39, column=6).alignment = Alignment(wrap_text=True, horizontal='center',
                                                                                vertical='center')
                            except Exception:
                                pass

            #purchase signature
                        if template_key == 'CR_2':
                            # fetch purchase signature for this PO
                            try:
                                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as _cur:
                                    _cur.execute("""
                                        SELECT
                                            cds.signed_by,
                                            cds.signed_at,
                                            ms.signature_path
                                        FROM cr_department_signatures cds
                                        JOIN master_signatures ms
                                          ON ms.username = cds.signed_by
                                         AND ms.department = cds.department
                                         AND ms.is_active = TRUE
                                        WHERE cds.po_key = %s
                                          AND cds.department = %s
                                        """, (po_key, 'purchase'))
                                    purch_sign = _cur.fetchone()
                            except Exception:
                                purch_sign = None

                            if purch_sign and purch_sign.get('signature_path'):
                                try:
                                    sig_rel_path = purch_sign["signature_path"].lstrip("/")
                                    sig_abs_path = os.path.join(app.root_path, sig_rel_path)

                                    if os.path.exists(sig_abs_path):
                                        tmp_path = None
                                        try:
                                            # Pillow-based processing: autocrop, scale to ~70% of span, then insert
                                            from PIL import Image as PILImage, ImageChops
                                            import tempfile

                                            pil = PILImage.open(sig_abs_path)

                                            # normalize transparency
                                            if pil.mode in ("RGBA", "LA") or (
                                                    pil.mode == "P" and "transparency" in pil.info):
                                                bg = PILImage.new("RGBA", pil.size, (255, 255, 255, 0))
                                                bg.paste(pil, mask=pil.convert("RGBA").split()[-1])
                                                pil = bg
                                            pil = pil.convert("RGBA")

                                            # trim transparent/blank margins
                                            try:
                                                bbox = ImageChops.invert(pil.convert("L")).getbbox()
                                                if bbox:
                                                    pil = pil.crop(bbox)
                                            except Exception:
                                                pass

                                            # target span: columns N..R (14..18), row 38
                                            start_col = 14  # N
                                            end_col = 18  # R
                                            start_row = 38
                                            end_row = 38

                                            from openpyxl.utils import get_column_letter
                                            def col_width_to_pixels(ws, col_idx):
                                                col_letter = get_column_letter(col_idx)
                                                w = ws.column_dimensions.get(col_letter).width
                                                if w is None:
                                                    w = 8.43
                                                if w < 1:
                                                    return int(w * 12)
                                                return int((w - 0.75) * 7 + 5)

                                            def row_height_to_pixels(ws, row_idx):
                                                h = ws.row_dimensions.get(row_idx).height
                                                if h is None:
                                                    h = 15
                                                return int(h * 96.0 / 72.0)

                                            span_width_px = sum(
                                                col_width_to_pixels(ws, c) for c in range(start_col, end_col + 1))
                                            span_height_px = sum(
                                                row_height_to_pixels(ws, r) for r in range(start_row, end_row + 1))
                                            if span_height_px == 0:
                                                span_height_px = row_height_to_pixels(ws, start_row)

                                            # scale to ~70% of span width
                                            target_w = max(1, int(span_width_px * 0.7))
                                            orig_w, orig_h = pil.size
                                            if orig_w > 0 and target_w < orig_w:
                                                target_h = int((target_w / orig_w) * orig_h)
                                                pil = pil.resize((target_w, target_h), PILImage.LANCZOS)

                                            tf = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
                                            tmp_path = tf.name
                                            pil.save(tmp_path, format="PNG")
                                            tf.close()

                                            img = Image(tmp_path)
                                            try:
                                                img.width, img.height = pil.size
                                            except Exception:
                                                pass

                                            # insert anchored at N38 then adjust anchor to span N..R and center
                                            ws.add_image(img, "N38")

                                            try:
                                                from openpyxl.utils.units import pixels_to_EMU
                                                anch = getattr(img, "anchor", None)
                                                if anch is not None and hasattr(anch, "_from"):
                                                    fr = anch._from
                                                    to = anch._to
                                                    # set anchor span (zero-based)
                                                    fr.col = start_col - 1
                                                    fr.row = start_row - 1
                                                    to.col = end_col
                                                    to.row = end_row

                                                    left_px = max(
                                                        int((span_width_px - (img.width or pil.size[0])) / 2.0), 0)
                                                    top_px = max(
                                                        int((span_height_px - (img.height or pil.size[1])) / 2.0), 0)
                                                    try:
                                                        fr.colOff = pixels_to_EMU(left_px)
                                                        fr.rowOff = pixels_to_EMU(top_px)
                                                    except Exception:
                                                        # ignore EMU assignment issues
                                                        pass
                                            except Exception:
                                                pass

                                        except Exception:
                                            # fallback if Pillow not available or processing fails
                                            try:
                                                img = Image(sig_abs_path)
                                                img.width = 120
                                                img.height = 50
                                                ws.add_image(img, "N38")
                                                # best-effort centering
                                                try:
                                                    from openpyxl.utils import get_column_letter
                                                    from openpyxl.utils.units import pixels_to_EMU
                                                    def col_width_to_pixels(ws, col_idx):
                                                        col_letter = get_column_letter(col_idx)
                                                        w = ws.column_dimensions.get(col_letter).width
                                                        if w is None:
                                                            w = 8.43
                                                        if w < 1:
                                                            return int(w * 12)
                                                        return int((w - 0.75) * 7 + 5)

                                                    def row_height_to_pixels(ws, row_idx):
                                                        h = ws.row_dimensions.get(row_idx).height
                                                        if h is None:
                                                            h = 15
                                                        return int(h * 96.0 / 72.0)

                                                    span_width_px = sum(
                                                        col_width_to_pixels(ws, c) for c in range(14, 19))
                                                    span_height_px = row_height_to_pixels(ws, 38)
                                                    anch = getattr(img, "anchor", None)
                                                    if anch is not None and hasattr(anch, "_from"):
                                                        fr = anch._from
                                                        to = anch._to
                                                        fr.col = 14 - 1
                                                        fr.row = 38 - 1
                                                        to.col = 18
                                                        to.row = 38
                                                        left_px = max(int((span_width_px - img.width) / 2.0), 0)
                                                        top_px = max(int((span_height_px - img.height) / 2.0), 0)
                                                        try:
                                                            fr.colOff = pixels_to_EMU(left_px)
                                                            fr.rowOff = pixels_to_EMU(top_px)
                                                        except Exception:
                                                            pass
                                                except Exception:
                                                    pass
                                            except Exception:
                                                pass
                                        finally:
                                            # cleanup temp file if created
                                            try:
                                                if tmp_path:
                                                    import os
                                                    os.unlink(tmp_path)
                                            except Exception:
                                                pass

                                except Exception as e:
                                    print("Purchase signature image error:", e)

                            # Write date + time into N39 (centered in that cell)
                            try:
                                if purch_sign:
                                    signed_at = purch_sign.get('signed_at')
                                    signed_by = purch_sign.get('signed_by', '')
                                else:
                                    signed_at = None
                                    signed_by = ''
                                dt_str = ''
                                if signed_at:
                                    try:
                                        dt_str = signed_at.strftime('%d-%m-%Y %H:%M')
                                    except Exception:
                                        dt_str = str(signed_at)
                                cell = ws.cell(row=39, column=14)  # N39
                                cell.value = f"{signed_by} | {dt_str}"
                                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                            except Exception:
                                pass

                        #special-process signature
                        if template_key == 'CR_2':
                            # fetch special-process signature for this PO
                            try:
                                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as _cur:
                                    _cur.execute("""
                                        SELECT
                                            cds.signed_by,
                                            cds.signed_at,
                                            ms.signature_path
                                        FROM cr_department_signatures cds
                                        JOIN master_signatures ms
                                          ON ms.username = cds.signed_by
                                         AND ms.department = cds.department
                                         AND ms.is_active = TRUE
                                        WHERE cds.po_key = %s
                                          AND cds.department = %s
                                        """, (po_key, 'special-process'))
                                    sp_sign = _cur.fetchone()
                            except Exception:
                                sp_sign = None

                            if sp_sign and sp_sign.get('signature_path'):
                                try:
                                    sig_rel_path = sp_sign["signature_path"].lstrip("/")
                                    sig_abs_path = os.path.join(app.root_path, sig_rel_path)

                                    if os.path.exists(sig_abs_path):
                                        tmp_path = None
                                        try:
                                            # Pillow-based processing: autocrop, scale to ~70% of span, then insert
                                            from PIL import Image as PILImage, ImageChops
                                            import tempfile

                                            pil = PILImage.open(sig_abs_path)

                                            # normalize transparency
                                            if pil.mode in ("RGBA", "LA") or (
                                                    pil.mode == "P" and "transparency" in pil.info):
                                                bg = PILImage.new("RGBA", pil.size, (255, 255, 255, 0))
                                                bg.paste(pil, mask=pil.convert("RGBA").split()[-1])
                                                pil = bg
                                            pil = pil.convert("RGBA")

                                            # trim transparent/blank margins
                                            try:
                                                bbox = ImageChops.invert(pil.convert("L")).getbbox()
                                                if bbox:
                                                    pil = pil.crop(bbox)
                                            except Exception:
                                                pass

                                            # target span: columns S..U (19..21), row 38
                                            start_col = 19  # S
                                            end_col = 21  # U
                                            start_row = 38
                                            end_row = 38

                                            from openpyxl.utils import get_column_letter
                                            def col_width_to_pixels(ws, col_idx):
                                                col_letter = get_column_letter(col_idx)
                                                w = ws.column_dimensions.get(col_letter).width
                                                if w is None:
                                                    w = 8.43
                                                if w < 1:
                                                    return int(w * 12)
                                                return int((w - 0.75) * 7 + 5)

                                            def row_height_to_pixels(ws, row_idx):
                                                h = ws.row_dimensions.get(row_idx).height
                                                if h is None:
                                                    h = 15
                                                return int(h * 96.0 / 72.0)

                                            span_width_px = sum(
                                                col_width_to_pixels(ws, c) for c in range(start_col, end_col + 1))
                                            span_height_px = sum(
                                                row_height_to_pixels(ws, r) for r in range(start_row, end_row + 1))
                                            if span_height_px == 0:
                                                span_height_px = row_height_to_pixels(ws, start_row)

                                            # resize image to ~70% of span width preserving aspect ratio
                                            target_w = max(1, int(span_width_px * 0.7))
                                            orig_w, orig_h = pil.size
                                            if orig_w > 0 and target_w < orig_w:
                                                target_h = int((target_w / orig_w) * orig_h)
                                                pil = pil.resize((target_w, target_h), PILImage.LANCZOS)

                                            # save processed image to temp file
                                            tf = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
                                            tmp_path = tf.name
                                            pil.save(tmp_path, format="PNG")
                                            tf.close()

                                            img = Image(tmp_path)
                                            try:
                                                img.width, img.height = pil.size
                                            except Exception:
                                                pass

                                            # insert anchored at S38 then adjust anchor to span S..U and center
                                            ws.add_image(img, "S38")

                                            try:
                                                from openpyxl.utils.units import pixels_to_EMU
                                                anch = getattr(img, "anchor", None)
                                                if anch is not None and hasattr(anch, "_from"):
                                                    fr = anch._from
                                                    to = anch._to
                                                    fr.col = start_col - 1
                                                    fr.row = start_row - 1
                                                    to.col = end_col
                                                    to.row = end_row

                                                    left_px = max(
                                                        int((span_width_px - (img.width or pil.size[0])) / 2.0), 0)
                                                    top_px = max(
                                                        int((span_height_px - (img.height or pil.size[1])) / 2.0), 0)
                                                    try:
                                                        fr.colOff = pixels_to_EMU(left_px)
                                                        fr.rowOff = pixels_to_EMU(top_px)
                                                    except Exception:
                                                        pass
                                            except Exception:
                                                pass

                                        except Exception:
                                            # fallback if Pillow not available or processing fails
                                            try:
                                                img = Image(sig_abs_path)
                                                img.width = 120
                                                img.height = 50
                                                ws.add_image(img, "S38")
                                                # best-effort centering
                                                try:
                                                    from openpyxl.utils import get_column_letter
                                                    from openpyxl.utils.units import pixels_to_EMU
                                                    def col_width_to_pixels(ws, col_idx):
                                                        col_letter = get_column_letter(col_idx)
                                                        w = ws.column_dimensions.get(col_letter).width
                                                        if w is None:
                                                            w = 8.43
                                                        if w < 1:
                                                            return int(w * 12)
                                                        return int((w - 0.75) * 7 + 5)

                                                    def row_height_to_pixels(ws, row_idx):
                                                        h = ws.row_dimensions.get(row_idx).height
                                                        if h is None:
                                                            h = 15
                                                        return int(h * 96.0 / 72.0)

                                                    span_width_px = sum(
                                                        col_width_to_pixels(ws, c) for c in range(19, 22))
                                                    span_height_px = row_height_to_pixels(ws, 38)
                                                    anch = getattr(img, "anchor", None)
                                                    if anch is not None and hasattr(anch, "_from"):
                                                        fr = anch._from
                                                        to = anch._to
                                                        fr.col = 19 - 1
                                                        fr.row = 38 - 1
                                                        to.col = 21
                                                        to.row = 38
                                                        left_px = max(int((span_width_px - img.width) / 2.0), 0)
                                                        top_px = max(int((span_height_px - img.height) / 2.0), 0)
                                                        try:
                                                            fr.colOff = pixels_to_EMU(left_px)
                                                            fr.rowOff = pixels_to_EMU(top_px)
                                                        except Exception:
                                                            pass
                                                except Exception:
                                                    pass
                                            except Exception:
                                                pass
                                        finally:
                                            # cleanup temp file if created
                                            try:
                                                if tmp_path:
                                                    import os
                                                    os.unlink(tmp_path)
                                            except Exception:
                                                pass

                                except Exception as e:
                                    print("Special Process signature image error:", e)

                            # Write date + time into S39 (centered)
                            try:
                                if sp_sign:
                                    signed_at = sp_sign.get('signed_at')
                                    signed_by = sp_sign.get('signed_by', '')
                                else:
                                    signed_at = None
                                    signed_by = ''
                                dt_str = ''
                                if signed_at:
                                    try:
                                        dt_str = signed_at.strftime('%d-%m-%Y %H:%M')
                                    except Exception:
                                        dt_str = str(signed_at)
                                cell = ws.cell(row=39, column=19)  # S39
                                cell.value = f"{signed_by} | {dt_str}"
                                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                            except Exception:
                                pass
                            #welding signature
                        if template_key == 'CR_2':
                            # fetch welding signature for this PO
                            try:
                                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as _cur:
                                    _cur.execute("""
                                        SELECT
                                            cds.signed_by,
                                            cds.signed_at,
                                            ms.signature_path
                                        FROM cr_department_signatures cds
                                        JOIN master_signatures ms
                                          ON ms.username = cds.signed_by
                                         AND ms.department = cds.department
                                         AND ms.is_active = TRUE
                                        WHERE cds.po_key = %s
                                          AND cds.department = %s
                                        """, (po_key, 'welding'))
                                    weld_sign = _cur.fetchone()
                            except Exception:
                                weld_sign = None

                            if weld_sign and weld_sign.get('signature_path'):
                                try:
                                    sig_rel_path = weld_sign["signature_path"].lstrip("/")
                                    sig_abs_path = os.path.join(app.root_path, sig_rel_path)

                                    if os.path.exists(sig_abs_path):
                                        tmp_path = None
                                        try:
                                            # Pillow-based processing: autocrop, scale to ~70% of span, then insert
                                            from PIL import Image as PILImage, ImageChops
                                            import tempfile

                                            pil = PILImage.open(sig_abs_path)

                                            # normalize transparency
                                            if pil.mode in ("RGBA", "LA") or (
                                                    pil.mode == "P" and "transparency" in pil.info):
                                                bg = PILImage.new("RGBA", pil.size, (255, 255, 255, 0))
                                                bg.paste(pil, mask=pil.convert("RGBA").split()[-1])
                                                pil = bg
                                            pil = pil.convert("RGBA")

                                            # trim transparent/blank margins
                                            try:
                                                bbox = ImageChops.invert(pil.convert("L")).getbbox()
                                                if bbox:
                                                    pil = pil.crop(bbox)
                                            except Exception:
                                                pass

                                            # target span: columns V..X (22..24), row 38
                                            start_col = 22  # V
                                            end_col = 24  # X
                                            start_row = 38
                                            end_row = 38

                                            from openpyxl.utils import get_column_letter
                                            def col_width_to_pixels(ws, col_idx):
                                                col_letter = get_column_letter(col_idx)
                                                w = ws.column_dimensions.get(col_letter).width
                                                if w is None:
                                                    w = 8.43
                                                if w < 1:
                                                    return int(w * 12)
                                                return int((w - 0.75) * 7 + 5)

                                            def row_height_to_pixels(ws, row_idx):
                                                h = ws.row_dimensions.get(row_idx).height
                                                if h is None:
                                                    h = 15
                                                return int(h * 96.0 / 72.0)

                                            span_width_px = sum(
                                                col_width_to_pixels(ws, c) for c in range(start_col, end_col + 1))
                                            span_height_px = sum(
                                                row_height_to_pixels(ws, r) for r in range(start_row, end_row + 1))
                                            if span_height_px == 0:
                                                span_height_px = row_height_to_pixels(ws, start_row)

                                            # scale to ~70% of span width
                                            target_w = max(1, int(span_width_px * 0.7))
                                            orig_w, orig_h = pil.size
                                            if orig_w > 0 and target_w < orig_w:
                                                target_h = int((target_w / orig_w) * orig_h)
                                                pil = pil.resize((target_w, target_h), PILImage.LANCZOS)

                                            tf = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
                                            tmp_path = tf.name
                                            pil.save(tmp_path, format="PNG")
                                            tf.close()

                                            img = Image(tmp_path)
                                            try:
                                                img.width, img.height = pil.size
                                            except Exception:
                                                pass

                                            # insert anchored at V38 then adjust anchor to span V..X and center
                                            ws.add_image(img, "V38")

                                            try:
                                                from openpyxl.utils.units import pixels_to_EMU
                                                anch = getattr(img, "anchor", None)
                                                if anch is not None and hasattr(anch, "_from"):
                                                    fr = anch._from
                                                    to = anch._to
                                                    # set anchor span (zero-based)
                                                    fr.col = start_col - 1
                                                    fr.row = start_row - 1
                                                    to.col = end_col
                                                    to.row = end_row

                                                    left_px = max(
                                                        int((span_width_px - (img.width or pil.size[0])) / 2.0), 0)
                                                    top_px = max(
                                                        int((span_height_px - (img.height or pil.size[1])) / 2.0), 0)
                                                    try:
                                                        fr.colOff = pixels_to_EMU(left_px)
                                                        fr.rowOff = pixels_to_EMU(top_px)
                                                    except Exception:
                                                        # ignore EMU assignment issues
                                                        pass
                                            except Exception:
                                                pass

                                        except Exception:
                                            # fallback if Pillow not available or processing fails
                                            try:
                                                img = Image(sig_abs_path)
                                                img.width = 120
                                                img.height = 50
                                                ws.add_image(img, "V38")
                                                # best-effort centering
                                                try:
                                                    from openpyxl.utils import get_column_letter
                                                    from openpyxl.utils.units import pixels_to_EMU
                                                    def col_width_to_pixels(ws, col_idx):
                                                        col_letter = get_column_letter(col_idx)
                                                        w = ws.column_dimensions.get(col_letter).width
                                                        if w is None:
                                                            w = 8.43
                                                        if w < 1:
                                                            return int(w * 12)
                                                        return int((w - 0.75) * 7 + 5)

                                                    def row_height_to_pixels(ws, row_idx):
                                                        h = ws.row_dimensions.get(row_idx).height
                                                        if h is None:
                                                            h = 15
                                                        return int(h * 96.0 / 72.0)

                                                    span_width_px = sum(
                                                        col_width_to_pixels(ws, c) for c in range(22, 25))
                                                    span_height_px = row_height_to_pixels(ws, 38)
                                                    anch = getattr(img, "anchor", None)
                                                    if anch is not None and hasattr(anch, "_from"):
                                                        fr = anch._from
                                                        to = anch._to
                                                        fr.col = 22 - 1
                                                        fr.row = 38 - 1
                                                        to.col = 24
                                                        to.row = 38
                                                        left_px = max(int((span_width_px - img.width) / 2.0), 0)
                                                        top_px = max(int((span_height_px - img.height) / 2.0), 0)
                                                        try:
                                                            fr.colOff = pixels_to_EMU(left_px)
                                                            fr.rowOff = pixels_to_EMU(top_px)
                                                        except Exception:
                                                            pass
                                                except Exception:
                                                    pass
                                            except Exception:
                                                pass
                                        finally:
                                            # cleanup temp file if created
                                            try:
                                                if tmp_path:
                                                    import os
                                                    os.unlink(tmp_path)
                                            except Exception:
                                                pass

                                except Exception as e:
                                    print("Welding signature image error:", e)

                            # Write date + time into V39 (centered in that cell)
                            try:
                                if weld_sign:
                                    signed_at = weld_sign.get('signed_at')
                                    signed_by = weld_sign.get('signed_by', '')
                                else:
                                    signed_at = None
                                    signed_by = ''
                                dt_str = ''
                                if signed_at:
                                    try:
                                        dt_str = signed_at.strftime('%d-%m-%Y %H:%M')
                                    except Exception:
                                        dt_str = str(signed_at)
                                cell = ws.cell(row=39, column=22)  # V39
                                cell.value = f"{signed_by} | {dt_str}"
                                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                            except Exception:
                                pass

                    #assembly/testing signature
                        if template_key == 'CR_2':
                            # fetch assembly/testing signature for this PO
                            try:
                                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as _cur:
                                    _cur.execute("""
                                        SELECT
                                            cds.signed_by,
                                            cds.signed_at,
                                            ms.signature_path
                                        FROM cr_department_signatures cds
                                        JOIN master_signatures ms
                                          ON ms.username = cds.signed_by
                                         AND ms.department = cds.department
                                         AND ms.is_active = TRUE
                                        WHERE cds.po_key = %s
                                          AND cds.department = %s
                                        """, (po_key, 'assembly'))
                                    assembly_sign = _cur.fetchone()
                            except Exception:
                                assembly_sign = None

                            if assembly_sign and assembly_sign.get('signature_path'):
                                try:
                                    sig_rel_path = assembly_sign["signature_path"].lstrip("/")
                                    sig_abs_path = os.path.join(app.root_path, sig_rel_path)

                                    if os.path.exists(sig_abs_path):
                                        tmp_path = None
                                        try:
                                            # Pillow-based processing: autocrop, scale to ~70% of span, then insert
                                            from PIL import Image as PILImage, ImageChops
                                            import tempfile

                                            pil = PILImage.open(sig_abs_path)

                                            # normalize transparency
                                            if pil.mode in ("RGBA", "LA") or (
                                                    pil.mode == "P" and "transparency" in pil.info):
                                                bg = PILImage.new("RGBA", pil.size, (255, 255, 255, 0))
                                                bg.paste(pil, mask=pil.convert("RGBA").split()[-1])
                                                pil = bg
                                            pil = pil.convert("RGBA")

                                            # trim transparent/blank margins
                                            try:
                                                bbox = ImageChops.invert(pil.convert("L")).getbbox()
                                                if bbox:
                                                    pil = pil.crop(bbox)
                                            except Exception:
                                                pass

                                            # target span: columns Y..AB (25..28), row 38
                                            start_col = 25  # Y
                                            end_col = 28  # AB
                                            start_row = 38
                                            end_row = 38

                                            from openpyxl.utils import get_column_letter
                                            def col_width_to_pixels(ws, col_idx):
                                                col_letter = get_column_letter(col_idx)
                                                w = ws.column_dimensions.get(col_letter).width
                                                if w is None:
                                                    w = 8.43
                                                if w < 1:
                                                    return int(w * 12)
                                                return int((w - 0.75) * 7 + 5)

                                            def row_height_to_pixels(ws, row_idx):
                                                h = ws.row_dimensions.get(row_idx).height
                                                if h is None:
                                                    h = 15
                                                return int(h * 96.0 / 72.0)

                                            span_width_px = sum(
                                                col_width_to_pixels(ws, c) for c in range(start_col, end_col + 1))
                                            span_height_px = sum(
                                                row_height_to_pixels(ws, r) for r in range(start_row, end_row + 1))
                                            if span_height_px == 0:
                                                span_height_px = row_height_to_pixels(ws, start_row)

                                            # resize image to ~70% of span width preserving aspect ratio
                                            target_w = max(1, int(span_width_px * 0.7))
                                            orig_w, orig_h = pil.size
                                            if orig_w > 0 and target_w < orig_w:
                                                target_h = int((target_w / orig_w) * orig_h)
                                                pil = pil.resize((target_w, target_h), PILImage.LANCZOS)

                                            # save processed image to temp file
                                            tf = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
                                            tmp_path = tf.name
                                            pil.save(tmp_path, format="PNG")
                                            tf.close()

                                            img = Image(tmp_path)
                                            try:
                                                img.width, img.height = pil.size
                                            except Exception:
                                                pass

                                            # insert anchored at Y38 then adjust anchor to span Y..AB and center
                                            ws.add_image(img, "Y38")

                                            try:
                                                from openpyxl.utils.units import pixels_to_EMU
                                                anch = getattr(img, "anchor", None)
                                                if anch is not None and hasattr(anch, "_from"):
                                                    fr = anch._from
                                                    to = anch._to
                                                    fr.col = start_col - 1
                                                    fr.row = start_row - 1
                                                    to.col = end_col
                                                    to.row = end_row

                                                    left_px = max(
                                                        int((span_width_px - (img.width or pil.size[0])) / 2.0), 0)
                                                    top_px = max(
                                                        int((span_height_px - (img.height or pil.size[1])) / 2.0), 0)
                                                    try:
                                                        fr.colOff = pixels_to_EMU(left_px)
                                                        fr.rowOff = pixels_to_EMU(top_px)
                                                    except Exception:
                                                        pass
                                            except Exception:
                                                pass

                                        except Exception:
                                            # fallback if Pillow not available or processing fails
                                            try:
                                                img = Image(sig_abs_path)
                                                img.width = 120
                                                img.height = 50
                                                ws.add_image(img, "Y38")
                                                # best-effort centering
                                                try:
                                                    from openpyxl.utils import get_column_letter
                                                    from openpyxl.utils.units import pixels_to_EMU
                                                    def col_width_to_pixels(ws, col_idx):
                                                        col_letter = get_column_letter(col_idx)
                                                        w = ws.column_dimensions.get(col_letter).width
                                                        if w is None:
                                                            w = 8.43
                                                        if w < 1:
                                                            return int(w * 12)
                                                        return int((w - 0.75) * 7 + 5)

                                                    def row_height_to_pixels(ws, row_idx):
                                                        h = ws.row_dimensions.get(row_idx).height
                                                        if h is None:
                                                            h = 15
                                                        return int(h * 96.0 / 72.0)

                                                    span_width_px = sum(
                                                        col_width_to_pixels(ws, c) for c in range(25, 29))
                                                    span_height_px = row_height_to_pixels(ws, 38)
                                                    anch = getattr(img, "anchor", None)
                                                    if anch is not None and hasattr(anch, "_from"):
                                                        fr = anch._from
                                                        to = anch._to
                                                        fr.col = 25 - 1
                                                        fr.row = 38 - 1
                                                        to.col = 28
                                                        to.row = 38
                                                        left_px = max(int((span_width_px - img.width) / 2.0), 0)
                                                        top_px = max(int((span_height_px - img.height) / 2.0), 0)
                                                        try:
                                                            fr.colOff = pixels_to_EMU(left_px)
                                                            fr.rowOff = pixels_to_EMU(top_px)
                                                        except Exception:
                                                            pass
                                                except Exception:
                                                    pass
                                            except Exception:
                                                pass
                                        finally:
                                            # cleanup temp file if created
                                            try:
                                                if tmp_path:
                                                    import os
                                                    os.unlink(tmp_path)
                                            except Exception:
                                                pass

                                except Exception as e:
                                    print("Assembly & Testing signature image error:", e)

                            # Write date + time into Y39 (centered)
                            try:
                                if assembly_sign:
                                    signed_at = assembly_sign.get('signed_at')
                                    signed_by = assembly_sign.get('signed_by', '')
                                else:
                                    signed_at = None
                                    signed_by = ''
                                dt_str = ''
                                if signed_at:
                                    try:
                                        dt_str = signed_at.strftime('%d-%m-%Y %H:%M')
                                    except Exception:
                                        dt_str = str(signed_at)
                                cell = ws.cell(row=39, column=25)  # Y39
                                cell.value = f"{signed_by} | {dt_str}"
                                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                            except Exception:
                                pass

                        #quality signature
                        if template_key == 'CR_3':
                            # fetch quality signature for this PO
                            try:
                                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as _cur:
                                    _cur.execute("""
                                        SELECT
                                            cds.signed_by,
                                            cds.signed_at,
                                            ms.signature_path
                                        FROM cr_department_signatures cds
                                        JOIN master_signatures ms
                                          ON ms.username = cds.signed_by
                                         AND ms.department = cds.department
                                         AND ms.is_active = TRUE
                                        WHERE cds.po_key = %s
                                          AND cds.department = %s
                                        """, (po_key, 'quality'))
                                    qc_sign = _cur.fetchone()
                            except Exception:
                                qc_sign = None

                            if qc_sign and qc_sign.get('signature_path'):
                                try:
                                    sig_rel_path = qc_sign["signature_path"].lstrip("/")
                                    sig_abs_path = os.path.join(app.root_path, sig_rel_path)

                                    if os.path.exists(sig_abs_path):
                                        tmp_path = None
                                        try:
                                            # Pillow-based processing: autocrop, scale to ~70% of span, then insert
                                            from PIL import Image as PILImage, ImageChops
                                            import tempfile

                                            pil = PILImage.open(sig_abs_path)

                                            # normalize transparency
                                            if pil.mode in ("RGBA", "LA") or (
                                                    pil.mode == "P" and "transparency" in pil.info):
                                                bg = PILImage.new("RGBA", pil.size, (255, 255, 255, 0))
                                                bg.paste(pil, mask=pil.convert("RGBA").split()[-1])
                                                pil = bg
                                            pil = pil.convert("RGBA")

                                            # trim transparent/blank margins
                                            try:
                                                bbox = ImageChops.invert(pil.convert("L")).getbbox()
                                                if bbox:
                                                    pil = pil.crop(bbox)
                                            except Exception:
                                                pass

                                            # target span: columns F..O (6..15), row 38
                                            start_col = 6  # F
                                            end_col = 15  # O
                                            start_row = 38
                                            end_row = 38

                                            from openpyxl.utils import get_column_letter
                                            def col_width_to_pixels(ws, col_idx):
                                                col_letter = get_column_letter(col_idx)
                                                w = ws.column_dimensions.get(col_letter).width
                                                if w is None:
                                                    w = 8.43
                                                if w < 1:
                                                    return int(w * 12)
                                                return int((w - 0.75) * 7 + 5)

                                            def row_height_to_pixels(ws, row_idx):
                                                h = ws.row_dimensions.get(row_idx).height
                                                if h is None:
                                                    h = 15
                                                return int(h * 96.0 / 72.0)

                                            span_width_px = sum(
                                                col_width_to_pixels(ws, c) for c in range(start_col, end_col + 1))
                                            span_height_px = sum(
                                                row_height_to_pixels(ws, r) for r in range(start_row, end_row + 1))
                                            if span_height_px == 0:
                                                span_height_px = row_height_to_pixels(ws, start_row)

                                            # resize image to ~70% of span width preserving aspect ratio
                                            target_w = max(1, int(span_width_px * 0.7))
                                            orig_w, orig_h = pil.size
                                            if orig_w > 0 and target_w < orig_w:
                                                target_h = int((target_w / orig_w) * orig_h)
                                                pil = pil.resize((target_w, target_h), PILImage.LANCZOS)

                                            # save processed image to temp file
                                            tf = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
                                            tmp_path = tf.name
                                            pil.save(tmp_path, format="PNG")
                                            tf.close()

                                            img = Image(tmp_path)
                                            try:
                                                img.width, img.height = pil.size
                                            except Exception:
                                                pass

                                            # insert anchored at F38 then adjust anchor to span F..O and center
                                            ws.add_image(img, "F38")

                                            try:
                                                from openpyxl.utils.units import pixels_to_EMU
                                                anch = getattr(img, "anchor", None)
                                                if anch is not None and hasattr(anch, "_from"):
                                                    fr = anch._from
                                                    to = anch._to
                                                    fr.col = start_col - 1
                                                    fr.row = start_row - 1
                                                    to.col = end_col
                                                    to.row = end_row

                                                    left_px = max(
                                                        int((span_width_px - (img.width or pil.size[0])) / 2.0), 0)
                                                    top_px = max(
                                                        int((span_height_px - (img.height or pil.size[1])) / 2.0), 0)
                                                    try:
                                                        fr.colOff = pixels_to_EMU(left_px)
                                                        fr.rowOff = pixels_to_EMU(top_px)
                                                    except Exception:
                                                        pass
                                            except Exception:
                                                pass

                                        except Exception:
                                            # fallback if Pillow not available or processing fails
                                            try:
                                                img = Image(sig_abs_path)
                                                img.width = 120
                                                img.height = 50
                                                ws.add_image(img, "F38")
                                                # best-effort centering
                                                try:
                                                    from openpyxl.utils import get_column_letter
                                                    from openpyxl.utils.units import pixels_to_EMU
                                                    def col_width_to_pixels(ws, col_idx):
                                                        col_letter = get_column_letter(col_idx)
                                                        w = ws.column_dimensions.get(col_letter).width
                                                        if w is None:
                                                            w = 8.43
                                                        if w < 1:
                                                            return int(w * 12)
                                                        return int((w - 0.75) * 7 + 5)

                                                    def row_height_to_pixels(ws, row_idx):
                                                        h = ws.row_dimensions.get(row_idx).height
                                                        if h is None:
                                                            h = 15
                                                        return int(h * 96.0 / 72.0)

                                                    span_width_px = sum(
                                                        col_width_to_pixels(ws, c) for c in range(6, 16))
                                                    span_height_px = row_height_to_pixels(ws, 38)
                                                    anch = getattr(img, "anchor", None)
                                                    if anch is not None and hasattr(anch, "_from"):
                                                        fr = anch._from
                                                        to = anch._to
                                                        fr.col = 6 - 1
                                                        fr.row = 38 - 1
                                                        to.col = 15
                                                        to.row = 38
                                                        left_px = max(int((span_width_px - img.width) / 2.0), 0)
                                                        top_px = max(int((span_height_px - img.height) / 2.0), 0)
                                                        try:
                                                            fr.colOff = pixels_to_EMU(left_px)
                                                            fr.rowOff = pixels_to_EMU(top_px)
                                                        except Exception:
                                                            pass
                                                except Exception:
                                                    pass
                                            except Exception:
                                                pass
                                        finally:
                                            # cleanup temp file if created
                                            try:
                                                if tmp_path:
                                                    import os
                                                    os.unlink(tmp_path)
                                            except Exception:
                                                pass

                                except Exception as e:
                                    print("Quality signature image error:", e)

                            # Write date + time into F39 (centered)
                            try:
                                if qc_sign:
                                    signed_at = qc_sign.get('signed_at')
                                    signed_by = qc_sign.get('signed_by', '')
                                else:
                                    signed_at = None
                                    signed_by = ''
                                dt_str = ''
                                if signed_at:
                                    try:
                                        dt_str = signed_at.strftime('%d-%m-%Y %H:%M')
                                    except Exception:
                                        dt_str = str(signed_at)
                                cell = ws.cell(row=39, column=6)  # F39
                                cell.value = f"{signed_by} | {dt_str}"
                                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                            except Exception:
                                pass

                        #painting/despatch signature
                        if template_key == 'CR_3':
                            # fetch painting/despatch signature for this PO
                            try:
                                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as _cur:
                                    _cur.execute("""
                                        SELECT
                                            cds.signed_by,
                                            cds.signed_at,
                                            ms.signature_path
                                        FROM cr_department_signatures cds
                                        JOIN master_signatures ms
                                          ON ms.username = cds.signed_by
                                         AND ms.department = cds.department
                                         AND ms.is_active = TRUE
                                        WHERE cds.po_key = %s
                                          AND cds.department = %s
                                        """, (po_key, 'painting'))
                                    paint_sign = _cur.fetchone()
                            except Exception:
                                paint_sign = None

                            if paint_sign and paint_sign.get('signature_path'):
                                try:
                                    sig_rel_path = paint_sign["signature_path"].lstrip("/")
                                    sig_abs_path = os.path.join(app.root_path, sig_rel_path)

                                    if os.path.exists(sig_abs_path):
                                        tmp_path = None
                                        try:
                                            # Pillow-based processing: autocrop, scale to ~70% of span, then insert
                                            from PIL import Image as PILImage, ImageChops
                                            import tempfile

                                            pil = PILImage.open(sig_abs_path)

                                            # normalize transparency
                                            if pil.mode in ("RGBA", "LA") or (
                                                    pil.mode == "P" and "transparency" in pil.info):
                                                bg = PILImage.new("RGBA", pil.size, (255, 255, 255, 0))
                                                bg.paste(pil, mask=pil.convert("RGBA").split()[-1])
                                                pil = bg
                                            pil = pil.convert("RGBA")

                                            # trim transparent/blank margins
                                            try:
                                                bbox = ImageChops.invert(pil.convert("L")).getbbox()
                                                if bbox:
                                                    pil = pil.crop(bbox)
                                            except Exception:
                                                pass

                                            # target span: columns P..T (16..20), row 38
                                            start_col = 16  # P
                                            end_col = 20  # T
                                            start_row = 38
                                            end_row = 38

                                            from openpyxl.utils import get_column_letter
                                            def col_width_to_pixels(ws, col_idx):
                                                col_letter = get_column_letter(col_idx)
                                                w = ws.column_dimensions.get(col_letter).width
                                                if w is None:
                                                    w = 8.43
                                                if w < 1:
                                                    return int(w * 12)
                                                return int((w - 0.75) * 7 + 5)

                                            def row_height_to_pixels(ws, row_idx):
                                                h = ws.row_dimensions.get(row_idx).height
                                                if h is None:
                                                    h = 15
                                                return int(h * 96.0 / 72.0)

                                            span_width_px = sum(
                                                col_width_to_pixels(ws, c) for c in range(start_col, end_col + 1))
                                            span_height_px = sum(
                                                row_height_to_pixels(ws, r) for r in range(start_row, end_row + 1))
                                            if span_height_px == 0:
                                                span_height_px = row_height_to_pixels(ws, start_row)

                                            # scale to ~70% of span width
                                            target_w = max(1, int(span_width_px * 0.7))
                                            orig_w, orig_h = pil.size
                                            if orig_w > 0 and target_w < orig_w:
                                                target_h = int((target_w / orig_w) * orig_h)
                                                pil = pil.resize((target_w, target_h), PILImage.LANCZOS)

                                            tf = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
                                            tmp_path = tf.name
                                            pil.save(tmp_path, format="PNG")
                                            tf.close()

                                            img = Image(tmp_path)
                                            try:
                                                img.width, img.height = pil.size
                                            except Exception:
                                                pass

                                            # insert anchored at P38 then adjust anchor to span P..T and center
                                            ws.add_image(img, "P38")

                                            try:
                                                from openpyxl.utils.units import pixels_to_EMU
                                                anch = getattr(img, "anchor", None)
                                                if anch is not None and hasattr(anch, "_from"):
                                                    fr = anch._from
                                                    to = anch._to
                                                    fr.col = start_col - 1
                                                    fr.row = start_row - 1
                                                    to.col = end_col
                                                    to.row = end_row

                                                    left_px = max(
                                                        int((span_width_px - (img.width or pil.size[0])) / 2.0), 0)
                                                    top_px = max(
                                                        int((span_height_px - (img.height or pil.size[1])) / 2.0), 0)
                                                    try:
                                                        fr.colOff = pixels_to_EMU(left_px)
                                                        fr.rowOff = pixels_to_EMU(top_px)
                                                    except Exception:
                                                        pass
                                            except Exception:
                                                pass

                                        except Exception:
                                            # fallback if Pillow not available or processing fails
                                            try:
                                                img = Image(sig_abs_path)
                                                img.width = 120
                                                img.height = 50
                                                ws.add_image(img, "P38")
                                                # best-effort centering
                                                try:
                                                    from openpyxl.utils import get_column_letter
                                                    from openpyxl.utils.units import pixels_to_EMU
                                                    def col_width_to_pixels(ws, col_idx):
                                                        col_letter = get_column_letter(col_idx)
                                                        w = ws.column_dimensions.get(col_letter).width
                                                        if w is None:
                                                            w = 8.43
                                                        if w < 1:
                                                            return int(w * 12)
                                                        return int((w - 0.75) * 7 + 5)

                                                    def row_height_to_pixels(ws, row_idx):
                                                        h = ws.row_dimensions.get(row_idx).height
                                                        if h is None:
                                                            h = 15
                                                        return int(h * 96.0 / 72.0)

                                                    span_width_px = sum(
                                                        col_width_to_pixels(ws, c) for c in range(16, 21))
                                                    span_height_px = row_height_to_pixels(ws, 38)
                                                    anch = getattr(img, "anchor", None)
                                                    if anch is not None and hasattr(anch, "_from"):
                                                        fr = anch._from
                                                        to = anch._to
                                                        fr.col = 16 - 1
                                                        fr.row = 38 - 1
                                                        to.col = 20
                                                        to.row = 38
                                                        left_px = max(int((span_width_px - img.width) / 2.0), 0)
                                                        top_px = max(int((span_height_px - img.height) / 2.0), 0)
                                                        try:
                                                            fr.colOff = pixels_to_EMU(left_px)
                                                            fr.rowOff = pixels_to_EMU(top_px)
                                                        except Exception:
                                                            pass
                                                except Exception:
                                                    pass
                                            except Exception:
                                                pass
                                        finally:
                                            # cleanup temp file if created
                                            try:
                                                if tmp_path:
                                                    import os
                                                    os.unlink(tmp_path)
                                            except Exception:
                                                pass

                                except Exception as e:
                                    print("Painting/Despatch (CR_3) signature image error:", e)

                            # Write date + time into P39 (centered)
                            try:
                                if paint_sign:
                                    signed_at = paint_sign.get('signed_at')
                                    signed_by = paint_sign.get('signed_by', '')
                                else:
                                    signed_at = None
                                    signed_by = ''
                                dt_str = ''
                                if signed_at:
                                    try:
                                        dt_str = signed_at.strftime('%d-%m-%Y %H:%M')
                                    except Exception:
                                        dt_str = str(signed_at)
                                cell = ws.cell(row=39, column=16)  # P39
                                cell.value = f"{signed_by} | {dt_str}"
                                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                            except Exception:
                                pass

                    #customer-service signature
                        if template_key == 'CR_3':
                            # fetch customer-service signature for this PO
                            try:
                                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as _cur:
                                    _cur.execute("""
                                        SELECT
                                            cds.signed_by,
                                            cds.signed_at,
                                            ms.signature_path
                                        FROM cr_department_signatures cds
                                        JOIN master_signatures ms
                                          ON ms.username = cds.signed_by
                                         AND ms.department = cds.department
                                         AND ms.is_active = TRUE
                                        WHERE cds.po_key = %s
                                          AND cds.department = %s
                                        """, (po_key, 'customer-service'))
                                    cs_sign = _cur.fetchone()
                            except Exception:
                                cs_sign = None

                            if cs_sign and cs_sign.get('signature_path'):
                                try:
                                    sig_rel_path = cs_sign["signature_path"].lstrip("/")
                                    sig_abs_path = os.path.join(app.root_path, sig_rel_path)

                                    if os.path.exists(sig_abs_path):
                                        tmp_path = None
                                        try:
                                            # Pillow-based processing: autocrop, scale to ~70% of span, then insert
                                            from PIL import Image as PILImage, ImageChops
                                            import tempfile

                                            pil = PILImage.open(sig_abs_path)

                                            # normalize transparency
                                            if pil.mode in ("RGBA", "LA") or (
                                                    pil.mode == "P" and "transparency" in pil.info):
                                                bg = PILImage.new("RGBA", pil.size, (255, 255, 255, 0))
                                                bg.paste(pil, mask=pil.convert("RGBA").split()[-1])
                                                pil = bg
                                            pil = pil.convert("RGBA")

                                            # trim transparent/blank margins
                                            try:
                                                bbox = ImageChops.invert(pil.convert("L")).getbbox()
                                                if bbox:
                                                    pil = pil.crop(bbox)
                                            except Exception:
                                                pass

                                            # target span: columns U..X (21..24), row 38
                                            start_col = 21  # U
                                            end_col = 24  # X
                                            start_row = 38
                                            end_row = 38

                                            from openpyxl.utils import get_column_letter
                                            def col_width_to_pixels(ws, col_idx):
                                                col_letter = get_column_letter(col_idx)
                                                w = ws.column_dimensions.get(col_letter).width
                                                if w is None:
                                                    w = 8.43
                                                if w < 1:
                                                    return int(w * 12)
                                                return int((w - 0.75) * 7 + 5)

                                            def row_height_to_pixels(ws, row_idx):
                                                h = ws.row_dimensions.get(row_idx).height
                                                if h is None:
                                                    h = 15
                                                return int(h * 96.0 / 72.0)

                                            span_width_px = sum(
                                                col_width_to_pixels(ws, c) for c in range(start_col, end_col + 1))
                                            span_height_px = sum(
                                                row_height_to_pixels(ws, r) for r in range(start_row, end_row + 1))
                                            if span_height_px == 0:
                                                span_height_px = row_height_to_pixels(ws, start_row)

                                            # resize image to ~70% of span width preserving aspect ratio
                                            target_w = max(1, int(span_width_px * 0.7))
                                            orig_w, orig_h = pil.size
                                            if orig_w > 0 and target_w < orig_w:
                                                target_h = int((target_w / orig_w) * orig_h)
                                                pil = pil.resize((target_w, target_h), PILImage.LANCZOS)

                                            # save processed image to temp file
                                            tf = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
                                            tmp_path = tf.name
                                            pil.save(tmp_path, format="PNG")
                                            tf.close()

                                            img = Image(tmp_path)
                                            try:
                                                img.width, img.height = pil.size
                                            except Exception:
                                                pass

                                            # insert anchored at U38 then adjust anchor to span U..X and center
                                            ws.add_image(img, "U38")

                                            try:
                                                from openpyxl.utils.units import pixels_to_EMU
                                                anch = getattr(img, "anchor", None)
                                                if anch is not None and hasattr(anch, "_from"):
                                                    fr = anch._from
                                                    to = anch._to
                                                    fr.col = start_col - 1
                                                    fr.row = start_row - 1
                                                    to.col = end_col
                                                    to.row = end_row

                                                    left_px = max(
                                                        int((span_width_px - (img.width or pil.size[0])) / 2.0), 0)
                                                    top_px = max(
                                                        int((span_height_px - (img.height or pil.size[1])) / 2.0), 0)
                                                    try:
                                                        fr.colOff = pixels_to_EMU(left_px)
                                                        fr.rowOff = pixels_to_EMU(top_px)
                                                    except Exception:
                                                        pass
                                            except Exception:
                                                pass

                                        except Exception:
                                            # fallback if Pillow not available or processing fails
                                            try:
                                                img = Image(sig_abs_path)
                                                img.width = 120
                                                img.height = 50
                                                ws.add_image(img, "U38")
                                                # best-effort centering
                                                try:
                                                    from openpyxl.utils import get_column_letter
                                                    from openpyxl.utils.units import pixels_to_EMU
                                                    def col_width_to_pixels(ws, col_idx):
                                                        col_letter = get_column_letter(col_idx)
                                                        w = ws.column_dimensions.get(col_letter).width
                                                        if w is None:
                                                            w = 8.43
                                                        if w < 1:
                                                            return int(w * 12)
                                                        return int((w - 0.75) * 7 + 5)

                                                    def row_height_to_pixels(ws, row_idx):
                                                        h = ws.row_dimensions.get(row_idx).height
                                                        if h is None:
                                                            h = 15
                                                        return int(h * 96.0 / 72.0)

                                                    span_width_px = sum(
                                                        col_width_to_pixels(ws, c) for c in range(21, 25))
                                                    span_height_px = row_height_to_pixels(ws, 38)
                                                    anch = getattr(img, "anchor", None)
                                                    if anch is not None and hasattr(anch, "_from"):
                                                        fr = anch._from
                                                        to = anch._to
                                                        fr.col = 21 - 1
                                                        fr.row = 38 - 1
                                                        to.col = 24
                                                        to.row = 38
                                                        left_px = max(int((span_width_px - img.width) / 2.0), 0)
                                                        top_px = max(int((span_height_px - img.height) / 2.0), 0)
                                                        try:
                                                            fr.colOff = pixels_to_EMU(left_px)
                                                            fr.rowOff = pixels_to_EMU(top_px)
                                                        except Exception:
                                                            pass
                                                except Exception:
                                                    pass
                                            except Exception:
                                                pass
                                        finally:
                                            # cleanup temp file if created
                                            try:
                                                if tmp_path:
                                                    import os
                                                    os.unlink(tmp_path)
                                            except Exception:
                                                pass

                                except Exception as e:
                                    print("Customer Service signature image error:", e)

                            # Write date + time into U39 (centered)
                            try:
                                if cs_sign:
                                    signed_at = cs_sign.get('signed_at')
                                    signed_by = cs_sign.get('signed_by', '')
                                else:
                                    signed_at = None
                                    signed_by = ''
                                dt_str = ''
                                if signed_at:
                                    try:
                                        dt_str = signed_at.strftime('%d-%m-%Y %H:%M')
                                    except Exception:
                                        dt_str = str(signed_at)
                                cell = ws.cell(row=39, column=21)  # U39
                                cell.value = f"{signed_by} | {dt_str}"
                                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                            except Exception:
                                pass
                            
                        #commercial signature
                        if template_key == 'CR_3':
                            # fetch commercial signature for this PO
                            try:
                                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as _cur:
                                    _cur.execute("""
                                        SELECT
                                            cds.signed_by,
                                            cds.signed_at,
                                            ms.signature_path
                                        FROM cr_department_signatures cds
                                        JOIN master_signatures ms
                                          ON ms.username = cds.signed_by
                                         AND ms.department = cds.department
                                         AND ms.is_active = TRUE
                                        WHERE cds.po_key = %s
                                          AND cds.department = %s
                                        """, (po_key, 'commercial'))
                                    comm_sign = _cur.fetchone()
                            except Exception:
                                comm_sign = None

                            if comm_sign and comm_sign.get('signature_path'):
                                try:
                                    sig_rel_path = comm_sign["signature_path"].lstrip("/")
                                    sig_abs_path = os.path.join(app.root_path, sig_rel_path)

                                    if os.path.exists(sig_abs_path):
                                        tmp_path = None
                                        try:
                                            # Pillow-based processing: autocrop, (optionally) scale to fit cell, then insert
                                            from PIL import Image as PILImage, ImageChops
                                            import tempfile

                                            pil = PILImage.open(sig_abs_path)

                                            # normalize transparency
                                            if pil.mode in ("RGBA", "LA") or (
                                                    pil.mode == "P" and "transparency" in pil.info):
                                                bg = PILImage.new("RGBA", pil.size, (255, 255, 255, 0))
                                                bg.paste(pil, mask=pil.convert("RGBA").split()[-1])
                                                pil = bg
                                            pil = pil.convert("RGBA")

                                            # trim transparent/blank margins
                                            try:
                                                bbox = ImageChops.invert(pil.convert("L")).getbbox()
                                                if bbox:
                                                    pil = pil.crop(bbox)
                                            except Exception:
                                                pass

                                            # target span: single column Y (25), row 38
                                            start_col = 25  # Y
                                            end_col = 25
                                            start_row = 38
                                            end_row = 38

                                            # helpers to estimate pixel sizes
                                            from openpyxl.utils import get_column_letter
                                            def col_width_to_pixels(ws, col_idx):
                                                col_letter = get_column_letter(col_idx)
                                                w = ws.column_dimensions.get(col_letter).width
                                                if w is None:
                                                    w = 8.43
                                                if w < 1:
                                                    return int(w * 12)
                                                return int((w - 0.75) * 7 + 5)

                                            def row_height_to_pixels(ws, row_idx):
                                                h = ws.row_dimensions.get(row_idx).height
                                                if h is None:
                                                    h = 15
                                                return int(h * 96.0 / 72.0)

                                            span_width_px = col_width_to_pixels(ws, start_col)
                                            span_height_px = row_height_to_pixels(ws, start_row)
                                            if span_height_px == 0:
                                                span_height_px = row_height_to_pixels(ws, start_row)

                                            # Make signature width a fraction of column width (e.g., 90% of the single column)
                                            target_w = max(1, int(span_width_px * 0.9))
                                            orig_w, orig_h = pil.size
                                            if orig_w > 0 and target_w < orig_w:
                                                target_h = int((target_w / orig_w) * orig_h)
                                                pil = pil.resize((target_w, target_h), PILImage.LANCZOS)

                                            # save processed image to temp file
                                            tf = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
                                            tmp_path = tf.name
                                            pil.save(tmp_path, format="PNG")
                                            tf.close()

                                            img = Image(tmp_path)
                                            try:
                                                img.width, img.height = pil.size
                                            except Exception:
                                                pass

                                            # insert anchored at Y38 then adjust anchor to the same single cell and center
                                            ws.add_image(img, "Y38")

                                            try:
                                                from openpyxl.utils.units import pixels_to_EMU
                                                anch = getattr(img, "anchor", None)
                                                if anch is not None and hasattr(anch, "_from"):
                                                    fr = anch._from
                                                    to = anch._to
                                                    # set anchor span (zero-based)
                                                    fr.col = start_col - 1
                                                    fr.row = start_row - 1
                                                    to.col = end_col
                                                    to.row = end_row

                                                    left_px = max(
                                                        int((span_width_px - (img.width or pil.size[0])) / 2.0), 0)
                                                    top_px = max(
                                                        int((span_height_px - (img.height or pil.size[1])) / 2.0), 0)
                                                    try:
                                                        fr.colOff = pixels_to_EMU(left_px)
                                                        fr.rowOff = pixels_to_EMU(top_px)
                                                    except Exception:
                                                        # different openpyxl versions may vary; ignore if fails
                                                        pass
                                            except Exception:
                                                pass

                                        except Exception:
                                            # fallback if Pillow not available or processing fails
                                            try:
                                                img = Image(sig_abs_path)
                                                # scale down a bit for single-cell placement
                                                img.width = 100
                                                img.height = 40
                                                ws.add_image(img, "Y38")
                                                # best-effort centering (attempt to set offsets)
                                                try:
                                                    from openpyxl.utils import get_column_letter
                                                    from openpyxl.utils.units import pixels_to_EMU
                                                    def col_width_to_pixels(ws, col_idx):
                                                        col_letter = get_column_letter(col_idx)
                                                        w = ws.column_dimensions.get(col_letter).width
                                                        if w is None:
                                                            w = 8.43
                                                        if w < 1:
                                                            return int(w * 12)
                                                        return int((w - 0.75) * 7 + 5)

                                                    def row_height_to_pixels(ws, row_idx):
                                                        h = ws.row_dimensions.get(row_idx).height
                                                        if h is None:
                                                            h = 15
                                                        return int(h * 96.0 / 72.0)

                                                    span_width_px = col_width_to_pixels(ws, 25)
                                                    span_height_px = row_height_to_pixels(ws, 38)
                                                    anch = getattr(img, "anchor", None)
                                                    if anch is not None and hasattr(anch, "_from"):
                                                        fr = anch._from
                                                        to = anch._to
                                                        fr.col = 25 - 1
                                                        fr.row = 38 - 1
                                                        to.col = 25
                                                        to.row = 38
                                                        left_px = max(int((span_width_px - img.width) / 2.0), 0)
                                                        top_px = max(int((span_height_px - img.height) / 2.0), 0)
                                                        try:
                                                            fr.colOff = pixels_to_EMU(left_px)
                                                            fr.rowOff = pixels_to_EMU(top_px)
                                                        except Exception:
                                                            pass
                                                except Exception:
                                                    pass
                                            except Exception:
                                                pass
                                        finally:
                                            # cleanup temp file if created
                                            try:
                                                if tmp_path:
                                                    import os
                                                    os.unlink(tmp_path)
                                            except Exception:
                                                pass

                                except Exception as e:
                                    print("Commercial signature image error:", e)

                            # Write date + time into Y39 (centered)
                            try:
                                if comm_sign:
                                    signed_at = comm_sign.get('signed_at')
                                    signed_by = comm_sign.get('signed_by', '')
                                else:
                                    signed_at = None
                                    signed_by = ''
                                dt_str = ''
                                if signed_at:
                                    try:
                                        dt_str = signed_at.strftime('%d-%m-%Y %H:%M')
                                    except Exception:
                                        dt_str = str(signed_at)
                                cell = ws.cell(row=39, column=25)  # Y39
                                cell.value = f"{signed_by} | {dt_str}"
                                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                            except Exception:
                                pass

                        buf = BytesIO()
                        wb.save(buf)
                        buf.seek(0)
                        zip_file.writestr(
                            f"{template_key}_{idx+1}_{safe_customer}.xlsx",
                            buf.read()
                        )

                    # ==================================================
                    # 2) CR COMMENTS EXCEL (ONCE PER PO)
                    # ==================================================
                    comments_buffer = build_cr_comments_excel(conn, form, form_id)
                    zip_file.writestr(
                        f"CR_Comments_{safe_customer}.xlsx",
                        comments_buffer.read()
                    )

        conn.close()

        zip_buffer.seek(0)
        response = make_response(zip_buffer.read())
        response.headers['Content-Type'] = 'application/zip'
        response.headers['Content-Disposition'] = 'attachment; filename=CR_Export.zip'
        return response

    except Exception as e:
        try:
            conn.close()
        except Exception:
            pass
        return jsonify({'error': str(e)}), 500

def build_ped_comments_excel(conn, form, form_id):
    import openpyxl
    from io import BytesIO

    template_path = "attached_assets/Comments.xlsx"
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    ws.title = "PED COMMENTS"

    # ---- Header ----
    ws['AB1'] = form['record_no'] or ''
    ws['AB2'] = form['record_date'] or ''
    ws['C3']  = form['customer'] or ''
    ws['F3']  = form['bid'] or ''
    ws['S3']  = form['po'] or ''
    ws['AC3'] = form['cr'] or ''

    # ---- Amendment Details ----
    ws['C26'] = form['amendment_details'] or ''

    # ---- Comments ----
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute("""
            SELECT department, username, comment_text
            FROM ped_comments
            WHERE ped_form_id = %s
            ORDER BY created_at
        """, (form_id,))
        comments = cur.fetchall()

    start_row = 5
    for idx, c in enumerate(comments, start=1):
        r = start_row + idx - 1
        ws[f"A{r}"] = idx
        ws[f"B{r}"] = c['department']
        ws[f"C{r}"] = "PED"
        ws[f"D{r}"] = c['username']
        ws[f"E{r}"] = c['comment_text']

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# --------------------------------------------------
# Helper to get PED signatures for a specific department
# --------------------------------------------------
def get_ped_department_signature(conn, po_key, department):
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute("""
            SELECT
                pds.signed_by,
                pds.signed_at,
                ms.signature_path
            FROM ped_department_signatures pds
            JOIN master_signatures ms
              ON ms.username = pds.signed_by
             AND ms.department = pds.department
             AND ms.is_active = TRUE
            WHERE pds.po_key = %s
              AND pds.department = %s
        """, (po_key, department))
        return cur.fetchone()


@app.route('/api/ped-export-excel', methods=['GET'])
@login_required
def export_ped_to_excel():
    import json
    import openpyxl
    from openpyxl.drawing.image import Image
    from openpyxl.styles import Alignment
    from io import BytesIO
    from flask import make_response, jsonify, request
    import os
    import zipfile

    # --------------------------------------------------
    # Helpers
    # --------------------------------------------------
    def build_merged_cell_map(ws):
        merged_map = {}
        for mr in ws.merged_cells.ranges:
            r0, c0 = mr.min_row, mr.min_col
            for r in range(mr.min_row, mr.max_row + 1):
                for c in range(mr.min_col, mr.max_col + 1):
                    merged_map[(r, c)] = (r0, c0)
        return merged_map

    def write_cell(ws, row, col, value, merged_map):
        if (row, col) in merged_map:
            ar, ac = merged_map[(row, col)]
            ws.cell(row=ar, column=ac).value = value
        else:
            ws.cell(row=row, column=col).value = value

    # Helper to fetch PED signature specifically
    def get_ped_signature(conn, po_key, department):
        try:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("""
                    SELECT
                        pds.signed_by,
                        pds.signed_at,
                        ms.signature_path
                    FROM ped_department_signatures pds
                    JOIN master_signatures ms
                      ON ms.username = pds.signed_by
                     AND ms.department = pds.department
                     AND ms.is_active = TRUE
                    WHERE pds.po_key = %s
                      AND pds.department = %s
                """, (po_key, department))
                return cur.fetchone()
        except Exception as e:
            print(f"Error fetching PED signature for {department}: {e}")
            return None

    # Helper to insert signature image & text
    def insert_signature(ws, sign_data, cell_ref, text_row, text_col):
        if not sign_data or not sign_data.get('signature_path'):
            return

        # 1. Insert Image
        try:
            sig_rel_path = sign_data["signature_path"].lstrip("/")
            sig_abs_path = os.path.join(app.root_path, sig_rel_path)

            if os.path.exists(sig_abs_path):
                img = Image(sig_abs_path)
                # Scale down slightly to fit typical cell
                img.width = 100
                img.height = 40
                ws.add_image(img, cell_ref)
        except Exception as e:
            print(f"Error inserting signature image: {e}")

        # 2. Insert Text (Name | Date)
        try:
            signed_at = sign_data.get('signed_at')
            signed_by = sign_data.get('signed_by', '')
            dt_str = ''
            if signed_at:
                try:
                    dt_str = signed_at.strftime('%d-%m-%Y')
                except Exception:
                    dt_str = str(signed_at)

            cell = ws.cell(row=text_row, column=text_col)
            cell.value = f"{signed_by} | {dt_str}"
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        except Exception as e:
            print(f"Error inserting signature text: {e}")

    templates = {
        'PED_1': 'attached_assets/PED 1_1763437023609.xlsx',
        'PED_2': 'attached_assets/PED 2_1763437023609.xlsx'
    }
    logo_path = 'attached_assets/GTN_LOGO_1762400078631.png'

    for p in templates.values():
        if not os.path.exists(p):
            return jsonify({'error': f'Missing template {p}'}), 404
    if not os.path.exists(logo_path):
        logo_path = None

    # Mapping for PED_2 columns
    ped2_remarks_map = {
        'special-process': 6,
        'welding': 11,
        'assembly': 13,
        'quality': 15,
        'painting': 20,
        'customer-service': 24,
        'commercial': 28
    }

    ped1_cycle_positions = [6, 7, 8, 9, 10, 11, 12, 13, 15, 19]

    ped1_header = {
        'record_no_col': 23,
        'record_date_col': 23,
        'customer_row_col': (3, 3),
        'bid_row_col': (3, 5),
        'po_row_col': (3, 12),
        'cr_row_col': (3, 22),
        'remarks_col': 22,
        'amendment_row_col': (40, 3)
    }

    ped2_header = {
        'record_no_col': 29,
        'record_date_col': 29,
        'customer_row_col': (3, 3),
        'bid_row_col': (3, 5),
        'po_row_col': (3, 13),
        'cr_row_col': (3, 27),
        'remarks_col': 29,
        'amendment_row_col': (40, 3)
    }

    po_ids_param = (request.args.get('po_ids') or '').strip()
    po_ids = []
    if po_ids_param:
        try:
            po_ids = [int(x) for x in po_ids_param.split(',') if x.strip().isdigit()]
        except Exception:
            po_ids = []

    conn = get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            if po_ids:
                placeholders = ",".join(["%s"] * len(po_ids))
                cur.execute(
                    f"SELECT customer, bid, po, cr FROM pos WHERE id IN ({placeholders})",
                    po_ids
                )
                pos_rows = cur.fetchall()

                po_keys = [f"{p['customer']}|{p['bid']}|{p['po']}|{p['cr']}" for p in pos_rows]
                if not po_keys:
                    conn.close()
                    return jsonify({'error': 'No matching POs found for given po_ids'}), 404

                key_placeholders = ",".join(["%s"] * len(po_keys))
                cur.execute(f"""
                    SELECT id, customer, bid, po, cr,
                           record_no, record_date, amendment_details
                    FROM ped_forms
                    WHERE po_key IN ({key_placeholders})
                    ORDER BY id
                """, po_keys)
                forms = cur.fetchall()
            else:
                cur.execute("""
                    SELECT id, customer, bid, po, cr,
                           record_no, record_date, amendment_details
                    FROM ped_forms
                    ORDER BY id
                """)
                forms = cur.fetchall()

        if not forms:
            conn.close()
            return jsonify({'error': 'No PED forms found for selected POs'}), 404

        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                for idx, form in enumerate(forms):
                    form_id = form['id']
                    po_key = f"{form['customer']}|{form['bid']}|{form['po']}|{form['cr']}"
                    safe_customer = ''.join(
                        c for c in (form['customer'] or 'Customer')
                        if c.isalnum() or c in (' ', '_', '-')
                    )[:30]

                    cur.execute("""
                        SELECT item_no, part_number, part_description,
                               rev, qty, ped_cycles, notes, remarks
                        FROM ped_form_rows
                        WHERE ped_form_id = %s
                        ORDER BY id
                    """, (form_id,))
                    rows = cur.fetchall()

                    # ================= PED_1 =================
                    wb1 = openpyxl.load_workbook(templates['PED_1'])
                    ws1 = wb1.active
                    mm1 = build_merged_cell_map(ws1)

                    if logo_path:
                        try:
                            img = Image(logo_path);
                            img.width = 80;
                            img.height = 60
                            ws1.add_image(img, 'A1')
                        except Exception:
                            pass

                    write_cell(ws1, 1, ped1_header['record_no_col'], form['record_no'], mm1)
                    write_cell(ws1, 2, ped1_header['record_date_col'], form['record_date'], mm1)
                    write_cell(ws1, *ped1_header['customer_row_col'], form['customer'], mm1)
                    write_cell(ws1, *ped1_header['bid_row_col'], form['bid'], mm1)
                    write_cell(ws1, *ped1_header['po_row_col'], form['po'], mm1)
                    write_cell(ws1, *ped1_header['cr_row_col'], form['cr'], mm1)
                    write_cell(ws1, *ped1_header['amendment_row_col'], form['amendment_details'], mm1)

                    r0 = 7  # CHANGED from 8 to 7
                    for i, row in enumerate(rows):
                        er = r0 + i
                        write_cell(ws1, er, 1, row['item_no'], mm1)
                        write_cell(ws1, er, 2, row['part_number'], mm1)
                        write_cell(ws1, er, 3, row['part_description'], mm1)
                        write_cell(ws1, er, 4, row['rev'], mm1)
                        write_cell(ws1, er, 5, row['qty'], mm1)

                        cycles = json.loads(row['ped_cycles']) if row['ped_cycles'] else []
                        for ci, col in enumerate(ped1_cycle_positions):
                            write_cell(ws1, er, col, cycles[ci] if ci < len(cycles) else '', mm1)

                        write_cell(ws1, er, ped1_header['remarks_col'], row['remarks'], mm1)

                    # --- PED_1 Signatures (Eng, Mfg, Mat, Purch) ---
                    # Using row 37 or 38 based on template footer space
                    # Adjust 'F37', 'S37' etc based on actual template layout

                    # Engineering
                    sign = get_ped_signature(conn, po_key, 'engineering')
                    insert_signature(ws1, sign, "F37", 38, 6)  # Image at F37, Text at F38 (col 6)

                    # Manufacturing
                    sign = get_ped_signature(conn, po_key, 'manufacturing')
                    insert_signature(ws1, sign, "M37", 38, 13)  # Adjust col as needed (M is 13)

                    # Materials
                    sign = get_ped_signature(conn, po_key, 'materials')
                    insert_signature(ws1, sign, "O37", 38, 15)  # O is 15

                    # Purchase
                    sign = get_ped_signature(conn, po_key, 'purchase')
                    insert_signature(ws1, sign, "S37", 38, 19)  # S is 19

                    b1 = BytesIO();
                    wb1.save(b1);
                    b1.seek(0)
                    zip_file.writestr(f"PED_1_{idx + 1}_{safe_customer}.xlsx", b1.read())

                    # ================= PED_2 =================
                    wb2 = openpyxl.load_workbook(templates['PED_2'])
                    ws2 = wb2.active
                    mm2 = build_merged_cell_map(ws2)

                    if logo_path:
                        try:
                            img = Image(logo_path);
                            img.width = 80;
                            img.height = 60
                            ws2.add_image(img, 'A1')
                        except Exception:
                            pass

                    write_cell(ws2, 1, ped2_header['record_no_col'], form['record_no'], mm2)
                    write_cell(ws2, 2, ped2_header['record_date_col'], form['record_date'], mm2)
                    write_cell(ws2, *ped2_header['customer_row_col'], form['customer'], mm2)
                    write_cell(ws2, *ped2_header['bid_row_col'], form['bid'], mm2)
                    write_cell(ws2, *ped2_header['po_row_col'], form['po'], mm2)
                    write_cell(ws2, *ped2_header['cr_row_col'], form['cr'], mm2)
                    write_cell(ws2, *ped2_header['amendment_row_col'], form['amendment_details'], mm2)

                    # Map internal keys to database keys
                    # DB keys: special-process, welding, assembly, quality, painting, customer-service, commercial
                    dept_db_keys = [
                        'special-process', 'welding', 'assembly', 'quality',
                        'painting', 'customer-service', 'commercial'
                    ]

                    # Columns in PED 2 template for signatures (Approximate)
                    # Special Process: F37 (col 6)
                    # Welding: K37 (col 11)
                    # Assembly: M37 (col 13)
                    # Quality: O37 (col 15)
                    # Painting: T37 (col 20)
                    # Cust Service: X37 (col 24)
                    # Commercial: AB37 (col 28)

                    sig_col_map = {
                        'special-process': ('F37', 6),
                        'welding': ('K37', 11),
                        'assembly': ('M37', 13),
                        'quality': ('O37', 15),
                        'painting': ('T37', 20),
                        'customer-service': ('X37', 24),
                        'commercial': ('AB37', 28)
                    }

                    # Write Rows
                    order = ['special-process', 'welding', 'assembly', 'quality',
                             'painting', 'customer-service', 'commercial']

                    r0 = 7  # CHANGED from 8 to 7
                    for i, row in enumerate(rows):
                        er = r0 + i
                        write_cell(ws2, er, 1, row['item_no'], mm2)
                        write_cell(ws2, er, 2, row['part_number'], mm2)
                        write_cell(ws2, er, 3, row['part_description'], mm2)
                        write_cell(ws2, er, 4, row['rev'], mm2)
                        write_cell(ws2, er, 5, row['qty'], mm2)

                        notes = json.loads(row['notes']) if row['notes'] else []
                        for ni, key in enumerate(order):
                            write_cell(ws2, er, ped2_remarks_map[key],
                                       notes[ni] if ni < len(notes) else '', mm2)

                        write_cell(ws2, er, ped2_header['remarks_col'], row['remarks'], mm2)

                    # --- PED_2 Signatures ---
                    for d_key in dept_db_keys:
                        if d_key in sig_col_map:
                            cell_ref, col_idx = sig_col_map[d_key]
                            sign = get_ped_signature(conn, po_key, d_key)
                            insert_signature(ws2, sign, cell_ref, 38, col_idx)

                    b2 = BytesIO();
                    wb2.save(b2);
                    b2.seek(0)
                    zip_file.writestr(f"PED_2_{idx + 1}_{safe_customer}.xlsx", b2.read())

                    # ================= PED COMMENTS (ONCE) =================
                    cbuf = build_ped_comments_excel(conn, form, form_id)
                    zip_file.writestr(f"PED_Comments_{safe_customer}.xlsx", cbuf.read())

        conn.close()
        zip_buffer.seek(0)
        resp = make_response(zip_buffer.read())
        resp.headers['Content-Type'] = 'application/zip'
        resp.headers['Content-Disposition'] = 'attachment; filename=PED_Export.zip'
        return resp

    except Exception as e:
        try:
            conn.close()
        except Exception:
            pass
        return jsonify({'error': str(e)}), 500


def build_lead_comments_excel(conn, form, form_id):
    import openpyxl
    from io import BytesIO

    template_path = "attached_assets/Comments.xlsx"
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    ws.title = "LEAD COMMENTS"

    # ---- Header ----
    ws['AB1'] = form['record_no'] or ''
    ws['AB2'] = form['record_date'] or ''
    ws['C3']  = form['customer'] or ''
    ws['F3']  = form['bid'] or ''
    ws['S3']  = form['po'] or ''
    ws['AC3'] = form['cr'] or ''
    ws['C26'] = form['general_remarks'] or ''

    # ---- Comments ----
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute("""
            SELECT department, username, comment_text
            FROM lead_comments
            WHERE lead_form_id = %s
            ORDER BY created_at
        """, (form_id,))
        comments = cur.fetchall()

    start_row = 5
    for idx, c in enumerate(comments, start=1):
        r = start_row + idx - 1
        ws[f"A{r}"] = idx
        ws[f"B{r}"] = c['department']
        ws[f"C{r}"] = "LEAD"
        ws[f"D{r}"] = c['username']
        ws[f"E{r}"] = c['comment_text']

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@app.route('/api/lead-export-excel', methods=['GET'])
@login_required
def export_lead_to_excel():
    import openpyxl
    from openpyxl.drawing.image import Image
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.units import pixels_to_EMU
    from io import BytesIO
    import zipfile
    from flask import make_response, jsonify, request
    import os
    from PIL import Image as PILImage, ImageChops
    import psycopg2.extras

    # --- Helpers for Signature Insertion (FIXED: use BytesIO, no temp files) ---
    def insert_signature(ws, img_path, start_col, end_col, row):
        if not img_path or not os.path.exists(img_path):
            return

        try:
            pil = PILImage.open(img_path)

            # normalize transparency
            if pil.mode in ("RGBA", "LA") or (pil.mode == "P" and "transparency" in pil.info):
                bg = PILImage.new("RGBA", pil.size, (255, 255, 255, 0))
                bg.paste(pil, mask=pil.convert("RGBA").split()[-1])
                pil = bg
            pil = pil.convert("RGBA")

            # trim whitespace
            try:
                bbox = ImageChops.invert(pil.convert("L")).getbbox()
                if bbox:
                    pil = pil.crop(bbox)
            except Exception:
                pass

            def col_width_to_pixels(ws, col_idx):
                col_letter = get_column_letter(col_idx)
                w = ws.column_dimensions.get(col_letter).width
                if w is None:
                    w = 8.43
                if w < 1:
                    return int(w * 12)
                return int((w - 0.75) * 7 + 5)

            def row_height_to_pixels(ws, row_idx):
                h = ws.row_dimensions.get(row_idx).height
                if h is None:
                    h = 15
                return int(h * 96.0 / 72.0)

            span_width_px = sum(col_width_to_pixels(ws, c) for c in range(start_col, end_col + 1))
            span_height_px = row_height_to_pixels(ws, row)
            if span_height_px == 0:
                span_height_px = 20

            target_w = max(1, int(span_width_px * 0.8))
            target_h = max(1, int(span_height_px * 0.9))

            orig_w, orig_h = pil.size
            if orig_w <= 0 or orig_h <= 0:
                return

            ratio = min(target_w / orig_w, target_h / orig_h)
            new_w = max(1, int(orig_w * ratio))
            new_h = max(1, int(orig_h * ratio))

            pil = pil.resize((new_w, new_h), PILImage.LANCZOS)

            img_bytes = BytesIO()
            pil.save(img_bytes, format="PNG")
            img_bytes.seek(0)

            img = Image(img_bytes)
            img.width = new_w
            img.height = new_h

            anchor_cell = get_column_letter(start_col) + str(row)
            ws.add_image(img, anchor_cell)

            center_x_px = (span_width_px - new_w) // 2
            center_y_px = (span_height_px - new_h) // 2

            try:
                anch = getattr(img, "anchor", None)
                if anch is not None and hasattr(anch, "_from"):
                    anch._from.col = start_col - 1
                    anch._from.row = row - 1
                    anch._from.colOff = pixels_to_EMU(max(0, center_x_px))
                    anch._from.rowOff = pixels_to_EMU(max(0, center_y_px))
            except Exception:
                pass

        except Exception as e:
            print(f"Error inserting signature {img_path}: {e}")

    def write_signature_text(ws, start_col, end_col, row, signed_by, signed_at):
        date_str = ''
        if signed_at:
            try:
                date_str = signed_at.strftime('%d-%m-%Y %H:%M')
            except Exception:
                date_str = str(signed_at)

        text = f"{signed_by} | {date_str}" if signed_by else ""

        try:
            ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
        except Exception:
            pass

        cell = ws.cell(row=row, column=start_col)
        cell.value = text
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def build_merged_cell_map(ws):
        merged_map = {}
        for mr in ws.merged_cells.ranges:
            r0, c0 = mr.min_row, mr.min_col
            for r in range(mr.min_row, mr.max_row + 1):
                for c in range(mr.min_col, mr.max_col + 1):
                    merged_map[(r, c)] = (r0, c0)
        return merged_map

    def write_cell(ws, row, col, value, merged_map):
        if (row, col) in merged_map:
            ar, ac = merged_map[(row, col)]
            ws.cell(row=ar, column=ac).value = value
        else:
            ws.cell(row=row, column=col).value = value

    template_path = 'attached_assets/Lead Form_1763437805952.xlsx'
    logo_path = 'attached_assets/GTN_LOGO_1762400078631.png'

    if not os.path.exists(template_path):
        return jsonify({'error': 'Lead Form template file not found'}), 404

    use_logo = os.path.exists(logo_path)

    # --- Signature Map for LEAD Form (UPDATED to match your HTML final dept keys) ---
    sig_map = {
        'css': {'start_col': 8, 'end_col': 9},                     # H-I
        'materials': {'start_col': 10, 'end_col': 11},             # J-K
        'technical-operations': {'start_col': 12, 'end_col': 14},  # L-N
        'quality': {'start_col': 15, 'end_col': 16},               # O-P
        'operations': {'start_col': 17, 'end_col': 19},            # Q-S
    }
    # Image row = 38, Text row = 39

    po_ids_param = (request.args.get('po_ids') or '').strip()
    po_ids = []
    if po_ids_param:
        try:
            po_ids = [int(x) for x in po_ids_param.split(',') if x.strip().isdigit()]
        except Exception:
            po_ids = []

    conn = get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            if po_ids:
                placeholders = ",".join(["%s"] * len(po_ids))
                cur.execute(
                    f"SELECT customer, bid, po, cr FROM pos WHERE id IN ({placeholders})",
                    po_ids
                )
                pos_rows = cur.fetchall()

                po_keys = [f"{p['customer']}|{p['bid']}|{p['po']}|{p['cr']}" for p in pos_rows]
                if not po_keys:
                    conn.close()
                    return jsonify({'error': 'No matching POs found for given po_ids'}), 404

                key_placeholders = ",".join(["%s"] * len(po_keys))
                cur.execute(f"""
                    SELECT id, customer, bid, po, cr,
                           record_no, record_date, prepared_by, general_remarks
                    FROM lead_forms
                    WHERE po_key IN ({key_placeholders})
                    ORDER BY id
                """, po_keys)
                forms = cur.fetchall()
            else:
                cur.execute("""
                    SELECT id, customer, bid, po, cr,
                           record_no, record_date, prepared_by, general_remarks
                    FROM lead_forms
                    ORDER BY id
                """)
                forms = cur.fetchall()

        if not forms:
            conn.close()
            return jsonify({'error': 'No Lead forms found for selected POs'}), 404

        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                for idx, form in enumerate(forms):
                    form_id = form['id']
                    po_key = f"{form['customer']}|{form['bid']}|{form['po']}|{form['cr']}"
                    safe_customer = ''.join(
                        c for c in (form['customer'] or 'Customer')
                        if c.isalnum() or c in (' ', '_', '-')
                    )[:30]

                    wb = openpyxl.load_workbook(template_path)
                    ws = wb.active
                    merged_map = build_merged_cell_map(ws)

                    if use_logo:
                        try:
                            logo_img = Image(logo_path)
                            logo_img.width = 80
                            logo_img.height = 60
                            ws.add_image(logo_img, 'A1')
                        except Exception:
                            pass

                    # ---- Header ----
                    write_cell(ws, 1, 16, form['record_no'] or 'SAL/R02/Y', merged_map)  # P1
                    write_cell(ws, 2, 16, form['record_date'] or '', merged_map)        # P2
                    write_cell(ws, 3, 3, form['customer'] or '', merged_map)            # C3
                    write_cell(ws, 3, 6, form['bid'] or '', merged_map)                 # F3
                    write_cell(ws, 3, 11, form['po'] or '', merged_map)                 # K3
                    write_cell(ws, 3, 16, form['cr'] or '', merged_map)                 # P3

                    # ---- Amendment / General Remarks ----
                    write_cell(ws, 37, 1, form['general_remarks'] or '', merged_map)

                    # ---- Prepared By (A39) ----
                    write_cell(ws, 39, 1, (form.get('prepared_by') or ''), merged_map)

                    # ---- Rows ----
                    cur.execute("""
                        SELECT item_no, part_number, part_description,
                               rev, qty, customer_required_date,
                               standard_lead_time, gtn_agreed_date, remarks
                        FROM lead_form_rows
                        WHERE lead_form_id = %s
                        ORDER BY id
                    """, (form_id,))
                    rows = cur.fetchall()

                    start_row = 6
                    for i, r in enumerate(rows):
                        er = start_row + i
                        write_cell(ws, er, 1, r['item_no'], merged_map)
                        write_cell(ws, er, 2, r['part_number'], merged_map)
                        write_cell(ws, er, 4, r['part_description'], merged_map)
                        write_cell(ws, er, 10, r['rev'], merged_map)
                        write_cell(ws, er, 11, r['qty'], merged_map)
                        write_cell(ws, er, 12, r['customer_required_date'], merged_map)
                        write_cell(ws, er, 14, r['standard_lead_time'], merged_map)
                        write_cell(ws, er, 15, r['gtn_agreed_date'], merged_map)
                        write_cell(ws, er, 16, r['remarks'], merged_map)

                    # ---- Signatures (NOW FROM LEAD table, not CR) ----
                    for dept, mapping in sig_map.items():
                        cur.execute("""
                            SELECT lds.signed_by, lds.signed_at, ms.signature_path
                            FROM lead_department_signatures lds
                            JOIN master_signatures ms
                              ON ms.username = lds.signed_by
                             AND ms.department = lds.department
                             AND ms.is_active = TRUE
                            WHERE lds.po_key = %s AND lds.department = %s
                        """, (po_key, dept))
                        sig_data = cur.fetchone()

                        if sig_data and sig_data.get('signature_path'):
                            full_path = os.path.join(app.root_path, sig_data['signature_path'].lstrip('/'))
                            insert_signature(ws, full_path, mapping['start_col'], mapping['end_col'], 38)
                            write_signature_text(
                                ws,
                                mapping['start_col'], mapping['end_col'], 39,
                                sig_data.get('signed_by'), sig_data.get('signed_at')
                            )

                    buf = BytesIO()
                    wb.save(buf)
                    buf.seek(0)
                    zip_file.writestr(f"Lead_Form_{idx + 1}_{safe_customer}.xlsx", buf.read())

                    # LEAD COMMENTS
                    comments_buf = build_lead_comments_excel(conn, form, form_id)
                    zip_file.writestr(f"Lead_Comments_{safe_customer}.xlsx", comments_buf.read())

        conn.close()
        zip_buffer.seek(0)
        response = make_response(zip_buffer.read())
        response.headers['Content-Type'] = 'application/zip'
        response.headers['Content-Disposition'] = 'attachment; filename=Lead_Export.zip'
        return response

    except Exception as e:
        try:
            conn.close()
        except Exception:
            pass
        return jsonify({'error': str(e)}), 500

# ---------- CR COMMENTS ----------

@app.route('/api/cr-comments/<int:form_id>', methods=['POST'])
@login_required
def post_cr_comment(form_id):
    data = request.get_json()
    comment_text = data.get('comment', '').strip()

    if not comment_text:
        return jsonify({'error': 'Comment text is required'}), 400

    username = session.get('username', 'unknown')
    department = session.get('user_department', 'unknown')

    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute('''
                INSERT INTO cr_comments (cr_form_id, username, department, comment_text)
                VALUES (%s, %s, %s, %s)
                RETURNING id, created_at
            ''', (form_id, username, department, comment_text))
            row = cur.fetchone()
        conn.commit()
        conn.close()

        return jsonify({
            'success': True,
            'comment': {
                'id': row[0],
                'username': username,
                'department': department,
                'text': comment_text,
                'createdAt': row[1]
            }
        })
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': str(e)}), 500

@app.route('/api/cr-comments/<int:form_id>', methods=['GET'])
@login_required
def get_cr_comments(form_id):
    conn = get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute('''
                SELECT id, username, department, comment_text, created_at
                FROM cr_comments
                WHERE cr_form_id = %s
                ORDER BY created_at ASC
            ''', (form_id,))
            comments = cur.fetchall()
        conn.close()

        return jsonify({
            'comments': [{
                'id': c['id'],
                'username': c['username'],
                'department': c['department'],
                'text': c['comment_text'],
                'createdAt': c['created_at']
            } for c in comments]
        })
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 500

@app.route('/api/ped-comments/<int:form_id>', methods=['GET'])
@login_required
def get_ped_comments(form_id):
    conn = get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute('''
                SELECT id, username, department, comment_text, created_at
                FROM ped_comments
                WHERE ped_form_id = %s
                ORDER BY created_at ASC
            ''', (form_id,))
            comments = cur.fetchall()
        conn.close()

        return jsonify({
            'comments': [{
                'id': c['id'],
                'username': c['username'],
                'department': c['department'],
                'text': c['comment_text'],
                'createdAt': c['created_at']
            } for c in comments]
        })
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 500

@app.route('/api/ped-comments/<int:form_id>', methods=['POST'])
@login_required
def post_ped_comment(form_id):
    data = request.get_json()
    comment_text = data.get('comment', '').strip()

    if not comment_text:
        return jsonify({'error': 'Comment text is required'}), 400

    username = session.get('username', 'unknown')
    department = session.get('user_department', 'unknown')

    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute('''
                INSERT INTO ped_comments (ped_form_id, username, department, comment_text)
                VALUES (%s, %s, %s, %s)
                RETURNING id, created_at
            ''', (form_id, username, department, comment_text))
            row = cur.fetchone()
        conn.commit()
        conn.close()

        return jsonify({
            'success': True,
            'comment': {
                'id': row[0],
                'username': username,
                'department': department,
                'text': comment_text,
                'createdAt': row[1]
            }
        })
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': str(e)}), 500

@app.route('/api/lead-comments/<int:form_id>', methods=['GET'])
@login_required
def get_lead_comments(form_id):
    conn = get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute('''
                SELECT id, username, department, comment_text, created_at
                FROM lead_comments
                WHERE lead_form_id = %s
                ORDER BY created_at ASC
            ''', (form_id,))
            comments = cur.fetchall()
        conn.close()

        return jsonify({
            'comments': [{
                'id': c['id'],
                'username': c['username'],
                'department': c['department'],
                'text': c['comment_text'],
                'createdAt': c['created_at']
            } for c in comments]
        })
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 500

@app.route('/api/lead-comments/<int:form_id>', methods=['POST'])
@login_required
def post_lead_comment(form_id):
    data = request.get_json()
    comment_text = data.get('comment', '').strip()

    if not comment_text:
        return jsonify({'error': 'Comment text is required'}), 400

    username = session.get('username', 'unknown')
    department = session.get('user_department', 'unknown')

    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute('''
                INSERT INTO lead_comments (lead_form_id, username, department, comment_text)
                VALUES (%s, %s, %s, %s)
                RETURNING id, created_at
            ''', (form_id, username, department, comment_text))
            row = cur.fetchone()
        conn.commit()
        conn.close()

        return jsonify({
            'success': True,
            'comment': {
                'id': row[0],
                'username': username,
                'department': department,
                'text': comment_text,
                'createdAt': row[1]
            }
        })
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': str(e)}), 500

# ---------- Notifications ----------

@app.route('/api/notifications', methods=['GET'])
@login_required
def get_notifications():
    user_id = session.get('user_id')
    limit = request.args.get('limit', 50, type=int)

    conn = get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute('''
                SELECT n.*, u.username AS actor_username
                  FROM notifications n
             LEFT JOIN users u ON n.actor_user_id = u.id
                 WHERE n.recipient_user_id = %s
                 ORDER BY n.created_at DESC
                 LIMIT %s
            ''', (user_id, limit))
            notifications = cur.fetchall()

            cur.execute('''
                SELECT COUNT(*) AS count
                  FROM notifications 
                 WHERE recipient_user_id = %s AND is_read = FALSE
            ''', (user_id,))
            unread_count = cur.fetchone()['count']

        conn.close()

        return jsonify({
            'notifications': [{
                'id': n['id'],
                'eventType': n['event_type'],
                'message': n['message'],
                'isRead': bool(n['is_read']),
                'createdAt': n['created_at'],
                'actorUsername': n['actor_username']
            } for n in notifications],
            'unreadCount': unread_count
        })
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 500

@app.route('/api/notifications/unread_count', methods=['GET'])
@login_required
def get_unread_count():
    user_id = session.get('user_id')

    conn = get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute('''
                SELECT COUNT(*) AS count
                  FROM notifications 
                 WHERE recipient_user_id = %s AND is_read = FALSE
            ''', (user_id,))
            unread_count = cur.fetchone()['count']

        conn.close()
        return jsonify({'unreadCount': unread_count})
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 500

@app.route('/api/notifications/<int:notification_id>/read', methods=['PUT'])
@login_required
def mark_notification_read(notification_id):
    user_id = session.get('user_id')

    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute('''
                UPDATE notifications 
                   SET is_read = TRUE 
                 WHERE id = %s AND recipient_user_id = %s
            ''', (notification_id, user_id))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': str(e)}), 500

@app.route('/api/notifications/mark_all_read', methods=['PUT'])
@login_required
def mark_all_read():
    user_id = session.get('user_id')

    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute('''
                UPDATE notifications 
                   SET is_read = TRUE 
                 WHERE recipient_user_id = %s AND is_read = FALSE
            ''', (user_id,))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': str(e)}), 500

# ---------- Email config ----------

@app.route('/api/email-config', methods=['GET'])
@login_required
def get_email_config():
    if not session.get('user_is_admin'):
        return jsonify({'error': 'Admin access required'}), 403

    conn = get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                'SELECT * FROM email_config ORDER BY id DESC LIMIT 1'
            )
            config = cur.fetchone()
        conn.close()

        if config:
            return jsonify({
                'smtp_host': config['smtp_host'],
                'smtp_port': config['smtp_port'],
                'smtp_username': config['smtp_username'],
                'from_email': config['from_email'],
                'use_tls': bool(config['use_tls']),
                'email_enabled': bool(config['email_enabled']),
                'has_password': bool(config['smtp_password'])
            })
        else:
            return jsonify({
                'smtp_host': '',
                'smtp_port': 587,
                'smtp_username': '',
                'from_email': '',
                'use_tls': True,
                'email_enabled': False,
                'has_password': False
            })
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 500

@app.route('/api/email-config', methods=['POST'])
@login_required
def save_email_config():
    if not session.get('user_is_admin'):
        return jsonify({'error': 'Admin access required'}), 403

    data = request.json
    username = session.get('username')

    conn = get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                'SELECT id, smtp_password FROM email_config ORDER BY id DESC LIMIT 1'
            )
            existing = cur.fetchone()

            smtp_password = data.get('smtp_password', '')
            if not smtp_password and existing:
                smtp_password = existing['smtp_password']
            elif smtp_password:
                smtp_password = encrypt_password(smtp_password)

            if not smtp_password:
                conn.close()
                return jsonify({'error': 'SMTP password is required'}), 400

            if existing:
                cur.execute('''
                    UPDATE email_config SET 
                        smtp_host = %s,
                        smtp_port = %s,
                        smtp_username = %s,
                        smtp_password = %s,
                        from_email = %s,
                        use_tls = %s,
                        email_enabled = %s,
                        updated_by = %s,
                        updated_at = CURRENT_TIMESTAMP
                     WHERE id = %s
                ''', (data.get('smtp_host'), data.get('smtp_port', 587),
                      data.get('smtp_username'), smtp_password,
                      data.get('from_email'),
                      True if data.get('use_tls', True) else False,
                      True if data.get('email_enabled', False) else False,
                      username, existing['id']))
            else:
                cur.execute('''
                    INSERT INTO email_config 
                      (smtp_host, smtp_port, smtp_username, smtp_password,
                       from_email, use_tls, email_enabled, updated_by)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                ''', (data.get('smtp_host'), data.get('smtp_port', 587),
                      data.get('smtp_username'), smtp_password,
                      data.get('from_email'),
                      True if data.get('use_tls', True) else False,
                      True if data.get('email_enabled', False) else False,
                      username))
        conn.commit()
        conn.close()

        return jsonify({'success': True, 'message': 'Email configuration saved successfully'})
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': str(e)}), 500

@app.route('/api/email-config/test', methods=['POST'])
@login_required
def test_email_config():
    if not session.get('user_is_admin'):
        return jsonify({'error': 'Admin access required'}), 403

    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart

    data = request.json
    test_email = data.get('test_email')

    if not test_email:
        return jsonify({'error': 'Test email address is required'}), 400

    conn = get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                'SELECT * FROM email_config ORDER BY id DESC LIMIT 1'
            )
            config = cur.fetchone()
        conn.close()

        if not config:
            return jsonify({
                'error':
                'No email configuration found. Please save configuration first.'
            }), 400

        msg = MIMEMultipart()
        msg['From'] = config['from_email']
        msg['To'] = test_email
        msg['Subject'] = 'GTN Engineering - Email Configuration Test'

        body = '''
        <html>
        <body>
            <h2>Email Configuration Test</h2>
            <p>This is a test email from GTN Engineering Contract Review Dashboard.</p>
            <p>If you received this email, your SMTP configuration is working correctly!</p>
            <p><strong>Configuration Details:</strong></p>
            <ul>
                <li>SMTP Host: {}</li>
                <li>SMTP Port: {}</li>
                <li>From Email: {}</li>
            </ul>
        </body>
        </html>
        '''.format(config['smtp_host'], config['smtp_port'],
                   config['from_email'])

        msg.attach(MIMEText(body, 'html'))

        decrypted_password = decrypt_password(config['smtp_password'])

        server = smtplib.SMTP(config['smtp_host'], config['smtp_port'])
        if config['use_tls']:
            server.starttls()
        server.login(config['smtp_username'], decrypted_password)
        server.send_message(msg)
        server.quit()

        return jsonify({
            'success': True,
            'message': f'Test email sent successfully to {test_email}'
        })
    except Exception as e:
        return jsonify(
            {'error': f'Failed to send test email: {str(e)}'}), 500


@app.route("/attached_assets/<path:filename>")
def attached_assets(filename):
    return send_from_directory("attached_assets", filename)

if __name__ == "__main__":
    init_db()
    app.run(host="0.0.0.0", port=1000, debug=True)